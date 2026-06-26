from flask import Flask, render_template, request, redirect, session, jsonify, flash
from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from functools import wraps
import gspread
from google.oauth2.service_account import Credentials
import time
from dotenv import load_dotenv
import os
import re
from db import db
from models import User, Customer, Cylinder, CylinderMaintenance, Scan, CustomerMap, BulkTank, Product, DuraGasHistory, SystemSetting
from apscheduler.schedulers.background import BackgroundScheduler
from werkzeug.security import generate_password_hash, check_password_hash
from concurrent.futures import ThreadPoolExecutor

# Initialize global background executor for Google Sheets
sheets_executor = ThreadPoolExecutor(max_workers=3)

def async_sheets_write(fn, *args, **kwargs):
    """Submits a Google Sheets API call to the background executor."""
    try:
        sheets_executor.submit(fn, *args, **kwargs)
    except Exception as e:
        print(f"[sheets_executor] Error submitting task: {e}")

load_dotenv()

app = Flask(__name__)
app.secret_key = 'cyl-tracker-secret-2026'

@app.template_filter('strip')
def strip_filter(s):
    if s is None:
        return ''
    if isinstance(s, str):
        return s.strip()
    return str(s).strip()


# SQLAlchemy Database Configuration
db_url = os.environ.get('DATABASE_URL')
if not db_url:
    print("[warning] DATABASE_URL environment variable is not set. Falling back to Google Sheets for operations.")
    db_url = 'sqlite:///:memory:'
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# Auto-create any missing tables on startup (safe — checkfirst=True skips existing tables)
with app.app_context():
    db.create_all()
    if os.environ.get('DATABASE_URL'):
        try:
            from sqlalchemy import text
            db.session.execute(text("ALTER TABLE scans ADD COLUMN IF NOT EXISTS gas_type VARCHAR(50);"))
            db.session.commit()
            print("[startup] Checked and added scans.gas_type column if missing.")
        except Exception as e:
            print("[startup] scans.gas_type alter check failed:", e)
    print("[startup] DB tables verified/created.")

# Session Security Configuration
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = not os.environ.get('FLASK_DEBUG') == '1'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)


@app.after_request
def add_header(response):
    # Prevent browser caching for admin and API endpoints
    if request.path.startswith('/admin') or request.path.startswith('/api'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response



SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

creds = Credentials.from_service_account_file(
    "credentials.json",
    scopes=SCOPES
)
client = gspread.authorize(creds)

SPREADSHEET_NAME    = "Cylinder Tracking"
SCAN_SHEET_NAME     = "Sheet1"
MAP_SHEET_NAME      = "Customer Map"
USERS_SHEET_NAME    = "Users"
CUSTOMER_SHEET_NAME = "Customers"
CYLINDER_SHEET_NAME = "Cylinders"
CYLINDER_MAINT_NAME = "Cylinder Maintenance"
BULK_TANKS_NAME     = "Bulk Tanks"
PRODUCTS_SHEET_NAME = "Products"

# Cache worksheet objects at startup to avoid roundtrip sheet lookup calls
try:
    doc = client.open(SPREADSHEET_NAME)
    scan_ws = doc.worksheet(SCAN_SHEET_NAME)
    map_ws = doc.worksheet(MAP_SHEET_NAME)
    users_ws = doc.worksheet(USERS_SHEET_NAME)
    customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
    sheet = scan_ws
    try:
        cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
    except Exception:
        cyl_ws = None
    try:
        cyl_maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
    except Exception:
        cyl_maint_ws = None
    try:
        bulk_tanks_ws = doc.worksheet(BULK_TANKS_NAME)
    except Exception:
        bulk_tanks_ws = None
    try:
        products_ws = doc.worksheet(PRODUCTS_SHEET_NAME)
    except Exception:
        products_ws = None
except Exception as e:
    print("Error caching Google Sheets worksheets:", e)
    doc = None
    scan_ws = None
    map_ws = None
    users_ws = None
    customer_ws = None
    sheet = None
    cyl_ws = None
    cyl_maint_ws = None
    bulk_tanks_ws = None
    products_ws = None

# ── Default products list (fallback if Products sheet not found) ──────────────
DEFAULT_PRODUCTS_CONFIG = [
    {'id': 'arg_pura',     'name': 'ARG Pura',      'gas_type': 'ARG', 'cylinder_type': 'Standard', 'gas_per_cyl': 7.0,    'unit': 'Cum', 'is_virtual': False},
    {'id': 'acm_90_10',   'name': 'ACM (90.10)_',  'gas_type': 'ACM', 'cylinder_type': 'Standard', 'gas_per_cyl': 6.3512, 'unit': 'Cum', 'is_virtual': False},
    {'id': 'co2_90_10',   'name': 'Co2 (90.10)_',  'gas_type': 'ACM', 'cylinder_type': 'Standard', 'gas_per_cyl': 1.35,   'unit': 'KG',  'is_virtual': True},
    {'id': 'co2_pure',    'name': 'Co2',            'gas_type': 'CO2', 'cylinder_type': 'Standard', 'gas_per_cyl': 30.0,   'unit': 'KG',  'is_virtual': False},
    {'id': 'n2_cyl',      'name': 'N2 Cyl',         'gas_type': 'N2',  'cylinder_type': 'Standard', 'gas_per_cyl': 7.0,    'unit': 'Cum', 'is_virtual': False},
    {'id': 'oxygen_pure', 'name': 'OXYGEN',          'gas_type': 'OXY', 'cylinder_type': 'Standard', 'gas_per_cyl': 7.0,    'unit': 'Cum', 'is_virtual': False},
    {'id': 'ahm_92_08',   'name': 'AHM(92.08)',     'gas_type': 'AHM', 'cylinder_type': 'Standard', 'gas_per_cyl': 6.92,   'unit': 'Cum', 'is_virtual': False},
    {'id': 'ahm_98_02',   'name': 'AHM (98.02)',    'gas_type': 'AHM', 'cylinder_type': 'Standard', 'gas_per_cyl': 6.98,   'unit': 'Cum', 'is_virtual': False},
    {'id': 'arg_dura',    'name': 'ARG Dura',       'gas_type': 'ARG', 'cylinder_type': 'Dura',     'gas_per_cyl': 0.0,    'unit': 'Cum', 'is_virtual': False},
    {'id': 'n2_dura',     'name': 'N2Dura',         'gas_type': 'N2',  'cylinder_type': 'Dura',     'gas_per_cyl': 0.88,   'unit': 'Cum', 'is_virtual': False},
    {'id': 'oxygen_dura', 'name': 'Oxygen Dura',    'gas_type': 'OXY', 'cylinder_type': 'Dura',     'gas_per_cyl': 0.0,    'unit': 'Cum', 'is_virtual': False},
]

def get_products_config():
    """Reads product rows from the Products database table, falling back to Google Sheets."""
    try:
        if os.environ.get('DATABASE_URL'):
            products = Product.query.all()
            if products:
                return [{
                    'id':            p.product_id,
                    'name':          p.name,
                    'gas_type':      p.gas_type,
                    'cylinder_type': p.cylinder_type,
                    'gas_per_cyl':   p.gas_per_cyl,
                    'unit':          p.unit,
                    'is_virtual':    p.is_virtual,
                } for p in products]
    except Exception as e:
        print("[db] get_products_config read error, falling back to Sheets:", e)

    global products_ws
    try:
        if products_ws is None and doc:
            try:
                products_ws = doc.worksheet(PRODUCTS_SHEET_NAME)
            except Exception:
                pass
        if products_ws is None:
            return DEFAULT_PRODUCTS_CONFIG

        rows = products_ws.get_all_values()
        if len(rows) < 2:
            return DEFAULT_PRODUCTS_CONFIG

        # Expected header order:
        # Product ID | Display Name | Gas Type | Cylinder Type | Gas Per Cylinder | Unit | Is Virtual?
        config = []
        for r in rows[1:]:  # skip header
            if len(r) < 6 or not r[0].strip():
                continue
            try:
                gas_per = float(r[4].strip()) if r[4].strip() else 0.0
            except ValueError:
                gas_per = 0.0
            is_virtual = str(r[6]).strip().upper() == 'TRUE' if len(r) > 6 else False
            config.append({
                'id':            r[0].strip(),
                'name':          r[1].strip(),
                'gas_type':      r[2].strip().upper(),
                'cylinder_type': r[3].strip().capitalize(),
                'gas_per_cyl':   gas_per,
                'unit':          r[5].strip(),
                'is_virtual':    is_virtual,
            })
        return config if config else DEFAULT_PRODUCTS_CONFIG
    except Exception as e:
        print("get_products_config error:", e)
        return DEFAULT_PRODUCTS_CONFIG

# Helper functions to fetch customer details from Google Sheets
def get_customer_names():
    if os.environ.get('DATABASE_URL'):
        try:
            customers = Customer.query.all()
            if customers:
                names = [c.name.strip() for c in customers if c.name.strip()]
                return sorted(list(set(names)))
        except Exception as e:
            print("[db] Error getting customer names from DB, falling back to Sheets:", e)

    now = time.time()
    if _data_cache['customer_names'] is not None and (now - _data_cache['customer_names_time']) < CACHE_TTL:
        return _data_cache['customer_names']
        print("[db] Error getting customer names from DB, falling back to Sheets:", e)

    try:
        if customer_ws is None:
            return []
        values = customer_ws.get_all_values()
        if len(values) < 2:
            return []
        # Column B is "Name" (index 1)
        names = [row[1].strip() for row in values[1:] if len(row) > 1 and row[1].strip()]
        result = sorted(list(set(names)))
        _data_cache['customer_names']      = result
        _data_cache['customer_names_time'] = now
        return result
    except Exception as e:
        print("Error getting customer names from sheet:", e)
        # Return stale data if available rather than an empty list
        if _data_cache['customer_names'] is not None:
            return _data_cache['customer_names']
        return []

def get_customer_emails():
    """Returns a dict of {customer_name: email}"""
    if os.environ.get('DATABASE_URL'):
        try:
            customers = Customer.query.all()
            if customers:
                out = {}
                for c in customers:
                    if c.name.strip():
                        out[c.name.strip()] = c.email.strip() if c.email else ''
                return out
        except Exception as e:
            print("[db] Error getting customer emails from DB, falling back to Sheets:", e)

    now = time.time()
    if _data_cache['customer_emails'] is not None and (now - _data_cache['customer_emails_time']) < CACHE_TTL:
        return _data_cache['customer_emails']
        print("[db] Error getting customer emails from DB, falling back to Sheets:", e)

    try:
        if customer_ws is None:
            return {}
        values = customer_ws.get_all_values()
        if len(values) < 2:
            return {}
        # Name is Column B (index 1), Email is Column C (index 2)
        out = {}
        for row in values[1:]:
            if len(row) > 1 and row[1].strip():
                name = row[1].strip()
                email = row[2].strip() if len(row) > 2 else ''
                out[name] = email
        _data_cache['customer_emails']      = out
        _data_cache['customer_emails_time'] = now
        return out
    except Exception as e:
        print("Error getting customer emails from sheet:", e)
        # Return stale data if available rather than an empty dict
        if _data_cache['customer_emails'] is not None:
            return _data_cache['customer_emails']
        return {}

# Helper to ensure customers sheet has phone column D and address column E
def ensure_customer_columns():
    """Ensures that the Customers sheet contains necessary columns (Phone in D, Address in E)"""
    global customer_ws
    try:
        if customer_ws is None:
            if doc:
                try: customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
                except Exception: return
            else:
                return
        if customer_ws:
            rows = customer_ws.get_all_values()
            if len(rows) > 0:
                headers = rows[0]
                if len(headers) < 4:
                    customer_ws.update_cell(1, 4, "Phone")
                if len(headers) < 5:
                    customer_ws.update_cell(1, 5, "Address")
    except Exception as e:
        print("Error ensuring customer columns:", e)

# Backward compatibility alias
def ensure_phone_column():
    ensure_customer_columns()

def rename_customer_in_sheets(old_name, new_name):
    """Cascades a customer name change across registry, log, and map sheets, and database."""
    global cyl_ws, map_ws, scan_ws
    old_u = old_name.strip().upper()
    new_n = new_name.strip()
    
    if os.environ.get('DATABASE_URL'):
        try:
            cust = Customer.query.filter(Customer.name.ilike(old_name.strip())).first()
            if cust:
                cust.name = new_n
            cyls = Cylinder.query.filter(Cylinder.location.ilike(old_name.strip())).all()
            for c in cyls:
                c.location = new_n
            scans = Scan.query.filter(Scan.customer.ilike(old_name.strip())).all()
            for s in scans:
                s.customer = new_n
            cmaps = CustomerMap.query.filter(CustomerMap.customer.ilike(old_name.strip())).all()
            for cm in cmaps:
                cm.customer = new_n
            db.session.commit()
            print(f"[db] Cascaded customer rename from '{old_name}' to '{new_name}' in DB.")
        except Exception as e:
            db.session.rollback()
            print("[db] Error cascading customer rename in DB:", e)

    # 1. Update Cylinders Registry (Column F: Current Location)
    try:
        if cyl_ws is None and doc:
            try: cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
            except Exception: pass
        if cyl_ws:
            rows = cyl_ws.get_all_values()
            for idx, r in enumerate(rows):
                if idx == 0: continue
                if len(r) >= 6 and r[5].strip().upper() == old_u:
                    cyl_ws.update_cell(idx + 1, 6, new_n)
    except Exception as e:
        print("Error renaming customer in Cylinders registry:", e)
        
    # 2. Update Customer Map (Column G: Customer Name)
    try:
        if map_ws is None and doc:
            try: map_ws = doc.worksheet(MAP_SHEET_NAME)
            except Exception: pass
        if map_ws:
            rows = map_ws.get_all_values()
            for idx, r in enumerate(rows):
                if idx == 0: continue
                if len(r) >= 7 and r[6].strip().upper() == old_u:
                    map_ws.update_cell(idx + 1, 7, new_n)
    except Exception as e:
        print("Error renaming customer in Customer Map:", e)
        
    # 3. Update Scan Log Sheet1 (Column F: Customer)
    try:
        if scan_ws is None and doc:
            try: scan_ws = doc.worksheet(SCAN_SHEET_NAME)
            except Exception: pass
        if scan_ws:
            rows = scan_ws.get_all_values()
            for idx, r in enumerate(rows):
                if idx == 0: continue
                if len(r) >= 6 and r[5].strip().upper() == old_u:
                    scan_ws.update_cell(idx + 1, 6, new_n)
    except Exception as e:
        print("Error renaming customer in Scan Log:", e)


# ── System Settings helpers ──────────────────────────────────────────────────
SETTINGS_SHEET_NAME = "Settings"

def get_setting(key, default_value):
    """Gets setting from DB first, then from Google Sheets, with fallback."""
    if os.environ.get('DATABASE_URL'):
        try:
            s = SystemSetting.query.filter_by(key=key).first()
            if s:
                return s.value
        except Exception as e:
            print("[settings] Error reading setting from DB:", e)
            
    # Try reading from Sheets if not found in DB
    global doc
    if doc:
        try:
            ws = doc.worksheet(SETTINGS_SHEET_NAME)
            records = ws.get_all_values()
            for r in records:
                if len(r) >= 2 and r[0].strip().lower() == key.strip().lower():
                    val = r[1].strip()
                    # Cache in DB if possible
                    if os.environ.get('DATABASE_URL') and val:
                        try:
                            s = SystemSetting(key=key, value=val)
                            db.session.add(s)
                            db.session.commit()
                        except Exception:
                            db.session.rollback()
                    return val
        except Exception:
            # Settings worksheet might not exist yet
            pass
            
    return default_value

def set_setting(key, value):
    """Sets setting in DB and updates Google Sheets in the background."""
    if os.environ.get('DATABASE_URL'):
        try:
            s = SystemSetting.query.filter_by(key=key).first()
            if not s:
                s = SystemSetting(key=key, value=value)
                db.session.add(s)
            else:
                s.value = value
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print("[settings] Error saving setting to DB:", e)
            
    # Background update to Google Sheets
    def background_save_setting_sheets():
        global doc
        if not doc:
            return
        try:
            try:
                ws = doc.worksheet(SETTINGS_SHEET_NAME)
            except Exception:
                # Create the worksheet if missing
                ws = doc.add_worksheet(title=SETTINGS_SHEET_NAME, rows=50, cols=2)
                ws.append_row(["Setting Key", "Setting Value"])
                
            rows = ws.get_all_values()
            row_num = None
            for idx, r in enumerate(rows):
                if len(r) > 0 and r[0].strip().lower() == key.strip().lower():
                    row_num = idx + 1
                    break
            if row_num:
                ws.update(f'B{row_num}', [[value]])
            else:
                ws.append_row([key, value])
        except Exception as se:
            print("[sheets] Error saving setting to Sheets:", se)

    async_sheets_write(background_save_setting_sheets)


# ── Cylinder Registry helpers ──────────────────────────────────────────────
def get_all_cylinders():
    """Returns list of dicts from Cylinders table, falling back to Sheets."""
    global cyl_ws
    if os.environ.get('DATABASE_URL'):
        try:
            cyls = Cylinder.query.all()
            return [{
                'uid'           : c.uid,
                'gas_type'      : c.gas_type or '',
                'cylinder_type' : c.cylinder_type or '',
                'owner'         : c.owner or '',
                'status'        : c.status or 'Active',
                'location'      : c.location or 'Depot',
                'last_activity' : c.last_activity_date or '',
            } for c in cyls]
        except Exception as e:
            print("[db] Error getting cylinders from DB, falling back to Sheets:", e)

    now = time.time()
    if _data_cache['cylinders'] is not None and (now - _data_cache['cylinders_time']) < CACHE_TTL:
        return _data_cache['cylinders']

    try:
        if cyl_ws is None:
            if doc:
                try:
                    cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                except Exception:
                    return []
            else:
                return []
        rows = cyl_ws.get_all_values()
        if len(rows) < 2:
            return []
        out = []
        for r in rows[1:]:
            if len(r) >= 1 and r[0].strip():
                out.append({
                    'uid'           : r[0].strip() if len(r) > 0 else '',
                    'gas_type'      : r[1].strip() if len(r) > 1 else '',
                    'cylinder_type' : r[2].strip() if len(r) > 2 else '',
                    'owner'         : r[3].strip() if len(r) > 3 else '',
                    'status'        : r[4].strip() if len(r) > 4 else 'Active',
                    'location'      : r[5].strip() if len(r) > 5 else 'Depot',
                    'last_activity' : r[6].strip() if len(r) > 6 else '',
                })
        _data_cache['cylinders']      = out
        _data_cache['cylinders_time'] = now
        return out
    except Exception as e:
        print("Error getting cylinders:", e)
        # Return stale data if available
        if _data_cache['cylinders'] is not None:
            return _data_cache['cylinders']
        return []

def get_all_maintenance():
    """Returns dict of {uid: maintenance_dict} from Cylinder Maintenance table, falling back to Sheets."""
    global cyl_maint_ws
    if os.environ.get('DATABASE_URL'):
        try:
            maints = CylinderMaintenance.query.all()
            out = {}
            for m in maints:
                if m.cylinder_uid:
                    out[m.cylinder_uid] = {
                        'uid'              : m.cylinder_uid,
                        'water_capacity'   : m.water_capacity or '',
                        'fill_pressure'    : m.fill_pressure or '',
                        'gas_capacity'     : m.gas_capacity or '',
                        'unit'             : m.unit or '',
                        'is_mixture'       : m.is_mixture or 'No',
                        'mix_ratio'        : m.mix_ratio or '',
                        'manufacture_date' : m.manufacture_date or '',
                        'last_hydro_date'  : m.last_hydro_date or '',
                        'next_hydro_due'   : m.next_hydro_due or '',
                        'hydro_test_status': m.hydro_test_status or '',
                        'cert_no'          : m.cert_no or '',
                        'is_uhp'           : m.is_uhp or 'No',
                    }
            return out
        except Exception as e:
            print("[db] Error getting maintenance data from DB, falling back to Sheets:", e)

    now = time.time()
    if _data_cache['maintenance'] is not None and (now - _data_cache['maintenance_time']) < CACHE_TTL:
        return _data_cache['maintenance']
        print("[db] Error getting maintenance data from DB, falling back to Sheets:", e)

    try:
        if cyl_maint_ws is None:
            if doc:
                try:
                    cyl_maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
                except Exception:
                    return {}
            else:
                return {}
        rows = cyl_maint_ws.get_all_values()
        if len(rows) < 2:
            return {}
        out = {}
        for r in rows[1:]:
            uid = r[0].strip() if len(r) > 0 else ''
            if not uid:
                continue
            out[uid] = {
                'uid'              : uid,
                'water_capacity'   : r[1].strip() if len(r) > 1 else '',
                'fill_pressure'    : r[2].strip() if len(r) > 2 else '',
                'gas_capacity'     : r[3].strip() if len(r) > 3 else '',
                'unit'             : r[4].strip() if len(r) > 4 else '',
                'is_mixture'       : r[5].strip() if len(r) > 5 else 'No',
                'mix_ratio'        : r[6].strip() if len(r) > 6 else '',
                'manufacture_date' : r[7].strip() if len(r) > 7 else '',
                'last_hydro_date'  : r[8].strip() if len(r) > 8 else '',
                'next_hydro_due'   : r[9].strip() if len(r) > 9 else '',
                'hydro_test_status': r[10].strip() if len(r) > 10 else '',
                'cert_no'          : r[11].strip() if len(r) > 11 else '',
                'is_uhp'           : r[12].strip() if len(r) > 12 else 'No',
            }
        _data_cache['maintenance']      = out
        _data_cache['maintenance_time'] = now
        return out
    except Exception as e:
        print("Error getting maintenance data:", e)
        # Return stale data if available
        if _data_cache['maintenance'] is not None:
            return _data_cache['maintenance']
        return {}
        return {}

def compute_hydro_badge(next_hydro_due_str):
    """Returns ('OK'|'Due Soon'|'Overdue'|'Not Set') based on next_hydro_due date string."""
    if not next_hydro_due_str:
        return 'Not Set'
    d = parse_date(next_hydro_due_str)
    if not d:
        return 'Not Set'
    today = date.today()
    delta = (d - today).days
    if delta < 0:
        return 'Overdue'
    elif delta <= 30:
        return 'Due Soon'
    else:
        return 'OK'

def merge_cylinder_data():
    """Merges Cylinders + Cylinder Maintenance sheets, computes hydro badge, and merges history/customer logs."""
    cylinders   = get_all_cylinders()
    maintenance = get_all_maintenance()
    
    # Load last customer mappings
    last_customers = {}
    try:
        if os.environ.get('DATABASE_URL'):
            scans = Scan.query.order_by(Scan.created_at.desc()).all()
            for s in scans:
                uid_upper = s.cylinder_uid.strip().upper()
                if uid_upper not in last_customers and s.customer:
                    last_customers[uid_upper] = s.customer
        else:
            scans = get_scan_rows()
            for s in reversed(scans):
                uid_upper = s['uid'].strip().upper()
                if uid_upper not in last_customers and s.get('customer'):
                    last_customers[uid_upper] = s['customer']
    except Exception as e:
        print("Error building last_customers map:", e)

    # Load latest DuraGasHistory mappings
    latest_history = {}
    try:
        if os.environ.get('DATABASE_URL'):
            histories = DuraGasHistory.query.order_by(DuraGasHistory.created_at.desc()).all()
            for h in histories:
                uid_upper = h.cylinder_uid.strip().upper()
                if uid_upper not in latest_history:
                    latest_history[uid_upper] = {
                        'previous_gas': h.previous_gas or '',
                        'purge_required': 'Yes' if h.purge_required else 'No'
                    }
    except Exception as e:
        print("Error building latest_history map:", e)

    merged = []
    for cyl in cylinders:
        maint = maintenance.get(cyl['uid'], {})
        hydro_badge = compute_hydro_badge(maint.get('next_hydro_due', ''))
        
        uid_upper = cyl['uid'].strip().upper()
        # Default Dura fields
        prev_gas = ''
        purge_req = 'No'
        last_cust = last_customers.get(uid_upper, '')
        
        hist = latest_history.get(uid_upper)
        if hist:
            prev_gas = hist['previous_gas']
            purge_req = hist['purge_required']
            
        merged.append({
            **cyl,
            **maint,
            'hydro_badge': hydro_badge,
            'previous_gas': prev_gas,
            'purge_required': purge_req,
            'last_customer': last_cust
        })
    return merged

def find_cylinder_rows(uid):
    """Returns (cyl_row_1indexed, maint_row_1indexed) for a given UID. Returns (None, None) if not found."""
    global cyl_ws, cyl_maint_ws
    cyl_row = None
    maint_row = None
    try:
        if cyl_ws:
            rows = cyl_ws.get_all_values()
            for i, r in enumerate(rows):
                if i == 0:
                    continue
                if r and r[0].strip().upper() == uid.strip().upper():
                    cyl_row = i + 1
                    break
    except Exception as e:
        print("Error finding cylinder row:", e)
    try:
        if cyl_maint_ws:
            rows = cyl_maint_ws.get_all_values()
            for i, r in enumerate(rows):
                if i == 0:
                    continue
                if r and r[0].strip().upper() == uid.strip().upper():
                    maint_row = i + 1
                    break
    except Exception as e:
        print("Error finding maintenance row:", e)
    return cyl_row, maint_row


# In-memory TTL data cache configurations
# TTL is 60 s — all write routes call clear_cache() immediately after mutating
# data, so the dashboard is never stale after an intentional change.
_data_cache = {
    'scans'              : None, 'scans_time'          : 0,
    'map'                : None, 'map_time'            : 0,
    'cylinders'          : None, 'cylinders_time'      : 0,
    'maintenance'        : None, 'maintenance_time'    : 0,
    'customer_names'     : None, 'customer_names_time' : 0,
    'customer_emails'    : None, 'customer_emails_time': 0,
}
CACHE_TTL = 60  # seconds (was 10 — safe to raise because writes always clear the cache)

def clear_cache():
    """Wipes every cached value so the next read fetches fresh data from Sheets."""
    for key in list(_data_cache.keys()):
        _data_cache[key] = 0 if key.endswith('_time') else None

def sheets_write_with_retry(fn, *args, retries=3, **kwargs):
    """Call a gspread write function with exponential-backoff retry.

    Usage:  sheets_write_with_retry(ws.append_row, row_data)
    Catches rate-limit / transient errors (HTTP 429 / 503) and retries
    up to `retries` times with 1 s, 2 s, 4 s waits between attempts.
    """
    for attempt in range(retries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            if attempt < retries - 1:
                wait = 2 ** attempt  # 1 s → 2 s → 4 s
                print(f"[sheets_write] error on attempt {attempt + 1}/{retries}, "
                      f"retrying in {wait}s: {e}")
                time.sleep(wait)
            else:
                raise



# ================================================================
#  AUTH HELPERS
# ================================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect('/login')
        if session['user']['role'] not in ['manager', 'owner']:
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def check_login(username, password):
    try:
        if os.environ.get('DATABASE_URL'):
            u = User.query.filter(db.func.lower(User.username) == username.lower()).first()
            if u:
                password_valid = False
                is_hashed = u.password.startswith('pbkdf2:') or u.password.startswith('scrypt:')
                
                if is_hashed:
                    password_valid = check_password_hash(u.password, password)
                else:
                    # Plain text comparison
                    password_valid = (u.password == password)
                    if password_valid:
                        # Auto-upgrade the plain text password to hashed format in database
                        try:
                            u.password = generate_password_hash(password)
                            db.session.commit()
                            print(f"[security] Successfully auto-upgraded password hash for user: {u.username}")
                        except Exception as upgrade_err:
                            db.session.rollback()
                            print(f"[security] Error auto-upgrading password hash for {u.username}:", upgrade_err)
                
                if password_valid:
                    return {
                        'username': u.username,
                        'role': u.role.lower() if u.role else 'driver',
                        'name': u.name or u.username
                    }
    except Exception as e:
        print("[db] Login check error, falling back to Sheets:", e)

    try:
        if users_ws is None:
            return None
        records  = users_ws.get_all_records()
        for r in records:
            u = str(r.get('Username', '')).strip()
            p = str(r.get('Password', '')).strip()
            if u.lower() == username.lower() and p == password:
                return {
                    'username': u,
                    'role'    : str(r.get('Role', 'driver')).strip().lower(),
                    'name'    : str(r.get('Name', u)).strip()
                }
    except Exception as e:
        print("Login check error:", e)
    return None


# ================================================================
#  DATE HELPERS
# ================================================================

def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def fmt_date(value):
    if not value:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d-%m-%Y')
    if isinstance(value, date):
        return value.strftime('%d-%m-%Y')
    return str(value).strip()


# ================================================================
#  DATA HELPERS  (all read from Google Sheets)
# ================================================================

def get_scan_rows():
    """List of dicts from Sheet1: date, time, driver, action, uid, customer, gas_type (cached)"""
    if os.environ.get('DATABASE_URL'):
        try:
            scans = Scan.query.all()
            if scans:
                return [{
                    'date': s.scan_date,
                    'time': s.scan_time or '',
                    'driver': s.driver or '',
                    'action': s.action,
                    'uid': s.cylinder_uid,
                    'customer': s.customer or '',
                    'gas_type': s.gas_type or ''
                } for s in scans]
        except Exception as e:
            print("[db] Error reading scans from DB, falling back to Sheets:", e)

    now = time.time()
    if _data_cache['scans'] is not None and (now - _data_cache['scans_time']) < CACHE_TTL:
        return _data_cache['scans']

    try:
        if scan_ws is None:
            return []
        rows = scan_ws.get_all_values()
        out  = []
        for r in rows[1:]:
            if len(r) >= 5 and r[4].strip():
                out.append({
                    'date'  : r[0].strip(),
                    'time'  : r[1].strip(),
                    'driver': r[2].strip(),
                    'action': r[3].strip(),
                    'uid'   : r[4].strip(),
                    'customer': r[5].strip() if len(r) > 5 else '',
                    'gas_type': r[6].strip() if len(r) > 6 else ''
                })
        _data_cache['scans'] = out
        _data_cache['scans_time'] = now
        return out
    except Exception as e:
        print("Error reading scans:", e)
        if _data_cache['scans'] is not None:
            return _data_cache['scans']
        return []

def get_map_rows():
    """List of dicts from Customer Map: date, time, driver, action, customer (cached)"""
    if os.environ.get('DATABASE_URL'):
        try:
            maps = CustomerMap.query.all()
            if maps:
                return [{
                    'date': m.scan_date,
                    'time': m.scan_time or '',
                    'driver': m.driver or '',
                    'action': m.action or '',
                    'customer': m.customer or ''
                } for m in maps]
        except Exception as e:
            print("[db] Error reading map from DB, falling back to Sheets:", e)

    now = time.time()
    if _data_cache['map'] is not None and (now - _data_cache['map_time']) < CACHE_TTL:
        return _data_cache['map']
        print("[db] Error reading map from DB, falling back to Sheets:", e)

    try:
        if map_ws is None:
            return []
        rows = map_ws.get_all_values()
        out  = []
        for r in rows[1:]:
            if len(r) >= 7 and r[6].strip():
                out.append({
                    'date'    : r[0].strip(),
                    'time'    : r[1].strip(),
                    'driver'  : r[2].strip(),
                    'action'  : r[3].strip(),
                    'customer': r[6].strip()
                })
        _data_cache['map'] = out
        _data_cache['map_time'] = now
        return out
    except Exception as e:
        print("Error reading map:", e)
        if _data_cache['map'] is not None:
            return _data_cache['map']
        return []

def build_batch_map():
    """dict: 'date||time||driver||action' → customer"""
    result = {}
    for r in get_map_rows():
        key = f"{r['date']}||{r['time']}||{r['driver']}||{r['action']}"
        result[key] = r['customer']
    return result

def build_events():
    """Sorted list of scan events enriched with customer name"""
    batch_map = build_batch_map()
    events    = []
    for r in get_scan_rows():
        customer = r.get('customer', '').strip()
        if not customer:
            key      = f"{r['date']}||{r['time']}||{r['driver']}||{r['action']}"
            customer = batch_map.get(key)
        if not customer:
            continue
        events.append({**r, 'customer': customer, 'date_obj': parse_date(r['date'])})
    events.sort(key=lambda x: (x['date_obj'] or date.min, x['time']))
    return events

def get_activity_events():
    """Sorted list of all activity scans from Sheet1, grouped by submission, enriched with Customer column or batch_map fallback"""
    try:
        if os.environ.get('DATABASE_URL'):
            scans = Scan.query.all()
            if scans:
                grouped = {}
                for s in scans:
                    date_str = s.scan_date
                    time_str = s.scan_time or ''
                    driver = s.driver or ''
                    action = s.action
                    customer = s.customer or ''
                    cust_display = customer if customer else ('Depot' if action == 'Filling' else '—')
                    
                    group_key = (date_str, time_str, driver, action, cust_display)
                    if group_key not in grouped:
                        grouped[group_key] = []
                    grouped[group_key].append({
                        'uid': s.cylinder_uid,
                        'gas_type': s.gas_type or ''
                    })
                
                events = []
                for (date_str, time_str, driver, action, cust_display), uids in grouped.items():
                    events.append({
                        'date': date_str,
                        'time': time_str,
                        'driver': driver,
                        'action': action,
                        'uids': uids,
                        'customer': cust_display,
                        'date_obj': parse_date(date_str)
                    })
                events.sort(key=lambda x: (x['date_obj'] or date.min, x['time']), reverse=True)
                return events
    except Exception as e:
        print("[db] Error getting activity events from DB, falling back to Sheets:", e)

    try:
        if scan_ws is None:
            return []
        rows = scan_ws.get_all_values()
        if len(rows) < 2:
            return []
        
        batch_map = build_batch_map()
        grouped = {}
        for r in rows[1:]:
            if len(r) >= 5 and r[4].strip():
                uid = r[4].strip()
                date_str = r[0].strip()
                time_str = r[1].strip()
                driver = r[2].strip()
                action = r[3].strip()
                gas_type = r[6].strip() if len(r) > 6 else ''
                
                # Retrieve from Column F (index 5) if present, otherwise fall back to batch_map
                customer = r[5].strip() if len(r) > 5 else ''
                if not customer:
                    key = f"{date_str}||{time_str}||{driver}||{action}"
                    customer = batch_map.get(key, '')
                
                # Standardize Customer name display
                cust_display = customer if customer else ('Depot' if action == 'Filling' else '—')
                
                group_key = (date_str, time_str, driver, action, cust_display)
                if group_key not in grouped:
                    grouped[group_key] = []
                grouped[group_key].append({
                    'uid': uid,
                    'gas_type': gas_type
                })
        
        events = []
        for (date_str, time_str, driver, action, cust_display), uids in grouped.items():
            events.append({
                'date': date_str,
                'time': time_str,
                'driver': driver,
                'action': action,
                'uids': uids,
                'customer': cust_display,
                'date_obj': parse_date(date_str)
            })
        
        # Sort newest first for chronological view
        events.sort(key=lambda x: (x['date_obj'] or date.min, x['time']), reverse=True)
        return events
    except Exception as e:
        print("Error getting activity events:", e)
        return []

def get_mapping_mismatches():
    """
    Finds mismatches between Delivery and Collection customers for all cylinders.
    A mismatch occurs when a cylinder is collected from Customer B, but was last
    delivered to Customer A (where A != B).
    """
    events = build_events()
    cylinder_owner = {}
    mismatches = []

    for ev in events:
        uid = ev['uid']
        cust = ev['customer']
        
        if ev['action'] == 'Delivery':
            cylinder_owner[uid] = {
                'customer': cust,
                'date': ev['date'],
                'time': ev['time'],
                'driver': ev['driver'],
                'action': 'Delivery'
            }
        elif ev['action'] == 'Collection':
            prev = cylinder_owner.pop(uid, None)
            if prev and prev['customer'].lower() != cust.lower():
                mismatches.append({
                    'uid': uid,
                    'delivery': prev,
                    'collection': {
                        'customer': cust,
                        'date': ev['date'],
                        'time': ev['time'],
                        'driver': ev['driver'],
                        'action': 'Collection'
                    }
                })
        elif ev['action'] == 'Filling':
            cylinder_owner.pop(uid, None)

    return mismatches


def build_outstanding():
    """Outstanding cylinders per customer"""
    events          = build_events()
    cylinder_owner  = {}
    customer_stats  = {}

    for ev in events:
        c = ev['customer']
        if c not in customer_stats:
            customer_stats[c] = {'total_delivered': 0, 'total_collected': 0, 'last_activity': ev['date']}
        customer_stats[c]['last_activity'] = ev['date']

        if ev['action'] == 'Delivery':
            cylinder_owner[ev['uid']] = c
            customer_stats[c]['total_delivered'] += 1
        elif ev['action'] == 'Collection':
            cylinder_owner.pop(ev['uid'], None)
            customer_stats[c]['total_collected'] += 1
        elif ev['action'] == 'Filling':
            cylinder_owner.pop(ev['uid'], None)

    customer_outstanding = {}
    for uid, cust in cylinder_owner.items():
        customer_outstanding.setdefault(cust, []).append(uid)

    result = []
    for cust, stats in customer_stats.items():
        uid_list = customer_outstanding.get(cust, [])
        result.append({
            'customer'       : cust,
            'total_delivered': stats['total_delivered'],
            'total_collected': stats['total_collected'],
            'outstanding'    : len(uid_list),
            'cylinder_uids'  : uid_list,              # list — for template iteration
            'uids'           : ', '.join(uid_list) if uid_list else '—',  # string fallback
            'last_activity'  : stats['last_activity']
        })

    result.sort(key=lambda x: x['outstanding'], reverse=True)
    return result

def build_aging():
    """Days outstanding per cylinder currently with a customer"""
    events                 = build_events()
    cylinder_owner         = {}
    cylinder_delivery_date = {}

    for ev in events:
        if ev['action'] == 'Delivery':
            cylinder_owner[ev['uid']]         = ev['customer']
            cylinder_delivery_date[ev['uid']] = ev['date_obj']
        elif ev['action'] in ('Collection', 'Filling'):
            cylinder_owner.pop(ev['uid'], None)
            cylinder_delivery_date.pop(ev['uid'], None)

    today  = date.today()
    result = []
    for uid, cust in cylinder_owner.items():
        d_date   = cylinder_delivery_date.get(uid)
        days_out = (today - d_date).days if d_date else None
        status   = (
            'overdue' if days_out and days_out > 20 else
            'warning' if days_out and days_out >= 10 else
            'ok'
        )
        result.append({
            'cylinder_uid' : uid,          # matches aging.html template
            'uid'          : uid,          # kept for backward compat
            'customer'     : cust,
            'delivered_on' : fmt_date(d_date),   # matches aging.html template
            'delivery_date': fmt_date(d_date),   # kept for backward compat
            'days_out'     : days_out,
            'status'       : status
        })

    result.sort(key=lambda x: (x['days_out'] or 0), reverse=True)
    return result

def build_rotation(from_date=None, to_date=None, gas_filter='', customer_filter='', direction_filter=''):
    """Builds cylinder rotation data — every in/out movement event in a date range.

    Direction logic:
      'out'  = Delivery (cylinder left depot to customer)
      'in'   = Collection OR Filling (cylinder is back at depot)
    """
    events = build_events()
    all_scan_rows = get_scan_rows()  # for Filling events (no customer in batch_map)

    # Master registry lookup for gas types
    cyls = get_all_cylinders()
    cyl_gas_map = {c['uid'].strip().upper(): c['gas_type'].strip().upper() for c in cyls if c.get('gas_type')}

    # Build a delivery-date lookup per UID so we can compute days_out for returns
    # uid -> list of (delivery_date, customer) tuples, chronological
    delivery_log = {}
    for ev in events:
        uid = ev['uid']
        if ev['action'] == 'Delivery':
            delivery_log.setdefault(uid, []).append((ev['date_obj'], ev.get('customer', '')))

    movements = []

    # ── Delivery + Collection from events (have customer) ──────────────────
    for ev in events:
        d = ev.get('date_obj')
        if from_date and d and d < from_date:
            continue
        if to_date and d and d > to_date:
            continue

        action = ev.get('action', '')
        uid    = ev.get('uid', '')
        uid_upper = uid.strip().upper()
        gas = ev.get('gas_type', '').strip().upper()
        if not gas:
            if uid_upper in cyl_gas_map:
                gas = cyl_gas_map[uid_upper]
            else:
                gas = uid.split('-')[0].upper() if '-' in uid else ''

        if gas_filter and gas != gas_filter.upper():
            continue

        customer = ev.get('customer', '')
        if customer_filter and customer.lower() != customer_filter.lower():
            continue

        if action == 'Delivery':
            direction = 'out'
        elif action == 'Collection':
            direction = 'in'
        else:
            continue   # Filling handled separately below

        if direction_filter and direction != direction_filter:
            continue

        # For Collection: find how many days out since last delivery to same customer
        days_out = None
        if action == 'Collection' and d:
            deliveries_for_uid = delivery_log.get(uid, [])
            # Last delivery to same customer before this collection date
            prior = [x for x in deliveries_for_uid if x[0] < d and x[1].lower() == customer.lower()]
            if prior:
                last_del = max(prior, key=lambda x: x[0])
                days_out = (d - last_del[0]).days

        movements.append({
            'uid'      : uid,
            'gas_type' : gas,
            'direction': direction,
            'action'   : action,
            'customer' : customer,
            'driver'   : ev.get('driver', ''),
            'date'     : ev.get('date', ''),
            'time'     : ev.get('time', ''),
            'date_obj' : d,
            'days_out' : days_out,
        })
    # Sort chronologically descending
    movements.sort(key=lambda x: (x['date_obj'] or date.min, x['time']), reverse=True)
    return movements


def build_cylinder_journey(uid):
    """Returns the full chronological journey of a single cylinder UID."""
    events    = build_events()
    all_rows  = get_scan_rows()
    journey   = []

    for ev in events:
        if ev.get('uid', '').strip().upper() != uid.strip().upper():
            continue
        journey.append({
            'action'  : ev['action'],
            'customer': ev.get('customer', ''),
            'driver'  : ev.get('driver', ''),
            'date'    : ev.get('date', ''),
            'time'    : ev.get('time', ''),
            'date_obj': ev.get('date_obj'),
        })

    for r in all_rows:
        if r.get('uid', '').strip().upper() != uid.strip().upper():
            continue
        if r.get('action') != 'Filling':
            continue
        d = parse_date(r.get('date', ''))
        journey.append({
            'action'  : 'Filling',
            'customer': '— Depot —',
            'driver'  : r.get('driver', ''),
            'date'    : r.get('date', ''),
            'time'    : r.get('time', ''),
            'date_obj': d,
        })

    journey.sort(key=lambda x: (x['date_obj'] or date.min, x['time']))
    return journey

def build_driver_stats():
    """Delivery / collection stats per driver"""
    scan_rows   = get_scan_rows()
    today       = date.today()
    week_start  = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    stats       = {}

    for r in scan_rows:
        d = r['driver']
        if d not in stats:
            stats[d] = {
                'driver'           : d,
                'total_deliveries' : 0,
                'total_collections': 0,
                'this_week'        : 0,
                'this_month'       : 0,
                'last_active'      : r['date']
            }
        stats[d]['last_active'] = r['date']
        rd = parse_date(r['date'])

        if r['action'] == 'Delivery':
            stats[d]['total_deliveries'] += 1
            if rd and rd >= week_start : stats[d]['this_week']  += 1
            if rd and rd >= month_start: stats[d]['this_month'] += 1
        elif r['action'] == 'Collection':
            stats[d]['total_collections'] += 1

    result = list(stats.values())
    result.sort(key=lambda x: x['total_deliveries'] + x['total_collections'], reverse=True)
    return result

def build_daily_movement():
    """Cylinders delivered vs collected per day (last 30 days)"""
    scan_rows = get_scan_rows()
    daily     = {}

    for r in scan_rows:
        d = r['date']
        if d not in daily:
            daily[d] = {'date': d, 'delivered': 0, 'collected': 0, 'date_obj': parse_date(d)}
        if   r['action'] == 'Delivery'  : daily[d]['delivered'] += 1  # matches movement.html
        elif r['action'] == 'Collection': daily[d]['collected'] += 1

    result = sorted(daily.values(), key=lambda x: x['date_obj'] or date.min)
    return result[-30:]

def get_cylinder_history(uid):
    """Full event history for a single cylinder UID"""
    batch_map = build_batch_map()
    history   = []

    for r in get_scan_rows():
        if r['uid'].upper() == uid.strip().upper():
            key      = f"{r['date']}||{r['time']}||{r['driver']}||{r['action']}"
            customer = batch_map.get(key, '(Not mapped yet)')
            history.append({**r, 'customer': customer})

    history.sort(key=lambda x: (parse_date(x['date']) or date.min, x['time']))
    return history

def get_cylinder_status(uid):
    uid_upper = uid.strip().upper()
    
    # Check if the UID is registered in Cylinders sheet
    registered = False
    try:
        cyls = get_all_cylinders()
        registered = any(c['uid'].strip().upper() == uid_upper for c in cyls)
    except Exception:
        registered = False
        
    scan_rows = get_scan_rows()
    history = [r for r in scan_rows if r['uid'].strip().upper() == uid_upper]
    
    if not history:
        return {'status': 'Empty', 'owner': None, 'date': None, 'registered': registered}
    
    history.sort(key=lambda x: (parse_date(x['date']) or date.min, x['time']))
    last_event = history[-1]
    action = last_event['action']
    
    if action == 'Filling':
        return {'status': 'Filled', 'owner': 'Depot', 'date': last_event['date'], 'registered': registered}
    elif action == 'Delivery':
        batch_map = build_batch_map()
        key = f"{last_event['date']}||{last_event['time']}||{last_event['driver']}||{last_event['action']}"
        customer = batch_map.get(key, '(Unknown Customer)')
        return {'status': 'Delivered', 'owner': customer, 'date': last_event['date'], 'registered': registered}
    elif action == 'Collection':
        return {'status': 'Empty', 'owner': 'Depot', 'date': last_event['date'], 'registered': registered}
        
    return {'status': 'Empty', 'owner': None, 'date': None, 'registered': registered}

@app.route('/api/cylinder_status/<uid>')
def api_cylinder_status(uid):
    status_data = get_cylinder_status(uid)
    return jsonify(status_data)


@app.route('/admin/api/mapping_mismatches')
@admin_required
def admin_api_mapping_mismatches():
    mismatches = get_mapping_mismatches()
    return jsonify({'mismatches': mismatches})


# ================================================================
#  AUTH ROUTES
# ================================================================

@app.route('/login', methods=['GET', 'POST'])
@app.route('/admin/login', methods=['GET', 'POST'])   # alias — login.html posts here
def login():
    if 'user' in session:
        if session['user']['role'] in ['manager', 'owner']:
            return redirect('/admin')
        else:
            return redirect('/')

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user     = check_login(username, password)
        if user:
            session.permanent = True
            session['user'] = user
            if user['role'] in ['manager', 'owner']:
                return redirect('/admin')
            else:
                return redirect('/')
        return render_template('login.html', error='Invalid username or password')
    return render_template('login.html', error=None)

@app.route('/logout')
@app.route('/admin/logout')
def logout():
    session.clear()
    return redirect('/login')


# ================================================================
#  ADMIN ROUTES
# ================================================================

@app.route('/admin')
@admin_required
def admin():
    return redirect('/admin/dashboard')

@app.route('/admin/sync_from_sheets', methods=['POST'])
@admin_required
def admin_sync_from_sheets():
    user = session.get('user')
    if user.get('role') not in ['manager', 'owner']:
        flash("Unauthorized: Only managers/owners can trigger a manual sync.", "danger")
        return redirect('/admin/dashboard')
        
    try:
        from sync import sync_sheets_to_db
        if doc:
            success = sync_sheets_to_db(doc)
            if success:
                clear_cache()
                flash("Successfully synced all tables (Scans, Cylinders, Customers, etc.) from Google Sheets to database!", "success")
            else:
                flash("Sync completed with warnings. Please check the logs.", "warning")
        else:
            flash("Sheets document is not connected. Unable to sync.", "danger")
    except Exception as e:
        print("[sync] Manual sync error:", e)
        flash(f"Sync error: {str(e)}", "danger")
        
    return redirect('/admin/dashboard')

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    outstanding  = build_outstanding()
    total_out    = sum(c['outstanding'] for c in outstanding)
    cust_count   = len([c for c in outstanding if c['outstanding'] > 0])
    high_count   = len([c for c in outstanding if c['outstanding'] > 10])

    today_str    = date.today().strftime('%d-%m-%Y')
    scan_rows    = get_scan_rows()
    # Individual cylinder-action rows for today
    today_rows   = [r for r in scan_rows if r['date'] == today_str]
    # today_scans = cylinder-level count so it always matches the hover breakdown sum
    today_scans  = len(today_rows)

    top_customers = [c for c in outstanding if c['outstanding'] > 0][:5]

    # ── Hover-insight hints (derived in-memory, no extra DB queries) ──

    # Card 1: top 3 customers by outstanding count
    hint_top_cylinders = [
        {'customer': c['customer'], 'count': c['outstanding']}
        for c in outstanding if c['outstanding'] > 0
    ][:3]

    # Card 2: top 3 most recently active customers (outstanding > 0)
    recent = sorted(
        [c for c in outstanding if c['outstanding'] > 0],
        key=lambda x: x['last_activity'], reverse=True
    )[:3]
    hint_recent_customers = [
        {'customer': c['customer'], 'last_activity': c['last_activity']}
        for c in recent
    ]

    # Card 3: customers with outstanding > 10
    hint_high_customers = [
        {'customer': c['customer'], 'count': c['outstanding']}
        for c in outstanding if c['outstanding'] > 10
    ][:5]

    # Card 4: today's breakdown — Deliveries / Collections / Fillings
    hint_today_split = {
        'deliveries' : sum(1 for r in today_rows if r.get('action') == 'Delivery'),
        'collections': sum(1 for r in today_rows if r.get('action') == 'Collection'),
        'fillings'   : sum(1 for r in today_rows if r.get('action') == 'Filling'),
    }


    return render_template('dashboard.html',
        user                  = session['user'],
        total_out             = total_out,
        cust_count            = cust_count,
        high_count            = high_count,
        today_scans           = today_scans,
        top_customers         = top_customers,
        hint_top_cylinders    = hint_top_cylinders,
        hint_recent_customers = hint_recent_customers,
        hint_high_customers   = hint_high_customers,
        hint_today_split      = hint_today_split,
    )

@app.route('/admin/activity')
@admin_required
def admin_activity():
    events = get_activity_events()
    drivers = sorted(list(set(e['driver'] for e in events if e.get('driver'))))
    today_str = date.today().strftime('%d-%m-%Y')
    return render_template('activity.html',
        user      = session['user'],
        events    = events,
        drivers   = drivers,
        today_str = today_str
    )

@app.route('/admin/activity/delete', methods=['POST'])
@admin_required
def delete_activity():
    user = session.get('user')
    if user.get('role') not in ['manager', 'owner']:
        flash("Unauthorized: Only managers/owners can delete activity records.", "danger")
        return redirect('/admin/activity')
        
    date_str = request.form.get('date', '').strip()
    time_str = request.form.get('time', '').strip()
    driver   = request.form.get('driver', '').strip()
    action   = request.form.get('action', '').strip()
    customer = request.form.get('customer', '').strip()

    if not date_str or not action:
        flash("Invalid activity parameters.", "danger")
        return redirect('/admin/activity')

    db_written = False
    uids_to_revert = []
    rows_to_del_data = []
    
    if os.environ.get('DATABASE_URL'):
        try:
            # Find scans to delete in DB
            query = Scan.query.filter(
                Scan.scan_date == date_str,
                Scan.action == action
            )
            if time_str:
                query = query.filter(Scan.scan_time == time_str)
            if driver:
                query = query.filter(Scan.driver == driver)
            
            scans_to_delete = query.all()
            matched_scans = []
            for s in scans_to_delete:
                s_cust = s.customer or ''
                s_cust_display = s_cust if s_cust else ('Depot' if s.action == 'Filling' else '—')
                if s_cust_display.lower() == customer.lower() or s_cust.lower() == customer.lower():
                    matched_scans.append(s)
                    
            if not matched_scans:
                flash("No matching activity scans found in database.", "warning")
                return redirect('/admin/activity')
                
            uids_to_revert = list(set(s.cylinder_uid for s in matched_scans))
            rows_to_del_data = [{
                'date': s.scan_date,
                'time': s.scan_time or '',
                'driver': s.driver or '',
                'action': s.action,
                'uid': s.cylinder_uid
            } for s in matched_scans]
            
            # Delete the scans
            for s in matched_scans:
                db.session.delete(s)
                
            # Revert Cylinder statuses
            for uid in uids_to_revert:
                all_prev = Scan.query.filter(
                    Scan.cylinder_uid.ilike(uid)
                ).all()
                remaining = [s for s in all_prev if s.id not in [m.id for m in matched_scans]]
                
                c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
                if c_db:
                    if remaining:
                        remaining.sort(key=lambda s: (parse_date(s.scan_date) or date.min, s.scan_time or ''), reverse=True)
                        latest = remaining[0]
                        if latest.action == 'Delivery':
                            c_db.status = 'Delivered'
                            c_db.location = latest.customer or 'Customer'
                        elif latest.action == 'Collection':
                            c_db.status = 'Empty'
                            c_db.location = 'Depot'
                        elif latest.action == 'Filling':
                            c_db.status = 'Filled'
                            c_db.location = 'Depot'
                        c_db.last_activity_date = latest.scan_date
                    else:
                        c_db.status = 'Active'
                        c_db.location = 'Depot'
                        c_db.last_activity_date = ''

            # Delete from CustomerMap
            cmap_query = CustomerMap.query.filter(
                CustomerMap.scan_date == date_str,
                CustomerMap.action == action
            )
            if time_str:
                cmap_query = cmap_query.filter(CustomerMap.scan_time == time_str)
            if driver:
                cmap_query = cmap_query.filter(CustomerMap.driver == driver)
                
            cmaps = cmap_query.all()
            for cmap in cmaps:
                cmap_cust = cmap.customer or ''
                cmap_cust_display = cmap_cust if cmap_cust else ('Depot' if cmap.action == 'Filling' else '—')
                if cmap_cust_display.lower() == customer.lower() or cmap_cust.lower() == customer.lower():
                    db.session.delete(cmap)

            db.session.commit()
            db_written = True
            
        except Exception as dbe:
            db.session.rollback()
            print("[db] Error deleting activity scans:", dbe)
            flash(f"Database error deleting activity: {str(dbe)}", "danger")
            return redirect('/admin/activity')

    # Build revert map for sheets
    revert_map = {}
    if os.environ.get('DATABASE_URL'):
        for uid in uids_to_revert:
            c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
            if c_db:
                revert_map[uid.upper()] = {
                    'status': c_db.status or 'Active',
                    'location': c_db.location or 'Depot',
                    'last_activity': c_db.last_activity_date or ''
                }
            else:
                revert_map[uid.upper()] = {
                    'status': 'Active',
                    'location': 'Depot',
                    'last_activity': ''
                }
    else:
        rows_to_del_data = [{
            'date': date_str,
            'time': time_str,
            'driver': driver,
            'action': action,
            'uid': ''
        }]

    # Offload Sheets deletes to background
    def background_delete_activity_sheets():
        try:
            global sheet, cyl_ws, doc
            if sheet is None and doc:
                try: sheet = doc.worksheet(SCAN_SHEET_NAME)
                except Exception: pass
            if cyl_ws is None and doc:
                try: cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                except Exception: pass

            if sheet and rows_to_del_data:
                all_rows = sheet.get_all_values()
                rows_to_delete_indices = []
                for idx, r in enumerate(all_rows):
                    if idx == 0:
                        continue
                    r_date   = r[0].strip() if len(r) > 0 else ''
                    r_time   = r[1].strip() if len(r) > 1 else ''
                    r_driver = r[2].strip() if len(r) > 2 else ''
                    r_action = r[3].strip() if len(r) > 3 else ''
                    r_uid    = r[4].strip() if len(r) > 4 else ''
                    r_cust   = r[5].strip() if len(r) > 5 else ''
                    
                    for d_row in rows_to_del_data:
                        date_matches = (r_date == d_row['date'])
                        action_matches = (r_action == d_row['action'])
                        time_matches = (not d_row['time'] or r_time == d_row['time'])
                        driver_matches = (not d_row['driver'] or r_driver.lower() == d_row['driver'].lower())
                        uid_matches = (not d_row['uid'] or r_uid.lower() == d_row['uid'].lower())
                        
                        if date_matches and action_matches and time_matches and driver_matches and uid_matches:
                            r_cust_display = r_cust if r_cust else ('Depot' if r_action == 'Filling' else '—')
                            if r_cust_display.lower() == customer.lower() or r_cust.lower() == customer.lower():
                                rows_to_delete_indices.append(idx + 1)
                                break
                
                rows_to_delete_indices.sort(reverse=True)
                for row_idx in rows_to_delete_indices:
                    sheets_write_with_retry(sheet.delete_rows, row_idx)

            if cyl_ws and revert_map:
                cyl_rows = cyl_ws.get_all_values()
                uid_row_map = {}
                for idx, r in enumerate(cyl_rows):
                    if idx == 0:
                        continue
                    if r and r[0].strip():
                        uid_row_map[r[0].strip().upper()] = idx + 1
                
                batch_data = []
                for uid_upper, state in revert_map.items():
                    row_num = uid_row_map.get(uid_upper)
                    if row_num:
                        batch_data.append({
                            'range': f'E{row_num}:G{row_num}',
                            'values': [[state['status'], state['location'], state['last_activity']]]
                        })
                if batch_data:
                    sheets_write_with_retry(cyl_ws.batch_update, batch_data)
                    
        except Exception as se:
            print("[sheets] Error mirroring activity deletion to Sheets:", se)

    async_sheets_write(background_delete_activity_sheets)

    flash("Activity record deleted successfully.", "success")
    return redirect('/admin/activity')

@app.route('/admin/daily_summary')
@admin_required
def admin_daily_summary():
    target_date_str = request.args.get('date', '').strip()
    if target_date_str:
        try:
            # HTML input type="date" yields YYYY-MM-DD
            parsed_dt = datetime.strptime(target_date_str, '%Y-%m-%d')
            selected_date_str = parsed_dt.strftime('%d-%m-%Y')
            display_date_input = target_date_str
        except ValueError:
            try:
                # Fallback if in DD-MM-YYYY format
                parsed_dt = datetime.strptime(target_date_str, '%d-%m-%Y')
                selected_date_str = target_date_str
                display_date_input = parsed_dt.strftime('%Y-%m-%d')
            except ValueError:
                selected_date = date.today()
                selected_date_str = selected_date.strftime('%d-%m-%Y')
                display_date_input = selected_date.strftime('%Y-%m-%d')
    else:
        selected_date = date.today()
        selected_date_str = selected_date.strftime('%d-%m-%Y')
        display_date_input = selected_date.strftime('%Y-%m-%d')

    scan_rows = get_scan_rows()
    today_rows = [r for r in scan_rows if r['date'] == selected_date_str]

    # Group by driver — track counts AND UID lists per action
    driver_stats = {}
    for r in today_rows:
        d = r['driver'].strip()
        if not d:
            continue
        if d not in driver_stats:
            driver_stats[d] = {
                'deliveries': 0, 'collections': 0, 'fillings': 0, 'total': 0,
                'uid_deliveries': [], 'uid_collections': [], 'uid_fillings': []
            }
        action = r['action'].strip()
        uid    = r.get('uid', '').strip()
        if action == 'Delivery':
            driver_stats[d]['deliveries'] += 1
            if uid:
                driver_stats[d]['uid_deliveries'].append(uid)
        elif action == 'Collection':
            driver_stats[d]['collections'] += 1
            if uid:
                driver_stats[d]['uid_collections'].append(uid)
        elif action == 'Filling':
            driver_stats[d]['fillings'] += 1
            if uid:
                driver_stats[d]['uid_fillings'].append(uid)
        driver_stats[d]['total'] += 1

    total_deliveries  = sum(v['deliveries']  for v in driver_stats.values())
    total_collections = sum(v['collections'] for v in driver_stats.values())
    total_fillings    = sum(v['fillings']    for v in driver_stats.values())
    total_scans       = len(today_rows)

    outstanding = build_outstanding()
    outstanding_snapshot = [c for c in outstanding if c['outstanding'] > 0]
    total_out = sum(c['outstanding'] for c in outstanding_snapshot)

    return render_template('daily_summary.html',
        user               = session['user'],
        today              = selected_date_str,
        display_date_input = display_date_input,
        driver_stats       = driver_stats,
        total_deliveries   = total_deliveries,
        total_collections  = total_collections,
        total_fillings     = total_fillings,
        total_scans        = total_scans,
        outstanding_snapshot = outstanding_snapshot,
        total_out          = total_out,
    )

@app.route('/admin/outstanding')
@admin_required
def admin_outstanding():
    data = build_outstanding()
    return render_template('outstanding.html',
        user = session['user'],
        data = data,
        total_out = sum(c['outstanding'] for c in data)
    )

@app.route('/admin/aging')
@admin_required
def admin_aging():
    data     = build_aging()
    overdue  = [r for r in data if r['status'] == 'overdue']
    warning  = [r for r in data if r['status'] == 'warning']
    ok       = [r for r in data if r['status'] == 'ok']
    return render_template('aging.html',
        user    = session['user'],
        data    = data,
        overdue = len(overdue),
        warning = len(warning),
        ok_count= len(ok)
    )

@app.route('/admin/rotation')
@admin_required
def admin_rotation():
    today_str = date.today().isoformat()
    from_str         = request.args.get('from', today_str)
    to_str           = request.args.get('to',   today_str)
    gas_filter       = request.args.get('gas',  '')
    customer_filter  = request.args.get('customer', '')
    direction_filter = request.args.get('direction', '')
    try:
        from_date = date.fromisoformat(from_str)
    except Exception:
        from_date = date.today()
    try:
        to_date = date.fromisoformat(to_str)
    except Exception:
        to_date = date.today()

    movements = build_rotation(from_date=from_date, to_date=to_date,
                               gas_filter=gas_filter, customer_filter=customer_filter,
                               direction_filter=direction_filter)

    in_count       = sum(1 for m in movements if m['direction'] == 'in')
    out_count      = sum(1 for m in movements if m['direction'] == 'out')
    days_out_vals  = [m['days_out'] for m in movements if m['days_out'] is not None]
    avg_days       = round(sum(days_out_vals) / len(days_out_vals), 1) if days_out_vals else 0
    customers      = sorted(get_customer_names())

    return render_template('rotation.html',
        user=session['user'],
        movements=movements,
        in_count=in_count,
        out_count=out_count,
        avg_days=avg_days,
        net_movement=out_count - in_count,
        from_str=from_str,
        to_str=to_str,
        gas_filter=gas_filter,
        customer_filter=customer_filter,
        direction_filter=direction_filter,
        customers=customers,
    )

@app.route('/admin/rotation/journey/<path:uid>')
@admin_required
def admin_rotation_journey(uid):
    from flask import jsonify
    journey = build_cylinder_journey(uid)
    for j in journey:
        j.pop('date_obj', None)
    return jsonify(journey)

# ── Products Config Editor ─────────────────────────────────────────────────────
@app.route('/admin/products')
@admin_required
def admin_products():
    """Show the Products Config editor — lets admin edit Table 1 rows."""
    config = get_products_config()
    source = 'sheet' if products_ws is not None else 'default'
    return render_template('products_config.html',
        user=session['user'],
        products=config,
        source=source,
    )

@app.route('/admin/products/save', methods=['POST'])
@admin_required
def admin_products_save():
    """Save all product rows back to the Products Google Sheet."""
    global products_ws
    from flask import jsonify as _jsonify

    # Collect all rows from form
    ids      = request.form.getlist('pid[]')
    names    = request.form.getlist('name[]')
    gas_types= request.form.getlist('gas_type[]')
    cyl_types= request.form.getlist('cylinder_type[]')
    gas_vals = request.form.getlist('gas_per_cyl[]')
    units    = request.form.getlist('unit[]')
    virtuals = request.form.getlist('is_virtual[]')   # checkboxes — only present if checked

    # Build a set of checked virtual product IDs
    virtual_set = set(virtuals)

    rows = []
    for i, pid in enumerate(ids):
        if not pid.strip():
            continue
        is_virt = 'TRUE' if pid.strip() in virtual_set else 'FALSE'
        try:
            gval = float(gas_vals[i]) if i < len(gas_vals) else 0.0
        except ValueError:
            gval = 0.0
        rows.append([
            pid.strip(),
            names[i].strip()      if i < len(names)     else '',
            gas_types[i].strip()  if i < len(gas_types)  else '',
            cyl_types[i].strip()  if i < len(cyl_types)  else 'Standard',
            gval,
            units[i].strip()      if i < len(units)      else 'Cum',
            is_virt,
        ])

    if not rows:
        return _jsonify({'ok': False, 'msg': 'No rows to save.'}), 400

    db_written = False
    if os.environ.get('DATABASE_URL'):
        try:
            Product.query.delete()
            for r in rows:
                p_db = Product(
                    product_id=r[0],
                    name=r[1],
                    gas_type=r[2],
                    cylinder_type=r[3],
                    gas_per_cyl=r[4],
                    unit=r[5],
                    is_virtual=(r[6] == 'TRUE')
                )
                db.session.add(p_db)
            db.session.commit()
            db_written = True
            print(f"[db] Saved {len(rows)} products to database.")
        except Exception as dbe:
            db.session.rollback()
            print("[db] Error saving products in DB:", dbe)
            from sqlalchemy.exc import OperationalError, InterfaceError
            is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
            if not is_connection_error:
                return _jsonify({'ok': False, 'msg': f"Database error: {str(dbe)}"}), 500

    try:
        # Ensure Products sheet exists
        if products_ws is None and doc:
            try:
                products_ws = doc.worksheet(PRODUCTS_SHEET_NAME)
            except Exception:
                products_ws = doc.add_worksheet(
                    title=PRODUCTS_SHEET_NAME, rows=100, cols=7)

        if products_ws is None:
            if not db_written:
                return _jsonify({'ok': False, 'msg': 'Cannot connect to Google Sheets and DB offline.'}), 500
        else:
            # Clear and rewrite
            products_ws.clear()
            header = ['Product ID', 'Display Name', 'Gas Type',
                      'Cylinder Type', 'Gas Per Cylinder', 'Unit', 'Is Virtual?']
            products_ws.append_row(header, value_input_option='RAW')
            products_ws.append_rows(rows, value_input_option='RAW')

        clear_cache()
        return _jsonify({'ok': True, 'msg': f'Saved {len(rows)} products successfully.'})
    except Exception as e:
        print("admin_products_save sheets error:", e)
        if not db_written:
            return _jsonify({'ok': False, 'msg': str(e)}), 500
        clear_cache()
        return _jsonify({'ok': True, 'msg': f'Saved {len(rows)} products to DB (failed mirroring to Sheets).'})

@app.route('/admin/drivers')
@admin_required
def admin_drivers():
    data = build_driver_stats()
    return render_template('drivers.html',
        user = session['user'],
        data = data
    )

@app.route('/admin/movement')
@admin_required
def admin_movement():
    data        = build_daily_movement()
    labels      = [r['date'] for r in data]
    delivered   = [r['delivered'] for r in data]   # fixed: was 'out'
    collected   = [r['collected'] for r in data]
    return render_template('movement.html',
        user      = session['user'],
        data      = data,
        labels    = labels,
        delivered = delivered,
        collected = collected
    )

@app.route('/admin/search')
@admin_required
def admin_search():
    uid     = request.args.get('uid', '').strip()
    history = get_cylinder_history(uid) if uid else []
    current = None
    if history:
        # Current status = last delivery not yet collected
        events = build_events()
        cyl_owner = {}
        for ev in events:
            if ev['uid'].upper() == uid.upper():
                if ev['action'] == 'Delivery':
                    cyl_owner['owner'] = ev['customer']
                    cyl_owner['since'] = ev['date']
                elif ev['action'] == 'Collection':
                    cyl_owner = {}
        current = cyl_owner if cyl_owner else None

    scan_rows = get_scan_rows()
    drivers = sorted(list(set(r['driver'].strip() for r in scan_rows if r.get('driver') and r['driver'].strip())))

    return render_template('search.html',
        user    = session['user'],
        uid     = uid,
        history = history,
        current = current,
        drivers = drivers
    )
def get_customer_receipt_history(customer_name):
    """Returns receipt send history for a customer from the Customer Map table, falling back to Sheets.
    Returns list of dicts: {date, time, driver, action, count, status}
    """
    try:
        if os.environ.get('DATABASE_URL'):
            maps = CustomerMap.query.filter(CustomerMap.customer.ilike(customer_name)).all()
            if maps:
                out = []
                for m in maps:
                    status = m.receipt_status or ''
                    if not status.lower().startswith('sent'):
                        continue
                    out.append({
                        'date':   m.scan_date,
                        'time':   m.scan_time or '',
                        'driver': m.driver or '',
                        'action': m.action or '',
                        'count':  str(m.count or 0),
                        'status': status,
                    })
                out.reverse()
                return out
    except Exception as e:
        print("[db] Error reading receipt history from DB, falling back to Sheets:", e)

    if map_ws is None:
        return []
    try:
        rows = map_ws.get_all_values()
        if len(rows) < 2:
            return []
        out = []
        for r in rows[1:]:
            if len(r) < 7:
                continue
            cust = r[6].strip() if len(r) > 6 else ''
            status = r[8].strip() if len(r) > 8 else ''
            if cust.lower() != customer_name.lower():
                continue
            if not status.lower().startswith('sent'):
                continue
            out.append({
                'date':   r[0].strip(),
                'time':   r[1].strip(),
                'driver': r[2].strip(),
                'action': r[3].strip(),
                'count':  r[4].strip() if len(r) > 4 else '0',
                'status': status,
            })
        out.reverse()
        return out
    except Exception as e:
        print('Error reading receipt history from sheet:', e)
        return []

def get_customer_map_batches():
    """Returns list of all Customer Map batches from PostgreSQL, falling back to Sheets."""
    try:
        if os.environ.get('DATABASE_URL'):
            maps = CustomerMap.query.all()
            if maps:
                out = [{
                    'row_num': None,
                    'date': m.scan_date,
                    'time': m.scan_time or '',
                    'driver': m.driver or '',
                    'action': m.action or '',
                    'count': str(m.count or 0),
                    'uids': m.uids or '',
                    'customer': m.customer or '',
                    'send_receipt': 'TRUE' if m.send_receipt else 'FALSE',
                    'status': m.receipt_status or ''
                } for m in maps]
                out.reverse()
                return out
    except Exception as e:
        print("[db] Error reading map batches from DB, falling back to Sheets:", e)

    if map_ws is None:
        return []
    try:
        rows = map_ws.get_all_values()
        if len(rows) < 2:
            return []
        out = []
        for idx, r in enumerate(rows[1:]):
            # Columns: Date(0), Time(1), Driver(2), Action(3), Count(4), UIDs(5), Customer(6), Send Receipt?(7), Receipt Status(8)
            if len(r) >= 4:
                out.append({
                    'row_num': idx + 2,
                    'date': r[0].strip(),
                    'time': r[1].strip(),
                    'driver': r[2].strip(),
                    'action': r[3].strip(),
                    'count': r[4].strip() if len(r) > 4 else '0',
                    'uids': r[5].strip() if len(r) > 5 else '',
                    'customer': r[6].strip() if len(r) > 6 else '',
                    'send_receipt': r[7].strip() if len(r) > 7 else 'FALSE',
                    'status': r[8].strip() if len(r) > 8 else ''
                })
        # Show newest batches first (reverse order)
        out.reverse()
        return out
    except Exception as e:
        print("Error reading Customer Map from sheet:", e)
        return []

@app.route('/admin/receipts')
@admin_required
def admin_receipts():
    batches = get_customer_map_batches()
    customers = get_customer_names()
    emails = get_customer_emails()
    
    # Calculate counts in Python to avoid template filter errors
    sent_count = sum(1 for b in batches if str(b.get('status', '')).startswith('Sent'))
    pending_count = sum(1 for b in batches if b.get('customer') and not str(b.get('status', '')).startswith('Sent'))
    unmapped_count = sum(1 for b in batches if not b.get('customer'))
    
    return render_template('receipts.html',
        user=session['user'],
        batches=batches,
        customers=customers,
        emails=emails,
        sent_count=sent_count,
        pending_count=pending_count,
        unmapped_count=unmapped_count
    )

@app.route('/admin/update_mapping', methods=['POST'])
@admin_required
def admin_update_mapping():
    data = request.get_json() or {}
    date_val = data.get('date', '').strip()
    time_val = data.get('time', '').strip()
    driver_val = data.get('driver', '').strip()
    action_val = data.get('action', '').strip()
    customer_val = data.get('customer', '').strip()

    if not (date_val and time_val and driver_val and action_val):
        return jsonify({'status': 'Error', 'message': 'Missing batch identifiers'})

    try:
        db_written = False
        if os.environ.get('DATABASE_URL'):
            try:
                # 1. Update CustomerMap
                cmap = CustomerMap.query.filter_by(
                    scan_date=date_val,
                    scan_time=time_val,
                    driver=driver_val,
                    action=action_val
                ).first()
                if cmap:
                    cmap.customer = customer_val

                # 2. Update individual Scan records
                scans_to_update = Scan.query.filter_by(
                    scan_date=date_val,
                    scan_time=time_val,
                    driver=driver_val,
                    action=action_val
                ).all()
                for s in scans_to_update:
                    s.customer = customer_val

                # 3. If Delivery, update location of UIDs in Cylinder registry table
                if action_val == 'Delivery':
                    uids = []
                    if cmap and cmap.uids:
                        uids = [u.strip().upper() for u in cmap.uids.split(',') if u.strip()]
                    else:
                        uids = [s.cylinder_uid.upper() for s in scans_to_update]
                    
                    for uid in uids:
                        c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
                        if c_db:
                            c_db.location = customer_val

                db.session.commit()
                db_written = True
                print(f"[db] Mapped customer {customer_val} to batch, individual scans, and cylinders in DB.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error mapping customer in DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return jsonify({'status': 'Error', 'message': f"Database error: {str(dbe)}"})

        # Mirror to Google Sheets
        try:
            if map_ws is not None:
                rows = map_ws.get_all_values()
                for idx, r in enumerate(rows):
                    if idx == 0:
                        continue
                    if len(r) >= 4:
                        if (r[0].strip() == date_val and 
                            r[1].strip() == time_val and 
                            r[2].strip() == driver_val and 
                            r[3].strip() == action_val):
                            map_ws.update_cell(idx + 1, 7, customer_val) # Column G is 7
                            break
        except Exception as se:
            print("[sheets] Error mirroring mapping update to Sheets:", se)
            if not db_written:
                raise se

        clear_cache()
        return jsonify({'status': 'Success', 'message': 'Customer mapped successfully'})
    except Exception as e:
        return jsonify({'status': 'Error', 'message': str(e)})

@app.route('/admin/send_receipt', methods=['POST'])
@admin_required
def admin_send_receipt():
    data = request.get_json() or {}
    date_val = data.get('date', '').strip()
    time_val = data.get('time', '').strip()
    driver_val = data.get('driver', '').strip()
    action_val = data.get('action', '').strip()

    if not (date_val and time_val and driver_val and action_val):
        return jsonify({'status': 'Error', 'message': 'Missing batch identifiers'})

    try:
        # Determine customer name from DB first
        customer_name = ""
        db_record = None
        if os.environ.get('DATABASE_URL'):
            try:
                db_record = CustomerMap.query.filter_by(
                    scan_date=date_val,
                    scan_time=time_val,
                    driver=driver_val,
                    action=action_val
                ).first()
                if db_record:
                    customer_name = db_record.customer or ""
            except Exception as dbe:
                print("[db] Error looking up customer map for receipt in DB:", dbe)

        # If not found in DB or DB offline, read from Sheets
        if not customer_name:
            if map_ws is not None:
                rows = map_ws.get_all_values()
                for idx, r in enumerate(rows):
                    if idx == 0:
                        continue
                    if len(r) >= 7:
                        if (r[0].strip() == date_val and 
                            r[1].strip() == time_val and 
                            r[2].strip() == driver_val and 
                            r[3].strip() == action_val):
                            customer_name = r[6].strip()
                            break

        if not customer_name:
            return jsonify({'status': 'Error', 'message': 'Please map a customer first'})

        # Validate if email exists for this customer
        emails = get_customer_emails()
        email = emails.get(customer_name, '').strip()
        if not email or '@' not in email:
            return jsonify({'status': 'Error', 'message': f"Missing Email: Please enter an email address for {customer_name} in the Customers page first."})

        # Database update first
        db_written = False
        if os.environ.get('DATABASE_URL'):
            try:
                if not db_record:
                    db_record = CustomerMap.query.filter_by(
                        scan_date=date_val,
                        scan_time=time_val,
                        driver=driver_val,
                        action=action_val
                    ).first()
                if db_record:
                    db_record.send_receipt = True
                    db_record.receipt_status = "Sending..."
                    db.session.commit()
                    db_written = True
                    print("[db] Triggered send receipt in database.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error triggering send receipt in DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return jsonify({'status': 'Error', 'message': f"Database error: {str(dbe)}"})

        # Sheets update
        try:
            if map_ws is not None:
                rows = map_ws.get_all_values()
                for idx, r in enumerate(rows):
                    if idx == 0:
                        continue
                    if len(r) >= 4:
                        if (r[0].strip() == date_val and 
                            r[1].strip() == time_val and 
                            r[2].strip() == driver_val and 
                            r[3].strip() == action_val):
                            map_ws.update(f"H{idx + 1}:I{idx + 1}", [[True, "Sending..."]], value_input_option='USER_ENTERED')
                            break
        except Exception as se:
            print("[sheets] Error mirroring send receipt trigger to Sheets:", se)
            if not db_written:
                raise se

        clear_cache()
        return jsonify({'status': 'Success', 'message': 'Receipt trigger sent'})
    except Exception as e:
        return jsonify({'status': 'Error', 'message': str(e)})

@app.route('/admin/receipt_status')
@admin_required
def admin_receipt_status():
    date_val = request.args.get('date', '').strip()
    time_val = request.args.get('time', '').strip()
    driver_val = request.args.get('driver', '').strip()
    action_val = request.args.get('action', '').strip()

    if not (date_val and time_val and driver_val and action_val):
        return jsonify({'status': 'Error', 'message': 'Missing batch identifiers'})

    try:
        db_record = None
        # 1. Fast path: check DB first
        if os.environ.get('DATABASE_URL'):
            try:
                db_record = CustomerMap.query.filter_by(
                    scan_date=date_val,
                    scan_time=time_val,
                    driver=driver_val,
                    action=action_val
                ).first()
                if db_record:
                    db_status = db_record.receipt_status or ''
                    # If DB already has a final status, return immediately
                    if db_status and db_status != 'Sending...' and not db_status.startswith('Error:'):
                        return jsonify({
                            'status': 'Success',
                            'send_receipt': 'TRUE' if db_record.send_receipt else 'FALSE',
                            'receipt_status': db_status
                        })
            except Exception as dbe:
                print("[db] Error reading receipt status from DB:", dbe)

        # 2. Slow path: read from Sheets (Apps Script writes here directly)
        if map_ws is not None:
            try:
                rows = map_ws.get_all_values()
                for idx, r in enumerate(rows[1:]):
                    if len(r) >= 4:
                        if (r[0].strip() == date_val and
                            r[1].strip() == time_val and
                            r[2].strip() == driver_val and
                            r[3].strip() == action_val):
                            send_receipt = r[7].strip() if len(r) > 7 else 'FALSE'
                            sheet_status = r[8].strip() if len(r) > 8 else ''

                            # Mirror final status back to DB immediately so future polls are fast
                            if sheet_status and sheet_status != 'Sending...' and db_record:
                                try:
                                    db_record.receipt_status = sheet_status
                                    db_record.send_receipt = send_receipt.upper() in ('TRUE', '1', 'YES')
                                    db.session.commit()
                                    print(f"[db] Mirrored receipt status from Sheets to DB: {sheet_status}")
                                except Exception as dbe2:
                                    db.session.rollback()
                                    print("[db] Error mirroring receipt status to DB:", dbe2)

                            return jsonify({
                                'status': 'Success',
                                'send_receipt': send_receipt,
                                'receipt_status': sheet_status
                            })
            except Exception as se:
                print("[sheets] Error reading receipt status from Sheets:", se)

        # 3. If DB had a Sending... record but Sheets is unavailable, return DB value
        if db_record:
            return jsonify({
                'status': 'Success',
                'send_receipt': 'TRUE' if db_record.send_receipt else 'FALSE',
                'receipt_status': db_record.receipt_status or ''
            })

        return jsonify({'status': 'Error', 'message': 'Batch row not found'})
    except Exception as e:
        return jsonify({'status': 'Error', 'message': str(e)})


# ── Inventory & Bulk Tanks Calculations ─────────────────────────

def parse_mix_ratio(mix_ratio_str, gas_type):
    """Parses mix ratio percentages from a string."""
    if not mix_ratio_str:
        if gas_type == 'ACM':
            return {'Argon': 0.9073, 'CO2': 0.1928}  # default ACM 90/10 ratio
        elif gas_type == 'AHM':
            return {'Argon': 0.9886}  # default AHM
        return {}
    
    mix_ratio_str = str(mix_ratio_str).upper()
    arg_match = re.search(r'(\d+)%\s*(?:ARG|ARGON)', mix_ratio_str)
    co2_match = re.search(r'(\d+)%\s*(?:CO2|CARBON)', mix_ratio_str)
    n2_match = re.search(r'(\d+)%\s*(?:N2|NITROGEN)', mix_ratio_str)
    oxy_match = re.search(r'(\d+)%\s*(?:O2|OXY|OXYGEN)', mix_ratio_str)
    
    parts = {}
    if arg_match: parts['Argon'] = float(arg_match.group(1)) / 100.0
    if co2_match: parts['CO2'] = float(co2_match.group(1)) / 100.0
    if n2_match: parts['N2'] = float(n2_match.group(1)) / 100.0
    if oxy_match: parts['Oxygen'] = float(oxy_match.group(1)) / 100.0
    
    if not parts:
        slash_match = re.search(r'(\d+)\s*/\s*(\d+)', mix_ratio_str)
        if slash_match:
            val1 = float(slash_match.group(1))
            val2 = float(slash_match.group(2))
            if gas_type == 'ACM':
                if val1 == 90 and val2 == 10:
                    return {'Argon': 0.9073, 'CO2': 0.1928}
                elif val1 == 80 and val2 == 20:
                    return {'Argon': 0.80, 'CO2': 0.20}
                else:
                    total = val1 + val2
                    return {'Argon': val1 / total, 'CO2': val2 / total}
            elif gas_type == 'AHM':
                total = val1 + val2
                return {'Argon': val1 / total}
                
    if not parts:
        if gas_type == 'ACM':
            return {'Argon': 0.9073, 'CO2': 0.1928}
        elif gas_type == 'AHM':
            return {'Argon': 0.9886}
            
    return parts

def get_tank_data_for_date(target_date_str):
    """Fetches opening stock, dead volume, capacity, and unit for bulk tanks."""
    global bulk_tanks_ws
    defaults = {
        'Argon': {'opening': 5664.03, 'dead_volume': 500.0, 'capacity': 10000.0, 'unit': 'Cum'},
        'CO2': {'opening': 10941.05, 'dead_volume': 200.0, 'capacity': 15000.0, 'unit': 'KG'},
        'N2': {'opening': 4271.20, 'dead_volume': 300.0, 'capacity': 8000.0, 'unit': 'Cum'},
        'Oxygen': {'opening': 9215.30, 'dead_volume': 400.0, 'capacity': 12000.0, 'unit': 'Cum'}
    }
    try:
        records = []
        db_loaded = False
        if os.environ.get('DATABASE_URL'):
            try:
                tanks = BulkTank.query.all()
                if tanks:
                    for t in tanks:
                        records.append({
                            'date': t.date,
                            'date_obj': parse_date(t.date),
                            'gas': t.gas,
                            'opening': t.opening,
                            'dead_volume': t.dead_volume,
                            'capacity': t.capacity,
                            'unit': t.unit
                        })
                    db_loaded = True
            except Exception as dbe:
                print("[db] Error reading bulk tanks from DB, falling back to Sheets:", dbe)

        if not db_loaded:
            if bulk_tanks_ws is None and doc:
                try: bulk_tanks_ws = doc.worksheet(BULK_TANKS_NAME)
                except Exception: pass
            if bulk_tanks_ws is None:
                return defaults
            
            rows = bulk_tanks_ws.get_all_values()
            if len(rows) < 2:
                return defaults
                
            for r in rows[1:]:
                if len(r) >= 6 and r[0].strip():
                    try:
                        records.append({
                            'date': r[0].strip(),
                            'date_obj': parse_date(r[0]),
                            'gas': r[1].strip(),
                            'opening': float(r[2].strip() or 0.0),
                            'dead_volume': float(r[3].strip() or 0.0),
                            'capacity': float(r[4].strip() or 0.0),
                            'unit': r[5].strip()
                        })
                    except ValueError:
                        continue
                    
        target_records = [rec for rec in records if rec['date'] == target_date_str]
        if target_records:
            out = {}
            for rec in target_records:
                out[rec['gas']] = {
                    'opening': rec['opening'],
                    'dead_volume': rec['dead_volume'],
                    'capacity': rec['capacity'],
                    'unit': rec['unit']
                }
            for g in defaults:
                if g not in out:
                    out[g] = defaults[g]
            return out
            
        # Rollover closing stock of previous date
        sorted_records = sorted([r for r in records if r['date_obj'] is not None], key=lambda x: x['date_obj'])
        if not sorted_records:
            return defaults
            
        target_date_obj = parse_date(target_date_str)
        prev_records = []
        if target_date_obj:
            prev_records = [r for r in sorted_records if r['date_obj'] < target_date_obj]
            
        if prev_records:
            latest_prev_date = prev_records[-1]['date']
            prev_date_records = [r for r in prev_records if r['date'] == latest_prev_date]
        else:
            latest_date = sorted_records[-1]['date']
            prev_date_records = [r for r in sorted_records if r['date'] == latest_date]
            latest_prev_date = latest_date
            
        prev_used = calculate_used_today(latest_prev_date)
        out = {}
        for rec in prev_date_records:
            used = prev_used.get(rec['gas'], 0.0)
            closing = max(0.0, rec['opening'] - used)
            out[rec['gas']] = {
                'opening': round(closing, 2),
                'dead_volume': rec['dead_volume'],
                'capacity': rec['capacity'],
                'unit': rec['unit']
            }
        for g in defaults:
            if g not in out:
                out[g] = defaults[g]
        return out
        
    except Exception as e:
        print("Error fetching bulk tank data:", e)
        return defaults

def calculate_used_today(target_date_str):
    """Calculates used gas today from fillings logged in scan log."""
    used = {'Argon': 0.0, 'CO2': 0.0, 'N2': 0.0, 'Oxygen': 0.0}
    try:
        fillings = []
        db_loaded = False
        if os.environ.get('DATABASE_URL'):
            try:
                scans = Scan.query.filter_by(scan_date=target_date_str, action='Filling').all()
                fillings = [s.cylinder_uid.strip().upper() for s in scans if s.cylinder_uid]
                db_loaded = True
            except Exception as dbe:
                print("[db] Error querying scans for calculate_used_today, falling back to Sheets:", dbe)

        if not db_loaded:
            if scan_ws is None and doc:
                try: scan_ws = doc.worksheet(SCAN_SHEET_NAME)
                except Exception: pass
            if scan_ws is None:
                return used
                
            rows = scan_ws.get_all_values()
            if len(rows) < 2:
                return used
                
            for r in rows[1:]:
                if len(r) >= 5:
                    if r[0].strip() == target_date_str and r[3].strip() == 'Filling':
                        fillings.append(r[4].strip().upper())
                        
        if not fillings:
            return used
            
        cylinders = get_all_cylinders()
        maintenance = get_all_maintenance()
        cyl_map = {c['uid'].upper(): c for c in cylinders}
        
        for uid in fillings:
            cyl = cyl_map.get(uid)
            if cyl:
                gas_type = cyl['gas_type'].upper()
                cyl_type = cyl['cylinder_type'].capitalize()
            else:
                cyl_type = 'Standard'
                if uid.startswith('ARG'): gas_type = 'ARG'
                elif uid.startswith('CO2'): gas_type = 'CO2'
                elif uid.startswith('N2'): gas_type = 'N2'
                elif uid.startswith('OXY'): gas_type = 'OXY'
                elif uid.startswith('ACM'): gas_type = 'ACM'
                else: continue
                
            maint = maintenance.get(uid, {})
            try:
                capacity = float(maint.get('gas_capacity') or 0.0) if maint.get('gas_capacity') else None
            except ValueError:
                capacity = None
                
            mix_ratio = maint.get('mix_ratio', '')
            
            if capacity is None:
                if gas_type == 'ARG':
                    capacity = 7.0 if cyl_type == 'Standard' else 0.88
                elif gas_type == 'CO2':
                    capacity = 30.0
                elif gas_type == 'N2':
                    capacity = 7.0 if cyl_type == 'Standard' else 0.88
                elif gas_type == 'OXY':
                    capacity = 7.0 if cyl_type == 'Standard' else 0.88
                elif gas_type == 'ACM':
                    capacity = 7.0
                elif gas_type == 'AHM':
                    capacity = 7.0
                else:
                    capacity = 0.0
                    
            if gas_type == 'ARG':
                used['Argon'] += capacity
            elif gas_type == 'CO2':
                used['CO2'] += capacity
            elif gas_type == 'N2':
                used['N2'] += capacity
            elif gas_type == 'OXY':
                used['Oxygen'] += capacity
            elif gas_type == 'ACM':
                ratios = parse_mix_ratio(mix_ratio, 'ACM')
                used['Argon'] += capacity * ratios.get('Argon', 0.80)
                used['CO2'] += capacity * ratios.get('CO2', 0.20)
            elif gas_type == 'AHM':
                ratios = parse_mix_ratio(mix_ratio, 'AHM')
                used['Argon'] += capacity * ratios.get('Argon', 0.92)
                
        for k in used:
            used[k] = round(used[k], 2)
        return used
    except Exception as e:
        print("Error calculating used today:", e)
        return used

def calculate_fleet_sizes():
    """Calculates fleet size (total cylinder count) by gas type."""
    cylinders = get_all_cylinders()
    fleet = {'Argon': 0, 'CO2': 0, 'N2': 0, 'Oxygen': 0}
    for c in cylinders:
        gas = c.get('gas_type', '').upper()
        if gas in ('ARG', 'ACM', 'AHM'):
            fleet['Argon'] += 1
        elif gas == 'CO2':
            fleet['CO2'] += 1
        elif gas in ('N2', 'N2D'):
            fleet['N2'] += 1
        elif gas == 'OXY':
            fleet['Oxygen'] += 1
    return fleet

def calculate_table1_filled_inventory():
    """Calculates Table 1 — Filled Cylinder Inventory."""
    cylinders = get_all_cylinders()
    maintenance = get_all_maintenance()
    
    products_config = get_products_config()
    
    t1_rows = {p['id']: {**p, 'filled_count': 0, 'total_gas': 0.0} for p in products_config}
    
    for c in cylinders:
        if c.get('status') != 'Filled':
            continue
            
        uid = c.get('uid', '').strip().upper()
        gas_type = c.get('gas_type', '').upper()
        cyl_type = c.get('cylinder_type', '').capitalize()
        maint = maintenance.get(uid, {})
        
        try:
            capacity = float(maint.get('gas_capacity') or 0.0) if maint.get('gas_capacity') else None
        except ValueError:
            capacity = None
            
        mix_ratio = maint.get('mix_ratio', '')
        
        pid = None
        if gas_type == 'ARG':
            pid = 'arg_pura' if cyl_type == 'Standard' else 'arg_dura'
        elif gas_type == 'CO2':
            pid = 'co2_pure'
        elif gas_type == 'N2':
            pid = 'n2_cyl' if cyl_type == 'Standard' else 'n2_dura'
        elif gas_type == 'OXY':
            pid = 'oxygen_pure' if cyl_type == 'Standard' else 'oxygen_dura'
        elif gas_type == 'ACM':
            pid = 'acm_90_10'
        elif gas_type == 'AHM':
            if '98' in mix_ratio:
                pid = 'ahm_98_02'
            else:
                pid = 'ahm_92_08'
                
        if pid:
            if pid == 'acm_90_10':
                t1_rows['acm_90_10']['filled_count'] += 1
                t1_rows['co2_90_10']['filled_count'] += 1
                cap_val = capacity if capacity is not None else 7.0
                t1_rows['acm_90_10']['total_gas'] += cap_val * 0.9073
                t1_rows['co2_90_10']['total_gas'] += cap_val * 0.1928
            else:
                t1_rows[pid]['filled_count'] += 1
                cap_val = capacity if capacity is not None else t1_rows[pid]['gas_per_cyl']
                t1_rows[pid]['total_gas'] += cap_val
                
    results = []
    total_physical_filled = 0
    total_cum = 0.0
    total_kg = 0.0
    
    for p in products_config:
        res = t1_rows[p['id']]
        res['total_gas'] = round(res['total_gas'], 3)
        results.append(res)
        
        if not res.get('is_virtual'):
            total_physical_filled += res['filled_count']
        if res['unit'] == 'Cum':
            total_cum += res['total_gas']
        elif res['unit'] == 'KG':
            total_kg += res['total_gas']
            
    return {
        'rows': results,
        'total_filled': total_physical_filled,
        'total_cum': round(total_cum, 3),
        'total_kg': round(total_kg, 3)
    }

def calculate_table2_bulk_inventory(target_date_str):
    """Calculates Table 2 — Bulk Tank Inventory."""
    tanks = get_tank_data_for_date(target_date_str)
    used = calculate_used_today(target_date_str)
    fleet = calculate_fleet_sizes()
    
    gases = [
        {'id': 'Argon', 'name': 'Argon'},
        {'id': 'CO2', 'name': 'CO2'},
        {'id': 'N2', 'name': 'N2'},
        {'id': 'Oxygen', 'name': 'Oxygen'}
    ]
    
    rows = []
    for g in gases:
        tank_data = tanks.get(g['id'], {'opening': 0.0, 'dead_volume': 0.0, 'capacity': 10000.0, 'unit': 'Cum'})
        used_today = used.get(g['id'], 0.0)
        closing = round(max(0.0, tank_data['opening'] - used_today), 2)
        usable = round(max(0.0, tank_data['opening'] - tank_data['dead_volume']), 2)
        
        rows.append({
            'gas': g['name'],
            'opening': tank_data['opening'],
            'used_today': used_today,
            'closing': closing,
            'usable': usable,
            'dead_volume': tank_data['dead_volume'],
            'capacity': tank_data['capacity'],
            'fleet': fleet.get(g['id'], 0),
            'unit': tank_data['unit']
        })
        
    return rows

def update_tank_opening_stock(date_str, gas_name, opening, capacity, dead_volume, unit):
    """Updates opening stock, capacity, and dead volume for a specific gas and date in database first, then mirror to Sheets."""
    global bulk_tanks_ws
    try:
        db_written = False
        if os.environ.get('DATABASE_URL'):
            try:
                bt = BulkTank.query.filter_by(date=date_str, gas=gas_name).first()
                if bt:
                    bt.opening = float(opening)
                    bt.dead_volume = float(dead_volume)
                    bt.capacity = float(capacity)
                    bt.unit = unit
                else:
                    bt = BulkTank(
                        date=date_str,
                        gas=gas_name,
                        opening=float(opening),
                        dead_volume=float(dead_volume),
                        capacity=float(capacity),
                        unit=unit
                    )
                    db.session.add(bt)
                db.session.commit()
                db_written = True
                print(f"[db] Updated bulk tank stock for {gas_name} on {date_str} in DB.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error updating bulk tank stock in DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return False

        # Mirror to Google Sheets
        try:
            if bulk_tanks_ws is None and doc:
                try: bulk_tanks_ws = doc.worksheet(BULK_TANKS_NAME)
                except Exception: pass
            
            if bulk_tanks_ws is not None:
                rows = bulk_tanks_ws.get_all_values()
                row_num = None
                for idx, r in enumerate(rows):
                    if idx == 0: continue
                    if len(r) >= 2 and r[0].strip() == date_str and r[1].strip() == gas_name:
                        row_num = idx + 1
                        break
                        
                row_data = [date_str, gas_name, float(opening), float(dead_volume), float(capacity), unit]
                if row_num:
                    bulk_tanks_ws.update(f'A{row_num}:F{row_num}', [row_data])
                else:
                    bulk_tanks_ws.append_row(row_data)
        except Exception as se:
            print("[sheets] Error mirroring bulk tank update to Sheets:", se)
            if not db_written:
                raise se

        return True
    except Exception as e:
        print("Error updating tank opening stock:", e)
        return False

def calculate_daily_dispatch_report(target_date_str):
    """Calculates customer-wise dispatches (deliveries) and collections on target_date_str"""
    global scan_ws
    empty_report = {
        'company_rows': [],
        'party_rows': [],
        'company_totals': {
            'dispatch': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
            'collection': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
            'dispatch_total': 0,
            'collection_total': 0
        },
        'party_totals': {
            'dispatch': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
            'collection': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
            'dispatch_total': 0,
            'collection_total': 0
        },
        'grand_totals': {
            'dispatch': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
            'collection': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
            'dispatch_total': 0,
            'collection_total': 0
        }
    }
    try:
        day_scans = []
        db_loaded = False
        if os.environ.get('DATABASE_URL'):
            try:
                scans = Scan.query.filter(
                    Scan.scan_date == target_date_str,
                    Scan.action.in_(['Delivery', 'Collection'])
                ).all()
                for s in scans:
                    day_scans.append({
                        'action': s.action,
                        'uid': s.cylinder_uid.strip().upper(),
                        'customer': s.customer.strip() if s.customer else ''
                    })
                db_loaded = True
            except Exception as dbe:
                print("[db] Error in calculate_daily_dispatch_report query, falling back to Sheets:", dbe)

        if not db_loaded:
            if scan_ws is None and doc:
                try: scan_ws = doc.worksheet(SCAN_SHEET_NAME)
                except Exception: pass
            if scan_ws is None:
                return empty_report
                
            rows = scan_ws.get_all_values()
            if len(rows) < 2:
                return empty_report
                
            for r in rows[1:]:
                if len(r) >= 6:
                    if r[0].strip() == target_date_str and r[3].strip() in ('Delivery', 'Collection'):
                        day_scans.append({
                            'action': r[3].strip(),
                            'uid': r[4].strip().upper(),
                            'customer': r[5].strip()
                        })
                    
        if not day_scans:
            return empty_report
            
        cylinders = get_all_cylinders()
        maint_data = get_all_maintenance()
        cyl_map = {c['uid'].upper(): c for c in cylinders}
        
        def make_empty_row():
            return {
                'ACM': 0, 'ARG': 0, 'CO2': 0, 'N2': 0, 'Oxy': 0, 'Helium': 0, 'DA': 0,
                'Dura': {'count': 0, 'gases': {}}
            }
            
        company_customers = {}
        party_customers = {}
        
        for s in day_scans:
            uid = s['uid']
            action = s['action'].lower()
            customer = s['customer'] or '(No Customer)'
            
            cyl = cyl_map.get(uid)
            
            owner = cyl.get('owner', '').strip() if cyl else ''
            if not owner or owner.upper() in ('COMPANY', 'DEPOT'):
                is_company_owned = True
            else:
                is_company_owned = False
                
            if cyl:
                gas_type = cyl['gas_type'].upper()
                cyl_type = cyl['cylinder_type'].capitalize()
            else:
                cyl_type = 'Standard'
                if 'DURA' in uid:
                    cyl_type = 'Dura'
                if uid.startswith('ARG'): gas_type = 'ARG'
                elif uid.startswith('CO2'): gas_type = 'CO2'
                elif uid.startswith('N2'): gas_type = 'N2'
                elif uid.startswith('OXY'): gas_type = 'OXY'
                elif uid.startswith('ACM'): gas_type = 'ACM'
                elif uid.startswith('HEL'): gas_type = 'HEL'
                elif uid.startswith('DA'): gas_type = 'DA'
                else: gas_type = 'OXY'
                
            col_key = None
            if cyl_type == 'Dura':
                col_key = 'Dura'
            else:
                if gas_type == 'ACM': col_key = 'ACM'
                elif gas_type == 'ARG': col_key = 'ARG'
                elif gas_type == 'CO2': col_key = 'CO2'
                elif gas_type in ('N2', 'N2D'): col_key = 'N2'
                elif gas_type == 'OXY': col_key = 'Oxy'
                elif 'HEL' in gas_type: col_key = 'Helium'
                elif gas_type == 'DA': col_key = 'DA'
                else: col_key = 'Oxy'
                
            group = company_customers if is_company_owned else party_customers
            
            if customer not in group:
                group[customer] = {
                    'dispatch': make_empty_row(),
                    'collection': make_empty_row()
                }
                
            act_key = 'dispatch' if action == 'delivery' else 'collection'
            row_dict = group[customer][act_key]
            
            if col_key == 'Dura':
                row_dict['Dura']['count'] += 1
                gas_symbol = 'Ar'
                if gas_type == 'N2': gas_symbol = 'N²'
                elif gas_type == 'OXY': gas_symbol = 'O²'
                elif gas_type == 'CO2': gas_symbol = 'CO²'
                
                row_dict['Dura']['gases'][gas_symbol] = row_dict['Dura']['gases'].get(gas_symbol, 0) + 1
            else:
                row_dict[col_key] += 1
                
        def format_dura(dura_dict):
            if dura_dict['count'] == 0:
                return ''
            parts = []
            for gas, cnt in sorted(dura_dict['gases'].items()):
                parts.append(f"{cnt} {gas}")
            if not parts:
                return str(dura_dict['count'])
            return " / ".join(parts)
            
        def convert_to_list(group_dict):
            out = []
            for cust, data in sorted(group_dict.items()):
                formatted_row = {
                    'customer': cust,
                    'dispatch': {k: (v if k != 'Dura' else format_dura(v)) for k, v in data['dispatch'].items()},
                    'collection': {k: (v if k != 'Dura' else format_dura(v)) for k, v in data['collection'].items()},
                    'raw_dispatch': {k: (v if k != 'Dura' else v['count']) for k, v in data['dispatch'].items()},
                    'raw_collection': {k: (v if k != 'Dura' else v['count']) for k, v in data['collection'].items()}
                }
                out.append(formatted_row)
            return out
            
        company_rows = convert_to_list(company_customers)
        party_rows = convert_to_list(party_customers)
        
        def calc_totals(rows_list):
            tot = {
                'dispatch': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
                'collection': {k: 0 for k in ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']},
                'dispatch_total': 0,
                'collection_total': 0
            }
            for r in rows_list:
                for k in tot['dispatch']:
                    tot['dispatch'][k] += r['raw_dispatch'][k]
                    tot['collection'][k] += r['raw_collection'][k]
            tot['dispatch_total'] = sum(tot['dispatch'].values())
            tot['collection_total'] = sum(tot['collection'].values())
            return tot
            
        company_totals = calc_totals(company_rows)
        party_totals = calc_totals(party_rows)
        
        grand_totals = {
            'dispatch': {k: company_totals['dispatch'][k] + party_totals['dispatch'][k] for k in company_totals['dispatch']},
            'collection': {k: company_totals['collection'][k] + party_totals['collection'][k] for k in company_totals['collection']},
            'dispatch_total': company_totals['dispatch_total'] + party_totals['dispatch_total'],
            'collection_total': company_totals['collection_total'] + party_totals['collection_total']
        }
        
        return {
            'company_rows': company_rows,
            'party_rows': party_rows,
            'company_totals': company_totals,
            'party_totals': party_totals,
            'grand_totals': grand_totals
        }
    except Exception as e:
        print("Error calculating daily dispatch report:", e)
        return empty_report



# ================================================================
#  CYLINDER REGISTRY ROUTES
# ================================================================


@app.route('/admin/cylinders')
@admin_required
def admin_cylinders():
    cylinders = merge_cylinder_data()
    products = get_products_config()
    gas_types = sorted(list(set(c['gas_type'] for c in cylinders if c.get('gas_type'))))
    statuses  = ['Empty', 'Filled', 'Delivered']
    return render_template('cylinders.html',
        user       = session['user'],
        cylinders  = cylinders,
        gas_types  = gas_types,
        statuses   = statuses,
        total      = len(cylinders),
        active     = sum(1 for c in cylinders if c.get('status') in ('Active', 'Empty', 'Filled', 'Delivered')),
        overdue    = sum(1 for c in cylinders if c.get('hydro_badge') == 'Overdue'),
        due_soon   = sum(1 for c in cylinders if c.get('hydro_badge') == 'Due Soon'),
    )

@app.route('/admin/cylinders/sync_sheets', methods=['POST'])
@admin_required
def admin_cylinders_sync_sheets():
    def run_sync():
        try:
            from app import doc, CYLINDER_SHEET_NAME, CYLINDER_MAINT_NAME, USERS_SHEET_NAME
            if not doc:
                return
            cylinders = Cylinder.query.all()
            maints = {m.cylinder_uid: m for m in CylinderMaintenance.query.all()}
            
            cyl_rows = []
            maint_rows = []
            
            for c in cylinders:
                cyl_rows.append([
                    c.uid, c.gas_type, c.cylinder_type, c.owner, c.status, c.location, c.last_activity_date
                ])
                m = maints.get(c.uid)
                if m:
                    maint_rows.append([
                        c.uid, m.water_capacity, m.fill_pressure, m.gas_capacity, m.unit, m.is_mixture, '', 
                        m.manufacture_date, m.last_hydro_date, m.next_hydro_due, m.hydro_test_status, m.cert_no, m.is_uhp
                    ])

            try:
                cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                cyl_ws.resize(1)
                if cyl_rows:
                    cyl_ws.append_rows(cyl_rows)
            except Exception as e:
                print("Error with Cylinders sheet:", e)
                
            try:
                maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
                maint_ws.resize(1)
                if maint_rows:
                    maint_ws.append_rows(maint_rows)
            except Exception as e:
                print("Error with Maintenance sheet:", e)

            try:
                users_ws = doc.worksheet(USERS_SHEET_NAME)
                users_ws.resize(1)
                users = User.query.all()
                user_rows = []
                for u in users:
                    user_rows.append([
                        u.username, u.password, u.role, u.name or u.username
                    ])
                if user_rows:
                    users_ws.append_rows(user_rows)
            except Exception as e:
                print("Error with Users sheet sync:", e)
                
        except Exception as e:
            print("[sheets] Error forcing sync:", e)

    async_sheets_write(run_sync)
    flash("Background sync to Google Sheets started! Please wait 15-30 seconds for it to fully populate.", "info")
    return redirect('/admin/cylinders')

@app.route('/admin/cylinders/bulk_delete', methods=['POST'])
@admin_required
def admin_cylinders_bulk_delete():
    data = request.get_json()
    if not data or 'uids' not in data:
        return "Invalid request format", 400
        
    uids_to_delete = [str(u).strip() for u in data['uids'] if str(u).strip()]
    if not uids_to_delete:
        return "No cylinders selected", 400
        
    global cyl_ws, cyl_maint_ws, doc
    
    try:
        # Delete from PostgreSQL
        if os.environ.get('DATABASE_URL'):
            try:
                CylinderMaintenance.query.filter(CylinderMaintenance.cylinder_uid.in_(uids_to_delete)).delete(synchronize_session=False)
                Cylinder.query.filter(Cylinder.uid.in_(uids_to_delete)).delete(synchronize_session=False)
                db.session.commit()
                print(f"[db] Bulk deleted {len(uids_to_delete)} cylinders from PostgreSQL.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error in bulk delete:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return f"Database error: {str(dbe)}", 500

        # Delete from Google Sheets (Background Sync)
        def background_bulk_delete_sheets():
            try:
                global cyl_ws, cyl_maint_ws, doc
                if cyl_ws is None and doc:
                    try: cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                    except Exception: pass
                if cyl_maint_ws is None and doc:
                    try: cyl_maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
                    except Exception: pass

                uids_lower = set([u.lower() for u in uids_to_delete])

                def bulk_delete_sheet(ws):
                    if not ws: return
                    rows = ws.get_all_values()
                    rows_to_del = []
                    for idx, r in enumerate(rows):
                        if idx == 0: continue
                        if len(r) > 0 and r[0].strip().lower() in uids_lower:
                            rows_to_del.append(idx + 1)
                    
                    if not rows_to_del: return
                    
                    rows_to_del.sort()
                    ranges = []
                    start = rows_to_del[0]
                    end = rows_to_del[0]
                    for r_idx in rows_to_del[1:]:
                        if r_idx == end + 1:
                            end = r_idx
                        else:
                            ranges.append((start, end))
                            start = r_idx
                            end = r_idx
                    ranges.append((start, end))
                    
                    # Sort ranges descending so deleting doesn't shift later ranges
                    ranges.sort(key=lambda x: x[0], reverse=True)
                    
                    requests = []
                    for r_start, r_end in ranges:
                        requests.append({
                            "deleteDimension": {
                                "range": {
                                    "sheetId": ws.id,
                                    "dimension": "ROWS",
                                    "startIndex": r_start - 1,
                                    "endIndex": r_end
                                }
                            }
                        })
                    try:
                        ws.spreadsheet.batch_update({"requests": requests})
                    except Exception as e:
                        print(f"[sheets] batch_update error on {ws.title}: {e}")

                bulk_delete_sheet(cyl_ws)
                bulk_delete_sheet(cyl_maint_ws)

            except Exception as se:
                print("[sheets] Error mirroring bulk delete to Sheets:", se)

        async_sheets_write(background_bulk_delete_sheets)

        flash(f"Successfully deleted {len(uids_to_delete)} cylinders.", "success")
        return "OK", 200
        
    except Exception as e:
        return str(e), 500

@app.route('/admin/cylinders/upload', methods=['POST'])
@admin_required
def admin_cylinders_upload():
    import pandas as pd
    from datetime import date
    from flask import flash
    global cyl_ws, cyl_maint_ws, doc
    
    if 'file' not in request.files:
        flash("No file provided.", "danger")
        return redirect('/admin/cylinders')
        
    file = request.files['file']
    if file.filename == '':
        flash("No file selected.", "danger")
        return redirect('/admin/cylinders')
        
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file)
        else:
            flash("Invalid file format. Please upload a CSV or Excel file.", "danger")
            return redirect('/admin/cylinders')
            
        # Standardize column names to handle whitespace/case
        df.columns = df.columns.str.strip()
        
        # Check required columns
        if 'Cylinder ID' not in df.columns:
            flash("The file must contain a 'Cylinder ID' column.", "danger")
            return redirect('/admin/cylinders')
            
        added = 0
        skipped = 0
        
        # Query existing uids to skip duplicates efficiently
        existing_uids = set(c.uid for c in Cylinder.query.all())
        
        cyl_rows_to_append = []
        cyl_maint_rows_to_append = []
        today_str = date.today().strftime('%d-%m-%Y')
        
        for _, row in df.iterrows():
            uid = str(row.get('Cylinder ID', '')).strip()
            
            if pd.isna(uid) or uid == '' or uid == 'nan':
                continue
                
            if uid in existing_uids:
                skipped += 1
                continue
                
            # Gas Type Mapping
            raw_gas = str(row.get('Gas Type', '')).strip() if 'Gas Type' in df.columns and not pd.isna(row.get('Gas Type')) else ''
            gas_map = {
                'ARGON': 'ARG', 'OXYGEN': 'OXY', 'NITROGEN': 'N2', 
                'CARBON DIOXIDE': 'CO2', 'HELIUM': 'Helium', 'ACETYLENE': 'DA'
            }
            gas_type = gas_map.get(raw_gas.upper(), raw_gas)
            
            # Helper to safely extract columns
            def get_col(*col_names, default=''):
                for c in col_names:
                    if c in df.columns:
                        val = row.get(c)
                        if not pd.isna(val):
                            return str(val).strip()
                return default

            water_cap = get_col('Water Capacity', default='')
            raw_cyl_type = get_col('Cylinder Type', default='')

            cyl_type = raw_cyl_type
            if not cyl_type:
                if '46.7' in water_cap:
                    cyl_type = 'Standard'
                else:
                    cyl_type = 'Standard' # Default to Standard if empty

            if 'DURA' in uid.upper() and cyl_type.upper() != 'DURA':
                cyl_type = 'Dura'
                
            owner = get_col('Owner', default='Depot')
            fill_pressure = get_col('Fill Pressure', 'Fill Pressure (bar)', default='')
            gas_capacity = get_col('Gas Capacity', default='')
            unit = get_col('Unit', default='')
            is_mixture = get_col('Is Mixture?', 'Is Mixture', default='No')
            manufacture_date = get_col('Manufacture Date', default='')
            last_hydro = get_col('Last Hydro Test Date', 'Last Hydro Date', default='')
            next_hydro = get_col('Next Hydro Test Due', 'Next Hydro Due', default='')
            hydro_status = get_col('Hydro Test Status', default='')
            is_uhp = get_col('Is UHP?', 'Is UHP (Ultra High Purity)?', 'Is UHP', default='No')
            cert_no = get_col('Test Certificate No.', 'Test Certificate No', 'Cert No', default='')
            
            status_val = get_col('Status', default='Active')
            location_val = get_col('Location', default='Depot')
            
            # 1. Add to Supabase Cylinder table
            new_cylinder = Cylinder(
                uid=uid,
                gas_type=gas_type,
                cylinder_type=cyl_type,
                owner=owner,
                status=status_val,
                location=location_val,
                last_activity_date=today_str
            )
            db.session.add(new_cylinder)
            
            # 2. Add to Supabase CylinderMaintenance table
            m_db = CylinderMaintenance(
                cylinder_uid=uid,
                water_capacity=water_cap if water_cap else raw_cyl_type, # Keep 46.7L in water capacity
                fill_pressure=fill_pressure,
                gas_capacity=gas_capacity,
                unit=unit,
                is_mixture=is_mixture,
                manufacture_date=manufacture_date,
                last_hydro_date=last_hydro,
                next_hydro_due=next_hydro,
                hydro_test_status=hydro_status,
                cert_no=cert_no,
                is_uhp=is_uhp
            )
            db.session.add(m_db)
            
            # 3. Prepare Sheets Rows
            cyl_rows_to_append.append([
                uid, gas_type, cyl_type, owner, status_val, location_val, today_str
            ])
            cyl_maint_rows_to_append.append([
                uid, water_cap if water_cap else raw_cyl_type, fill_pressure, gas_capacity, unit, is_mixture, '', manufacture_date, last_hydro, next_hydro, hydro_status, cert_no, is_uhp
            ])
            
            existing_uids.add(uid)
            added += 1
            
        db.session.commit()
        
        # 4. Mirror to Google Sheets using background sync
        if added > 0:
            def background_upload_sheets():
                try:
                    global cyl_ws, cyl_maint_ws, doc
                    if cyl_ws is None and doc:
                        try: cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                        except Exception: pass
                    if cyl_maint_ws is None and doc:
                        try: cyl_maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
                        except Exception: pass
                        
                    if cyl_ws:
                        cyl_ws.append_rows(cyl_rows_to_append)
                    if cyl_maint_ws:
                        cyl_maint_ws.append_rows(cyl_maint_rows_to_append)
                except Exception as se:
                    print("[sheets] Error mirroring bulk upload to Sheets:", se)

            async_sheets_write(background_upload_sheets)
            flash(f"Successfully added {added} new cylinders to Database and syncing to Sheets! Skipped {skipped} duplicates.", "success")
        else:
            flash(f"No new cylinders added. Skipped {skipped} duplicates.", "info")
            
    except Exception as e:
        db.session.rollback()
        flash(f"Error processing file: {str(e)}", "danger")
        
    return redirect('/admin/cylinders')

@app.route('/admin/cylinders/add', methods=['GET', 'POST'])
@admin_required
def admin_cylinders_add():
    global cyl_ws, cyl_maint_ws
    products = get_products_config()
    gas_types = sorted(list(set(p['gas_type'] for p in products if p.get('gas_type'))))
    cylinder_types = sorted(list(set(p['cylinder_type'] for p in products if p.get('cylinder_type'))))

    if request.method == 'POST':
        data = request.form
        uid = data.get('uid', '').strip()
        if not uid:
            return render_template('cylinders_form.html',
                user=session['user'], mode='add',
                error='Cylinder UID is required.', form=data,
                gas_types=gas_types, cylinder_types=cylinder_types)
        # Check duplicate
        existing = get_all_cylinders()
        if any(c['uid'].upper() == uid.upper() for c in existing):
            return render_template('cylinders_form.html',
                user=session['user'], mode='add',
                error=f'Cylinder UID "{uid}" already exists.', form=data,
                gas_types=gas_types, cylinder_types=cylinder_types)
        try:
            db_written = False
            # Write to database (PostgreSQL) first
            if os.environ.get('DATABASE_URL'):
                try:
                    c_db = Cylinder(
                        uid=uid,
                        gas_type=data.get('gas_type', '').strip(),
                        cylinder_type=data.get('cylinder_type', '').strip(),
                        owner=data.get('owner', 'Depot').strip(),
                        status=data.get('status', 'Active').strip(),
                        location=data.get('location', 'Depot').strip(),
                        last_activity_date=date.today().strftime('%d-%m-%Y')
                    )
                    db.session.add(c_db)
                    
                    m_db = CylinderMaintenance(
                        cylinder_uid=uid,
                        water_capacity=data.get('water_capacity', '').strip(),
                        fill_pressure=data.get('fill_pressure', '').strip(),
                        gas_capacity=data.get('gas_capacity', '').strip(),
                        unit=data.get('unit', '').strip(),
                        is_mixture=data.get('is_mixture', 'No').strip(),
                        mix_ratio=data.get('mix_ratio', '').strip(),
                        manufacture_date=data.get('manufacture_date', '').strip(),
                        last_hydro_date=data.get('last_hydro_date', '').strip(),
                        next_hydro_due=data.get('next_hydro_due', '').strip(),
                        hydro_test_status=data.get('hydro_test_status', '').strip(),
                        cert_no=data.get('cert_no', '').strip(),
                        is_uhp='Yes' if data.get('is_uhp') == 'Yes' else 'No'
                    )
                    db.session.add(m_db)
                    db.session.commit()
                    db_written = True
                    print(f"[db] Added cylinder {uid} to PostgreSQL database.")
                except Exception as dbe:
                    db.session.rollback()
                    print("[db] Error adding cylinder to DB:", dbe)
                    from sqlalchemy.exc import OperationalError, InterfaceError
                    is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                    if not is_connection_error:
                        return render_template('cylinders_form.html',
                            user=session['user'], mode='add',
                            error=f"Database validation error: {str(dbe)}", form=data,
                            gas_types=gas_types, cylinder_types=cylinder_types)

            # Add to Google Sheets (Background Sync)
            def background_add_cylinder_sheets():
                try:
                    global cyl_ws, cyl_maint_ws, doc
                    if cyl_ws is None and doc:
                        try: cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                        except Exception: pass
                    if cyl_maint_ws is None and doc:
                        try: cyl_maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
                        except Exception: pass

                    if cyl_ws:
                        cyl_ws.append_row([
                            uid,
                            data.get('gas_type', '').strip(),
                            data.get('cylinder_type', '').strip(),
                            data.get('owner', 'Depot').strip(),
                            data.get('status', 'Active').strip(),
                            data.get('location', 'Depot').strip(),
                            date.today().strftime('%d-%m-%Y'),
                        ])
                    if cyl_maint_ws:
                        cyl_maint_ws.append_row([
                            uid,
                            data.get('water_capacity', '').strip(),
                            data.get('fill_pressure', '').strip(),
                            data.get('gas_capacity', '').strip(),
                            data.get('unit', '').strip(),
                            data.get('is_mixture', 'No').strip(),
                            data.get('mix_ratio', '').strip(),
                            data.get('manufacture_date', '').strip(),
                            data.get('last_hydro_date', '').strip(),
                            data.get('next_hydro_due', '').strip(),
                            data.get('hydro_test_status', '').strip(),
                            data.get('cert_no', '').strip(),
                            'Yes' if data.get('is_uhp') == 'Yes' else 'No',
                        ])
                except Exception as se:
                    print("[sheets] Error mirroring cylinder write to Sheets:", se)
                    if not db_written:
                        raise se

            async_sheets_write(background_add_cylinder_sheets)

            flash(f"Cylinder '{uid}' added successfully!", "success")
            return redirect('/admin/cylinders')
        except Exception as e:
            flash(f"Error adding cylinder: {str(e)}", "danger")
            return render_template('cylinders_form.html',
                user=session['user'], mode='add',
                error=str(e), form=data,
                gas_types=gas_types, cylinder_types=cylinder_types)
    return render_template('cylinders_form.html',
        user=session['user'], mode='add', error=None, form={},
        gas_types=gas_types, cylinder_types=cylinder_types)

@app.route('/admin/cylinders/<uid>/delete', methods=['POST'])
@admin_required
def admin_cylinders_delete(uid):
    global cyl_ws, cyl_maint_ws, doc
    try:
        # Delete from PostgreSQL
        if os.environ.get('DATABASE_URL'):
            try:
                c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
                if c_db:
                    db.session.delete(c_db)
                m_db = CylinderMaintenance.query.filter(CylinderMaintenance.cylinder_uid.ilike(uid)).first()
                if m_db:
                    db.session.delete(m_db)
                db.session.commit()
                print(f"[db] Deleted cylinder {uid} from PostgreSQL.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error deleting cylinder from DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    flash(f"Database error deleting cylinder: {str(dbe)}", "danger")
                    return redirect('/admin/cylinders')

        # Delete from Google Sheets (Background Sync)
        def background_delete_cylinder_sheets():
            try:
                global cyl_ws, cyl_maint_ws, doc
                if cyl_ws is None and doc:
                    try: cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                    except Exception: pass
                if cyl_maint_ws is None and doc:
                    try: cyl_maint_ws = doc.worksheet(CYLINDER_MAINT_NAME)
                    except Exception: pass

                if cyl_ws:
                    rows = cyl_ws.get_all_values()
                    for idx, r in enumerate(rows):
                        if idx == 0: continue
                        if len(r) > 0 and r[0].strip().lower() == uid.lower():
                            cyl_ws.delete_rows(idx + 1)
                            break
                if cyl_maint_ws:
                    mrows = cyl_maint_ws.get_all_values()
                    for idx, r in enumerate(mrows):
                        if idx == 0: continue
                        if len(r) > 0 and r[0].strip().lower() == uid.lower():
                            cyl_maint_ws.delete_rows(idx + 1)
                            break
            except Exception as se:
                print("[sheets] Error mirroring cylinder deletion to Sheets:", se)

        async_sheets_write(background_delete_cylinder_sheets)

        flash(f"Cylinder {uid} deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting cylinder: {str(e)}", "danger")
        
    return redirect('/admin/cylinders')

@app.route('/admin/cylinders/<uid>/edit', methods=['GET', 'POST'])
@admin_required
def admin_cylinders_edit(uid):
    cylinders   = get_all_cylinders()
    maintenance = get_all_maintenance()
    cyl  = next((c for c in cylinders if c['uid'].upper() == uid.upper()), None)
    maint = maintenance.get(uid, maintenance.get(uid.upper(), {}))
    if not cyl:
        return redirect('/admin/cylinders')

    products = get_products_config()
    gas_types = sorted(list(set(p['gas_type'] for p in products if p.get('gas_type'))))
    cylinder_types = sorted(list(set(p['cylinder_type'] for p in products if p.get('cylinder_type'))))

    if request.method == 'POST':
        data = request.form
        try:
            db_written = False
            # Update database first
            if os.environ.get('DATABASE_URL'):
                try:
                    c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
                    if c_db:
                        c_db.gas_type = data.get('gas_type', '').strip()
                        c_db.cylinder_type = data.get('cylinder_type', '').strip()
                        c_db.owner = data.get('owner', '').strip()
                        c_db.status = data.get('status', 'Active').strip()
                        c_db.location = data.get('location', '').strip()
                    
                    m_db = CylinderMaintenance.query.filter(CylinderMaintenance.cylinder_uid.ilike(uid)).first()
                    if not m_db:
                        m_db = CylinderMaintenance(cylinder_uid=uid)
                        db.session.add(m_db)
                    
                    m_db.water_capacity = data.get('water_capacity', '').strip()
                    m_db.fill_pressure = data.get('fill_pressure', '').strip()
                    m_db.gas_capacity = data.get('gas_capacity', '').strip()
                    m_db.unit = data.get('unit', '').strip()
                    m_db.is_mixture = data.get('is_mixture', 'No').strip()
                    m_db.mix_ratio = data.get('mix_ratio', '').strip()
                    m_db.manufacture_date = data.get('manufacture_date', '').strip()
                    m_db.last_hydro_date = data.get('last_hydro_date', '').strip()
                    m_db.next_hydro_due = data.get('next_hydro_due', '').strip()
                    m_db.hydro_test_status = data.get('hydro_test_status', '').strip()
                    m_db.cert_no = data.get('cert_no', '').strip()
                    m_db.is_uhp = 'Yes' if data.get('is_uhp') == 'Yes' else 'No'
                    
                    db.session.commit()
                    db_written = True
                    print(f"[db] Updated cylinder {uid} in PostgreSQL database.")
                except Exception as dbe:
                    db.session.rollback()
                    print("[db] Error updating cylinder in DB:", dbe)
                    from sqlalchemy.exc import OperationalError, InterfaceError
                    is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                    if not is_connection_error:
                        merged = {**cyl, **maint}
                        return render_template('cylinders_form.html',
                            user=session['user'], mode='edit',
                            error=f"Database validation error: {str(dbe)}", form=merged, uid=uid,
                            gas_types=gas_types, cylinder_types=cylinder_types)

            # Mirror edit to Sheets (Background Sync)
            def background_edit_cylinder_sheets():
                try:
                    cyl_row, maint_row = find_cylinder_rows(uid)
                    if cyl_row and cyl_ws:
                        cyl_ws.update(f'A{cyl_row}:G{cyl_row}', [[
                            uid,
                            data.get('gas_type', '').strip(),
                            data.get('cylinder_type', '').strip(),
                            data.get('owner', '').strip(),
                            data.get('status', 'Active').strip(),
                            data.get('location', '').strip(),
                            cyl.get('last_activity', ''),
                        ]])
                    if maint_row and cyl_maint_ws:
                        cyl_maint_ws.update(f'A{maint_row}:M{maint_row}', [[
                            uid,
                            data.get('water_capacity', '').strip(),
                            data.get('fill_pressure', '').strip(),
                            data.get('gas_capacity', '').strip(),
                            data.get('unit', '').strip(),
                            data.get('is_mixture', 'No').strip(),
                            data.get('mix_ratio', '').strip(),
                            data.get('manufacture_date', '').strip(),
                            data.get('last_hydro_date', '').strip(),
                            data.get('next_hydro_due', '').strip(),
                            data.get('hydro_test_status', '').strip(),
                            data.get('cert_no', '').strip(),
                            'Yes' if data.get('is_uhp') == 'Yes' else 'No',
                        ]])
                except Exception as se:
                    print("[sheets] Error mirroring cylinder edit to Sheets:", se)
                    if not db_written:
                        raise se
                        
            async_sheets_write(background_edit_cylinder_sheets)
            
            flash("Cylinder updated successfully!", "success")
            return redirect('/admin/cylinders')
        except Exception as e:
            flash(f"Error updating cylinder: {str(e)}", "danger")
            merged = {**cyl, **maint}
            return render_template('cylinders_form.html',
                user=session['user'], mode='edit',
                error=str(e), form=merged, uid=uid,
                gas_types=gas_types, cylinder_types=cylinder_types)

    merged = {**cyl, **maint}
    return render_template('cylinders_form.html',
        user=session['user'], mode='edit', error=None, form=merged, uid=uid,
        gas_types=gas_types, cylinder_types=cylinder_types)

@app.route('/admin/cylinders/<uid>')
@admin_required
def admin_cylinder_detail(uid):
    cylinders   = get_all_cylinders()
    maintenance = get_all_maintenance()
    cyl   = next((c for c in cylinders if c['uid'].upper() == uid.upper()), None)
    if not cyl:
        return redirect('/admin/cylinders')
    maint = maintenance.get(uid, maintenance.get(uid.upper(), {}))
    hydro_badge = compute_hydro_badge(maint.get('next_hydro_due', ''))
    history = get_cylinder_history(uid)
    return render_template('cylinder_detail.html',
        user        = session['user'],
        cyl         = cyl,
        maint       = maint,
        hydro_badge = hydro_badge,
        history     = history,
    )


@app.route('/admin/cylinders/mark_collected', methods=['POST'])
@admin_required
def admin_mark_collected():
    uid = request.form.get('uid', '').strip()
    customer = request.form.get('customer', '').strip()
    redirect_url = request.form.get('redirect_url', '/admin/dashboard').strip()
    
    if not uid:
        return redirect(redirect_url)
        
    driver_type = request.form.get('driver_type', 'select')
    if driver_type == 'custom':
        driver = request.form.get('driver_custom', '').strip()
        if not driver:
            driver = "Admin (Manual)"
    else:
        driver = request.form.get('driver_select', '').strip()
        if not driver:
            driver = "Admin (Manual)"
            
    now = datetime.now()
    # Resolve current gas type of the cylinder at scan time
    current_gas = ''
    try:
        if os.environ.get('DATABASE_URL'):
            c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
            if c_db:
                current_gas = c_db.gas_type or ''
        if not current_gas:
            sheet_cyls = get_all_cylinders()
            match = next((c for c in sheet_cyls if c['uid'].strip().upper() == uid.upper()), None)
            if match:
                current_gas = match.get('gas_type', '')
    except Exception as ge:
        print("[scan] Error looking up gas type for collection scan:", ge)

    # Columns: Date, Time, Driver, Action, UID, Customer, Gas Type (Column G)
    row_to_append = [
        now.strftime('%d-%m-%Y'),
        now.strftime('%H:%M:%S'),
        driver,
        'Collection',
        uid,
        customer,
        current_gas
    ]
    db_written = False
    try:
        if os.environ.get('DATABASE_URL'):
            try:
                with db.session.no_autoflush:
                    scan = Scan(
                        scan_date=now.strftime('%d-%m-%Y'),
                        scan_time=now.strftime('%H:%M:%S'),
                        driver=driver,
                        action='Collection',
                        cylinder_uid=uid,
                        customer=customer,
                        gas_type=current_gas
                    )
                    db.session.add(scan)
                    
                    c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
                    if c_db:
                        c_db.status = 'Empty'
                        c_db.location = 'Depot'
                        c_db.last_activity_date = now.strftime('%d-%m-%Y')
                        
                    # Save CustomerMap batch record
                    cmap = CustomerMap(
                        scan_date=now.strftime('%d-%m-%Y'),
                        scan_time=now.strftime('%H:%M:%S'),
                        driver=driver,
                        action='Collection',
                        count=1,
                        uids=uid,
                        customer=customer,
                        send_receipt=False,
                        receipt_status=''
                    )
                    db.session.add(cmap)
                    
                db.session.commit()
                db_written = True
                print(f"[db] Logged manual collection scan and updated cylinder {uid} status in DB.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error logging manual collection scan in DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return f"Database error logging manual collection: {str(dbe)}", 500

        # Mirror to Google Sheets
        try:
            global sheet, map_ws
            if sheet:
                sheets_write_with_retry(sheet.append_rows, [row_to_append])
                
            if map_ws:
                map_row = [
                    now.strftime('%d-%m-%Y'),
                    now.strftime('%H:%M:%S'),
                    driver,
                    'Collection',
                    1,
                    uid,
                    customer,
                    'FALSE',
                    ''
                ]
                sheets_write_with_retry(map_ws.append_rows, [map_row])
                
            # Update Cylinders registry sheet
            global cyl_ws
            if cyl_ws is None and doc:
                try:
                    cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                except Exception:
                    cyl_ws = None
            if cyl_ws:
                cyl_rows = cyl_ws.get_all_values()
                row_num = None
                for idx, r in enumerate(cyl_rows):
                    if idx == 0:
                        continue
                    if r and r[0].strip().upper() == uid.upper():
                        row_num = idx + 1
                        break
                if row_num:
                    today_str = now.strftime('%d-%m-%Y')
                    cyl_ws.update(f'E{row_num}:G{row_num}', [['Empty', 'Depot', today_str]])
        except Exception as se:
            print("[sheets] Error mirroring manual collection to Sheets:", se)
            if not db_written:
                raise se
    except Exception as e:
        print("Manual collection log write / registry update error:", e)
        return str(e), 500
        
    clear_cache()
    return redirect(redirect_url)


@app.route('/admin/customers/bulk_collect', methods=['POST'])
@admin_required
def admin_bulk_collect():
    customer = request.form.get('customer', '').strip()
    uids_raw = request.form.get('uids', '').strip()
    redirect_url = request.form.get('redirect_url', '/admin/dashboard').strip()
    
    if not uids_raw or not customer:
        return redirect(redirect_url)
        
    uids = [u.strip().upper() for u in uids_raw.split(',') if u.strip()]
    if not uids:
        return redirect(redirect_url)
        
    driver_type = request.form.get('driver_type', 'select')
    if driver_type == 'custom':
        driver = request.form.get('driver_custom', '').strip()
        if not driver:
            driver = "Admin (Manual)"
    else:
        driver = request.form.get('driver_select', '').strip()
        if not driver:
            driver = "Admin (Manual)"
            
    now = datetime.now()
    today_str = now.strftime('%d-%m-%Y')
    time_str = now.strftime('%H:%M:%S')
    
    # 1. Resolve gas types for the cylinders
    cyl_gas_map = {}
    try:
        if os.environ.get('DATABASE_URL'):
            cyls = Cylinder.query.filter(Cylinder.uid.in_(uids)).all()
            for c in cyls:
                cyl_gas_map[c.uid.strip().upper()] = c.gas_type or ''
        # Fallback to Sheets mapping for missing ones
        missing_uids = [u for u in uids if u not in cyl_gas_map]
        if missing_uids:
            sheet_cyls = get_all_cylinders()
            for c in sheet_cyls:
                uid_up = c['uid'].strip().upper()
                if uid_up in missing_uids:
                    cyl_gas_map[uid_up] = c.get('gas_type', '')
    except Exception as ge:
        print("[scan] Error looking up gas types for bulk collection scans:", ge)

    # 2. Create Scan rows to append
    rows_to_append = []
    for uid in uids:
        gas = cyl_gas_map.get(uid.strip().upper(), '')
        rows_to_append.append([
            today_str,
            time_str,
            driver,
            'Collection',
            uid,
            customer,
            gas
        ])
        
    # 3. CustomerMap batch row to append
    # Columns in Sheet 2: Date, Time, Driver, Action, Count, UIDs, Customer, Send Receipt?, Status
    map_row = [
        today_str,
        time_str,
        driver,
        'Collection',
        len(uids),
        ', '.join(uids),
        customer,
        'FALSE',
        ''
    ]
    
    db_written = False
    try:
        if os.environ.get('DATABASE_URL'):
            try:
                with db.session.no_autoflush:
                    # Save individual scans
                    for uid in uids:
                        gas = cyl_gas_map.get(uid.strip().upper(), '')
                        scan = Scan(
                            scan_date=today_str,
                            scan_time=time_str,
                            driver=driver,
                            action='Collection',
                            cylinder_uid=uid,
                            customer=customer,
                            gas_type=gas
                        )
                        db.session.add(scan)
                        
                        # Update Cylinder registry
                        c_db = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
                        if c_db:
                            c_db.status = 'Empty'
                            c_db.location = 'Depot'
                            c_db.last_activity_date = today_str
                            
                    # Save CustomerMap batch record
                    cmap = CustomerMap(
                        scan_date=today_str,
                        scan_time=time_str,
                        driver=driver,
                        action='Collection',
                        count=len(uids),
                        uids=', '.join(uids),
                        customer=customer,
                        send_receipt=False,
                        receipt_status=''
                    )
                    db.session.add(cmap)
                    
                db.session.commit()
                db_written = True
                print(f"[db] Logged bulk collection scans ({len(uids)}) and updated cylinder status in DB.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error logging bulk collection in DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return f"Database error logging bulk collection: {str(dbe)}", 500
                    
        # Mirror to Google Sheets
        try:
            global sheet, map_ws, cyl_ws
            if sheet and rows_to_append:
                sheets_write_with_retry(sheet.append_rows, rows_to_append)
            if map_ws:
                sheets_write_with_retry(map_ws.append_rows, [map_row])
                
            # Update Cylinders registry sheet
            if cyl_ws is None and doc:
                try:
                    cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                except Exception:
                    cyl_ws = None
            if cyl_ws:
                cyl_rows = cyl_ws.get_all_values()
                uid_row_map = {r[0].strip().upper(): idx + 1 for idx, r in enumerate(cyl_rows) if r and r[0].strip()}
                
                # Perform cell updates
                for uid in uids:
                    row_num = uid_row_map.get(uid)
                    if row_num:
                        cyl_ws.update(f'E{row_num}:G{row_num}', [['Empty', 'Depot', today_str]])
        except Exception as se:
            print("[sheets] Error mirroring bulk collection to Sheets:", se)
            if not db_written:
                raise se
    except Exception as e:
        print("Bulk collection log write / registry update error:", e)
        return str(e), 500
        
    clear_cache()
    return redirect(redirect_url)


# ================================================================
#  EXISTING ROUTES (unchanged)
# ================================================================

# ================================================================
#  CUSTOMER PROFILE ROUTES
# ================================================================

def get_all_customer_info():
    """Returns list of dicts from Customers database table, falling back to Sheets: {id, name, email, phone, address, cold_call_done}"""
    try:
        if os.environ.get('DATABASE_URL'):
            customers = Customer.query.all()
            return [{
                'id'             : c.customer_id,
                'name'           : c.name,
                'email'          : c.email or '',
                'phone'          : c.phone or '',
                'address'        : c.address or '',
                'cold_call_done' : bool(c.cold_call_done),
            } for c in customers]
    except Exception as e:
        print('[db] Error getting customer info from DB, falling back to Sheets:', e)

    try:
        if customer_ws is None:
            return []
        values = customer_ws.get_all_values()
        if len(values) < 2:
            return []
        out = []
        for i, row in enumerate(values[1:], start=1):
            if len(row) < 2 or not row[1].strip():
                continue
            is_done = False
            if len(row) > 5:
                val = row[5].strip().upper()
                is_done = (val in ('ON', 'TRUE', '1', 'YES'))
            out.append({
                'id'             : row[0].strip() if len(row) > 0 else f'C{str(i).zfill(3)}',
                'name'           : row[1].strip() if len(row) > 1 else '',
                'email'          : row[2].strip() if len(row) > 2 else '',
                'phone'          : row[3].strip() if len(row) > 3 else '',
                'address'        : row[4].strip() if len(row) > 4 else '',
                'cold_call_done' : is_done,
            })
        return out
    except Exception as e:
        print('Error getting customer info from sheet:', e)
        return []

def build_customer_outstanding_detail(customer_name):
    """
    Returns list of dicts for cylinders currently with this customer:
    {uid, gas_type, delivered_on, days_out, overdue}
    """
    events = build_events()
    cylinder_owner = {}
    cylinder_delivery_date = {}

    for ev in events:
        if ev['action'] == 'Delivery':
            cylinder_owner[ev['uid']] = ev['customer']
            cylinder_delivery_date[ev['uid']] = ev['date_obj']
        elif ev['action'] in ('Collection', 'Filling'):
            cylinder_owner.pop(ev['uid'], None)
            cylinder_delivery_date.pop(ev['uid'], None)

    # Build gas type map from Cylinders registry (if available)
    gas_map = {}
    try:
        cyls = get_all_cylinders()
        gas_map = {c['uid'].upper(): c.get('gas_type', '') for c in cyls}
    except Exception:
        pass

    today = date.today()
    result = []
    for uid, cust in cylinder_owner.items():
        if cust.lower() != customer_name.lower():
            continue
        d_date   = cylinder_delivery_date.get(uid)
        days_out = (today - d_date).days if d_date else None
        result.append({
            'uid'         : uid,
            'gas_type'    : gas_map.get(uid.upper(), '—'),
            'delivered_on': fmt_date(d_date),
            'days_out'    : days_out,
            'overdue'     : days_out is not None and days_out > 30,
        })

    result.sort(key=lambda x: x['days_out'] or 0, reverse=True)
    return result

def build_customer_history(customer_name):
    """
    All delivery/collection events for a customer, newest first.
    Returns list of dicts: {date, time, driver, action, uids, count}
    """
    batch_map = build_batch_map()
    scan_rows = get_scan_rows()

    # Collect all batches that belong to this customer
    matching_batches = set()
    for key, cust in batch_map.items():
        if cust.lower() == customer_name.lower():
            matching_batches.add(key)

    # Group scan rows by batch key
    batches = {}
    for r in scan_rows:
        key = f"{r['date']}||{r['time']}||{r['driver']}||{r['action']}"
        if key not in matching_batches:
            continue
        if key not in batches:
            batches[key] = {
                'date'    : r['date'],
                'time'    : r['time'],
                'driver'  : r['driver'],
                'action'  : r['action'],
                'uids'    : [],
                'date_obj': parse_date(r['date']),
            }
        batches[key]['uids'].append(r['uid'])

    result = []
    for b in batches.values():
        result.append({
            **b,
            'count': len(b['uids']),
            'uids_str': ', '.join(b['uids']),
        })

    result.sort(key=lambda x: (x['date_obj'] or date.min, x['time']), reverse=True)
    return result

def build_customer_statement(customer_name):
    """
    Running balance statement for a customer — like a bank statement for cylinders.
    Returns list of dicts: {date, time, driver, action, change, balance, uids_str}
    """
    history = build_customer_history(customer_name)
    # history is newest first — reverse for running balance
    history_asc = list(reversed(history))
    balance = 0
    statement = []
    for event in history_asc:
        if event['action'] == 'Delivery':
            change  = +event['count']
        elif event['action'] == 'Collection':
            change  = -event['count']
        else:
            change  = 0
        balance += change
        statement.append({
            **event,
            'change' : change,
            'balance': balance,
        })
    # Return newest first for display
    return list(reversed(statement))

def build_gas_type_breakdown(outstanding_detail):
    """
    Returns dict of {gas_type: count} from outstanding cylinders.
    """
    breakdown = {}
    for c in outstanding_detail:
        gt = c.get('gas_type') or '—'
        breakdown[gt] = breakdown.get(gt, 0) + 1
    return dict(sorted(breakdown.items()))


# ================================================================
#  USER MANAGEMENT ROUTES
# ================================================================

@app.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.order_by(User.username.asc()).all()
    return render_template('users_list.html', user=session['user'], users=users)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@admin_required
def admin_users_add():
    if request.method == 'POST':
        name     = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role     = request.form.get('role', 'driver').strip().lower()

        if not username or not password or not name:
            return render_template('users_form.html', user=session['user'], mode='add', error="Name, username, and password are required.", form=request.form)

        # Check unique username
        existing = User.query.filter(db.func.lower(User.username) == username.lower()).first()
        if existing:
            return render_template('users_form.html', user=session['user'], mode='add', error=f"Username '{username}' is already taken.", form=request.form)

        if role not in ['driver', 'filler', 'manager', 'owner']:
            return render_template('users_form.html', user=session['user'], mode='add', error="Invalid role selected.", form=request.form)

        try:
            hashed_pw = generate_password_hash(password)
            new_user = User(username=username, name=name, password=hashed_pw, role=role)
            db.session.add(new_user)
            db.session.commit()
            
            # Local background sync function
            def background_add_user_sheets():
                try:
                    global users_ws, doc
                    if users_ws is None and doc:
                        try: users_ws = doc.worksheet(USERS_SHEET_NAME)
                        except Exception: pass
                    if users_ws:
                        users_ws.append_row([
                            username,
                            hashed_pw,
                            role,
                            name
                        ])
                except Exception as se:
                    print("[sheets] Error syncing user addition:", se)
            
            async_sheets_write(background_add_user_sheets)
            flash(f"User '{username}' created successfully.", "success")
            return redirect('/admin/users')
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error adding user: {str(e)}", "danger")
            return render_template('users_form.html', user=session['user'], mode='add', error=f"Database error: {str(e)}", form=request.form)

    return render_template('users_form.html', user=session['user'], mode='add', error=None, form={})

@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def admin_users_edit(user_id):
    target_user = User.query.filter_by(id=user_id).first()
    if not target_user:
        flash("User not found.", "danger")
        return redirect('/admin/users')

    if request.method == 'POST':
        name     = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role     = request.form.get('role', 'driver').strip().lower()

        if not username or not name:
            return render_template('users_form.html', user=session['user'], mode='edit', user_id=user_id, error="Name and username are required.", form=request.form)

        # Check unique username
        existing = User.query.filter(db.func.lower(User.username) == username.lower()).filter(User.id != user_id).first()
        if existing:
            return render_template('users_form.html', user=session['user'], mode='edit', user_id=user_id, error=f"Username '{username}' is already taken.", form=request.form)

        if role not in ['driver', 'filler', 'manager', 'owner']:
            return render_template('users_form.html', user=session['user'], mode='edit', user_id=user_id, error="Invalid role selected.", form=request.form)

        old_username = target_user.username
        
        try:
            target_user.name = name
            target_user.username = username
            target_user.role = role
            
            if password:
                hashed_pw = generate_password_hash(password)
                target_user.password = hashed_pw
            else:
                hashed_pw = target_user.password
                
            db.session.commit()
            
            # Local background sync function
            def background_edit_user_sheets():
                try:
                    global users_ws, doc
                    if users_ws is None and doc:
                        try: users_ws = doc.worksheet(USERS_SHEET_NAME)
                        except Exception: pass
                    if users_ws:
                        rows = users_ws.get_all_values()
                        for idx, r in enumerate(rows):
                            if idx == 0: continue
                            if len(r) > 0 and r[0].strip().lower() == old_username.strip().lower():
                                users_ws.update(f'A{idx+1}:D{idx+1}', [[username, hashed_pw, role, name]])
                                break
                except Exception as se:
                    print("[sheets] Error syncing user update:", se)
            
            async_sheets_write(background_edit_user_sheets)
            
            # If editing yourself, update session to reflect new role or username
            if session['user']['username'].lower() == old_username.lower():
                session['user']['username'] = username
                session['user']['role'] = role
                session['user']['name'] = name
                
            flash(f"User '{username}' updated successfully.", "success")
            return redirect('/admin/users')
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating user: {str(e)}", "danger")
            return render_template('users_form.html', user=session['user'], mode='edit', user_id=user_id, error=f"Database error: {str(e)}", form=request.form)

    # Prepopulate form
    form_data = {
        'name': target_user.name,
        'username': target_user.username,
        'role': target_user.role
    }
    return render_template('users_form.html', user=session['user'], mode='edit', user_id=user_id, error=None, form=form_data)

@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def admin_users_delete(user_id):
    target_user = User.query.filter_by(id=user_id).first()
    if not target_user:
        flash("User not found.", "danger")
        return redirect('/admin/users')
    
    # Prevent self-deletion
    if target_user.username.lower() == session['user']['username'].lower():
        flash("You cannot delete your own user account.", "danger")
        return redirect('/admin/users')
        
    username = target_user.username
    try:
        db.session.delete(target_user)
        db.session.commit()
        
        # Local background sync function
        def background_delete_user_sheets():
            try:
                global users_ws, doc
                if users_ws is None and doc:
                    try: users_ws = doc.worksheet(USERS_SHEET_NAME)
                    except Exception: pass
                if users_ws:
                    rows = users_ws.get_all_values()
                    for idx, r in enumerate(rows):
                        if idx == 0: continue
                        if len(r) > 0 and r[0].strip().lower() == username.strip().lower():
                            users_ws.delete_rows(idx + 1)
                            break
            except Exception as se:
                print("[sheets] Error syncing user deletion:", se)
                
        async_sheets_write(background_delete_user_sheets)
        flash(f"User '{username}' deleted successfully.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Database error deleting user: {str(e)}", "danger")
        
    return redirect('/admin/users')


@app.route('/admin/customers')
@admin_required
def admin_customers():
    # All customers from Customers sheet
    all_info   = get_all_customer_info()
    info_map   = {c['name'].lower(): c for c in all_info}
    outstanding = build_outstanding()

    # Build aging data for overdue flags
    aging = build_aging()
    overdue_customers = set()
    for a in aging:
        if a['status'] == 'overdue':
            overdue_customers.add(a['customer'].lower())

    # Merge: customer sheet info + outstanding stats
    result = []
    seen = set()
    for o in outstanding:
        name  = o['customer']
        name_l = name.lower()
        seen.add(name_l)
        info = info_map.get(name_l, {})
        result.append({
            'name'            : name,
            'email'           : info.get('email', ''),
            'phone'           : info.get('phone', ''),
            'id'              : info.get('id', ''),
            'outstanding'     : o['outstanding'],
            'total_delivered' : o['total_delivered'],
            'total_collected' : o['total_collected'],
            'last_activity'   : o['last_activity'],
            'has_overdue'     : name_l in overdue_customers,
        })

    # Include customers from Customers sheet who have no scans yet
    for info in all_info:
        if info['name'].lower() not in seen:
            result.append({
                'name'           : info['name'],
                'email'          : info['email'],
                'phone'          : info['phone'],
                'id'             : info['id'],
                'outstanding'    : 0,
                'total_delivered': 0,
                'total_collected': 0,
                'last_activity'  : '—',
                'has_overdue'    : False,
            })

    overdue_customers_list = sorted([
        c['name'] for c in result if c['has_overdue']
    ])

    return render_template('customers_list.html',
        user      = session['user'],
        customers = result,
        total     = len(result),
        with_outstanding = sum(1 for c in result if c['outstanding'] > 0),
        overdue_count    = len(overdue_customers),
        overdue_customers_list = overdue_customers_list,
    )


@app.route('/admin/customers/<path:customer_name>')
@admin_required
def admin_customer_profile(customer_name):
    # Decode URL-encoded name
    from urllib.parse import unquote
    customer_name = unquote(customer_name)

    # Basic info from Customers sheet
    all_info = get_all_customer_info()
    info = next((c for c in all_info if c['name'].lower() == customer_name.lower()), {})

    # Outstanding detail
    outstanding_detail = build_customer_outstanding_detail(customer_name)
    overdue_cyls  = [c for c in outstanding_detail if c['overdue']]

    # Gas type breakdown
    gas_breakdown = build_gas_type_breakdown(outstanding_detail)

    # History + statement
    history   = build_customer_history(customer_name)
    statement = build_customer_statement(customer_name)

    # Summary stats
    total_delivered = sum(e['count'] for e in history if e['action'] == 'Delivery')
    total_collected = sum(e['count'] for e in history if e['action'] == 'Collection')

    scan_rows = get_scan_rows()
    drivers = sorted(list(set(r['driver'].strip() for r in scan_rows if r.get('driver') and r['driver'].strip())))

    receipt_history = get_customer_receipt_history(customer_name)

    return render_template('customer_profile.html',
        user               = session['user'],
        info               = info,
        customer_name      = customer_name,
        outstanding_detail = outstanding_detail,
        outstanding_count  = len(outstanding_detail),
        overdue_cyls       = overdue_cyls,
        gas_breakdown      = gas_breakdown,
        history            = history,
        statement          = statement,
        total_delivered    = total_delivered,
        total_collected    = total_collected,
        drivers            = drivers,
        receipt_history    = receipt_history,
    )


# ================================================================
#  CUSTOMER ADD / EDIT ROUTES
# ================================================================

@app.route('/admin/customers/add', methods=['GET', 'POST'])
@admin_required
def admin_customers_add():
    global customer_ws
    ensure_customer_columns()
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        
        if not name:
            return render_template('customers_form.html',
                user=session['user'], mode='add',
                error='Customer Name is required.', form=request.form)
                
        # Check duplicate
        existing = get_customer_names()
        if any(x.lower() == name.lower() for x in existing):
            return render_template('customers_form.html',
                user=session['user'], mode='add',
                error=f'Customer "{name}" already exists.', form=request.form)
                
        try:
            # Generate ID
            all_info = get_all_customer_info()
            max_id = 0
            for c in all_info:
                id_str = c.get('id', '')
                if id_str.startswith('C'):
                    try:
                        val = int(id_str[1:])
                        if val > max_id: max_id = val
                    except ValueError: pass
            new_id = f"C{str(max_id + 1).zfill(3)}"

            db_written = False
            if os.environ.get('DATABASE_URL'):
                try:
                    cust_db = Customer(
                        customer_id=new_id,
                        name=name,
                        email=email,
                        phone=phone,
                        address=address
                    )
                    db.session.add(cust_db)
                    db.session.commit()
                    db_written = True
                    print(f"[db] Added customer {name} to PostgreSQL.")
                except Exception as dbe:
                    db.session.rollback()
                    print("[db] Error adding customer to DB:", dbe)
                    from sqlalchemy.exc import OperationalError, InterfaceError
                    is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                    if not is_connection_error:
                        return render_template('customers_form.html',
                            user=session['user'], mode='add',
                            error=f"Database validation error: {str(dbe)}", form=request.form)

            # Mirror to Google Sheets (Background Sync)
            def background_add_customer_sheets():
                try:
                    global customer_ws, doc
                    if customer_ws is None and doc:
                        customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
                        
                    if customer_ws:
                        # Find the first row where Name (Column B) is empty
                        all_rows = customer_ws.get_all_values()
                        empty_row_idx = None
                        for idx, r in enumerate(all_rows):
                            if idx == 0: continue # Skip header
                            # Check if Column B (Name) is empty
                            if len(r) < 2 or not r[1].strip():
                                empty_row_idx = idx + 1 # 1-based row index
                                break
                        
                        if empty_row_idx:
                            existing_row = all_rows[empty_row_idx - 1]
                            existing_id = existing_row[0].strip() if len(existing_row) > 0 else ""
                            new_id_sheets = existing_id if existing_id else new_id
                            customer_ws.update(f'A{empty_row_idx}:E{empty_row_idx}', [[new_id_sheets, name, email, phone, address]])
                        else:
                            customer_ws.append_row([new_id, name, email, phone, address])
                except Exception as se:
                    print("[sheets] Error mirroring customer write to Sheets:", se)
                    if not db_written:
                        raise se

            async_sheets_write(background_add_customer_sheets)

            clear_cache()
            flash("Customer added successfully!", "success")
            return redirect('/admin/customers')
        except Exception as e:
            flash(f"Error adding customer: {str(e)}", "danger")
            return render_template('customers_form.html',
                user=session['user'], mode='add',
                error=str(e), form=request.form)
                
    return render_template('customers_form.html',
        user=session['user'], mode='add', error=None, form={})

@app.route('/admin/customers/<path:customer_name>/edit', methods=['GET', 'POST'])
@admin_required
def admin_customers_edit(customer_name):
    from urllib.parse import unquote
    customer_name = unquote(customer_name).strip()
    global customer_ws
    ensure_customer_columns()
    
    all_info = get_all_customer_info()
    cust = next((c for c in all_info if c['name'].lower() == customer_name.lower()), None)
    if not cust:
        return redirect('/admin/customers')
        
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        
        if not name:
            return render_template('customers_form.html',
                user=session['user'], mode='edit', uid=customer_name,
                error='Customer Name is required.', form=request.form)
                
        # Check duplicate if name is changing
        if name.lower() != customer_name.lower():
            existing = get_customer_names()
            if any(x.lower() == name.lower() for x in existing):
                return render_template('customers_form.html',
                    user=session['user'], mode='edit', uid=customer_name,
                    error=f'Customer Name "{name}" is already taken.', form=request.form)
                    
        try:
            db_written = False
            if os.environ.get('DATABASE_URL'):
                try:
                    c_db = Customer.query.filter(Customer.name.ilike(customer_name)).first()
                    if c_db:
                        c_db.name = name
                        c_db.email = email
                        c_db.phone = phone
                        c_db.address = address
                        db.session.commit()
                        db_written = True
                        print(f"[db] Updated customer {name} in PostgreSQL.")
                except Exception as dbe:
                    db.session.rollback()
                    print("[db] Error updating customer in DB:", dbe)
                    from sqlalchemy.exc import OperationalError, InterfaceError
                    is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                    if not is_connection_error:
                        return render_template('customers_form.html',
                            user=session['user'], mode='edit', uid=customer_name,
                            error=f"Database validation error: {str(dbe)}", form=request.form)

            # Mirror edit to Sheets (Background Sync)
            def background_edit_customer_sheets():
                try:
                    global customer_ws, doc
                    if customer_ws is None and doc:
                        customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
                        
                    if customer_ws is not None:
                        rows = customer_ws.get_all_values()
                        row_num = None
                        for idx, r in enumerate(rows):
                            if idx == 0: continue
                            if len(r) > 1 and r[1].strip().lower() == customer_name.lower():
                                row_num = idx + 1
                                break
                                
                        if row_num:
                            # If name changed, cascade
                            if name.lower() != customer_name.lower():
                                rename_customer_in_sheets(customer_name, name)
                            customer_ws.update(f'A{row_num}:E{row_num}', [[cust.get('id', ''), name, email, phone, address]])
                except Exception as se:
                    print("[sheets] Error mirroring customer edit to Sheets:", se)
                    if not db_written:
                        raise se

            async_sheets_write(background_edit_customer_sheets)

            clear_cache()
            flash("Customer updated successfully!", "success")
            return redirect(f'/admin/customers/{name}')
        except Exception as e:
            flash(f"Error updating customer: {str(e)}", "danger")
            return render_template('customers_form.html',
                user=session['user'], mode='edit', uid=customer_name,
                error=str(e), form=request.form)
                
    return render_template('customers_form.html',
        user=session['user'], mode='edit', uid=customer_name, error=None, form=cust)

@app.route('/admin/customers/<path:customer_name>/delete', methods=['POST'])
@admin_required
def admin_customers_delete(customer_name):
    from urllib.parse import unquote
    customer_name = unquote(customer_name).strip()
    global customer_ws
    try:
        db_deleted = False
        if os.environ.get('DATABASE_URL'):
            try:
                c_db = Customer.query.filter(Customer.name.ilike(customer_name)).first()
                if c_db:
                    db.session.delete(c_db)
                    db.session.commit()
                    db_deleted = True
                    print(f"[db] Deleted customer {customer_name} from PostgreSQL.")
            except Exception as dbe:
                db.session.rollback()
                print("[db] Error deleting customer from DB:", dbe)
                from sqlalchemy.exc import OperationalError, InterfaceError
                is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
                if not is_connection_error:
                    return f"Database error deleting customer: {str(dbe)}", 500

        # Mirror delete to Sheets (Background Sync)
        def background_delete_customer_sheets():
            try:
                global customer_ws, doc
                if customer_ws is None and doc:
                    customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
                if customer_ws is not None:
                    rows = customer_ws.get_all_values()
                    row_num = None
                    for idx, r in enumerate(rows):
                        if idx == 0: continue
                        if len(r) > 1 and r[1].strip().lower() == customer_name.lower():
                            row_num = idx + 1
                            break
                    if row_num:
                        customer_ws.update(f'B{row_num}:E{row_num}', [["", "", "", ""]])
            except Exception as se:
                print("[sheets] Error mirroring customer deletion to Sheets:", se)
                if not db_deleted:
                    raise se

        async_sheets_write(background_delete_customer_sheets)
        clear_cache()
        return redirect('/admin/customers')
    except Exception as e:
        print("Error deleting customer:", e)
        return str(e), 500



# ================================================================
#  CUSTOMER COLD CALL CHECKLIST ROUTES
# ================================================================

@app.route('/admin/cold_calls')
@admin_required
def admin_cold_calls():
    customers = get_all_customer_info()
    return render_template('cold_call_checklist.html', user=session['user'], customers=customers)


@app.route('/admin/api/cold_call/toggle', methods=['POST'])
@admin_required
def admin_api_cold_call_toggle():
    global customer_ws
    data = request.get_json() or {}
    cust_name = data.get('customer_name', '').strip()
    if not cust_name:
        return jsonify({'success': False, 'error': 'Customer name required'}), 400

    new_state = False
    db_written = False
    
    # 1. Update database
    if os.environ.get('DATABASE_URL'):
        try:
            cust = Customer.query.filter(Customer.name.ilike(cust_name)).first()
            if cust:
                cust.cold_call_done = not cust.cold_call_done
                new_state = cust.cold_call_done
                db.session.commit()
                db_written = True
        except Exception as e:
            db.session.rollback()
            print("[db] Error toggling cold call in DB:", e)
            
    # 2. Update Sheets
    try:
        if customer_ws is None and doc:
            customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
        if customer_ws is not None:
            rows = customer_ws.get_all_values()
            row_num = None
            for idx, r in enumerate(rows):
                if idx == 0: continue
                if len(r) > 1 and r[1].strip().lower() == cust_name.lower():
                    row_num = idx + 1
                    # If we don't have db, determine next state from Sheets row
                    if not db_written:
                        current_val = r[5].strip().upper() if len(r) > 5 else 'OFF'
                        new_state = (current_val not in ('ON', 'TRUE', '1', 'YES'))
                    break
            if row_num:
                # Ensure the sheet has F1 header as "Cold Call Done" if missing
                if len(rows[0]) < 6:
                    customer_ws.update_cell(1, 6, "Cold Call Done")
                customer_ws.update(f'F{row_num}', [["ON" if new_state else "OFF"]])
    except Exception as e:
        print("[sheets] Error mirroring cold call toggle to Sheets:", e)
        if not db_written:
            return jsonify({'success': False, 'error': str(e)}), 500

    clear_cache()
    return jsonify({'success': True, 'cold_call_done': new_state})


@app.route('/admin/api/cold_call/reset_all', methods=['POST'])
@admin_required
def admin_api_cold_call_reset_all():
    global customer_ws
    db_written = False
    
    # 1. Reset database
    if os.environ.get('DATABASE_URL'):
        try:
            Customer.query.update({Customer.cold_call_done: False})
            db.session.commit()
            db_written = True
        except Exception as e:
            db.session.rollback()
            print("[db] Error resetting cold calls in DB:", e)

    # 2. Reset Sheets
    try:
        if customer_ws is None and doc:
            customer_ws = doc.worksheet(CUSTOMER_SHEET_NAME)
        if customer_ws is not None:
            rows = customer_ws.get_all_values()
            if len(rows) > 1:
                # Ensure header F1 exists
                if len(rows[0]) < 6:
                    customer_ws.update_cell(1, 6, "Cold Call Done")
                
                # Build list of updates to make it faster
                updates = []
                for idx in range(2, len(rows) + 1):
                    updates.append({
                        'range': f'F{idx}',
                        'values': [['OFF']]
                    })
                customer_ws.batch_update(updates)
    except Exception as e:
        print("[sheets] Error resetting cold calls in Sheets:", e)
        if not db_written:
            return jsonify({'success': False, 'error': str(e)}), 500

    clear_cache()
    return jsonify({'success': True})



# ================================================================
#  DURA CYLINDER FILL API ROUTE
# ================================================================

@app.route('/admin/api/dura_fill', methods=['POST'])
@admin_required
def admin_api_dura_fill():
    global cyl_ws
    data = request.get_json() or {}
    uid = data.get('uid', '').strip()
    fill_new = bool(data.get('fill_new_gas', False))
    new_gas = data.get('new_gas', '').strip()
    purge_ack = bool(data.get('purge_acknowledged', False))

    if not uid:
        return jsonify({'success': False, 'error': 'Cylinder UID required'}), 400

    now = datetime.now()
    date_str = now.strftime('%d-%m-%Y')
    time_str = now.strftime('%H:%M:%S')
    operator_name = session.get('user', {}).get('name', session.get('user', {}).get('username', 'Admin'))

    db_written = False
    old_gas = ''

    # 1. Fetch current cylinder and update DB (or create if missing)
    if os.environ.get('DATABASE_URL'):
        try:
            cyl = Cylinder.query.filter(Cylinder.uid.ilike(uid)).first()
            if cyl:
                old_gas = cyl.gas_type or ''
                if fill_new and new_gas:
                    cyl.gas_type = new_gas
            else:
                # Cylinder not in DB yet — get old gas from Sheets for history, then insert cylinder row
                if cyl_ws is None and doc:
                    try:
                        cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                    except Exception:
                        pass
                if cyl_ws is not None:
                    try:
                        sheet_cyls = get_all_cylinders()
                        match = next((c for c in sheet_cyls if c['uid'].strip().upper() == uid.upper()), None)
                        if match:
                            old_gas = match.get('gas_type', '')
                    except Exception:
                        pass
                # Insert the cylinder into DB so future updates work
                new_cyl = Cylinder(
                    uid=uid,
                    gas_type=new_gas if (fill_new and new_gas) else old_gas,
                    cylinder_type='Dura',
                    owner='Depot',
                    status='Active',
                    location='Depot'
                )
                db.session.add(new_cyl)

            # Always log to dura_gas_history
            hist = DuraGasHistory(
                cylinder_uid=uid,
                gas_filled=new_gas if fill_new else old_gas,
                previous_gas=old_gas,
                purge_required=(old_gas.upper() != new_gas.upper()) if (fill_new and old_gas) else False,
                purge_acknowledged=purge_ack,
                operator=operator_name,
                fill_date=date_str,
                fill_time=time_str
            )
            db.session.add(hist)
            db.session.commit()
            db_written = True
            print(f"[db] Refill registered for {uid} in DB.")
        except Exception as e:
            db.session.rollback()
            print("[db] Error updating Dura refill in DB:", e)
            # Return error immediately — don't pretend success if DB write failed
            return jsonify({'success': False, 'error': f'Database write failed: {str(e)}'}), 500

    # 2. Sync to Sheets (best-effort — DB is source of truth)
    try:
        if cyl_ws is None and doc:
            cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
        if cyl_ws is not None:
            # Read directly from sheet to get correct row number (not from get_all_cylinders
            # which may return DB rows in a different order than the sheet)
            sheet_rows = cyl_ws.get_all_values()
            row_num = None
            for idx, r in enumerate(sheet_rows):
                if idx == 0:
                    continue  # skip header
                if r and r[0].strip().upper() == uid.upper():
                    row_num = idx + 1  # 1-based
                    break

            if row_num:
                target_gas = new_gas if fill_new else old_gas
                cyl_ws.update(f'B{row_num}', [[target_gas]])
                print(f"[sheets] Updated cylinder {uid} Gas Type to {target_gas} on Sheet row {row_num}.")
    except Exception as e:
        print("[sheets] Error updating Dura refill on Sheets:", e)
        # Don't fail — DB is already updated successfully

    clear_cache()
    return jsonify({'success': True})



@app.route('/admin/customers/<path:customer_name>/offer')
@admin_required
def admin_customer_offer_form(customer_name):
    from urllib.parse import unquote
    customer_name = unquote(customer_name).strip()
    global customer_ws
    ensure_customer_columns()
    
    # Get customer details
    all_info = get_all_customer_info()
    cust = next((c for c in all_info if c['name'].lower() == customer_name.lower()), None)
    if not cust:
        return redirect('/admin/customers')
        
    # Generate sequential Quotation Number: NAG/26-27/XXXX
    try:
        curr_num = int(get_setting('last_quotation_number', '6674'))
    except (ValueError, TypeError):
        curr_num = 6674
    next_num = curr_num + 1
    set_setting('last_quotation_number', str(next_num))
    q_no = f"NAG/26-27/{next_num}"
    
    # Format today's date with suffix (e.g. 25th May 2026)
    today = date.today()
    day = today.day
    if 4 <= day <= 20 or 24 <= day <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][day % 10 - 1]
    date_str = f"{day}{suffix} {today.strftime('%B %Y')}"
    
    # Default products matching the sample exactly
    default_products = [
        {"name": "Argon (5.5 Grade)", "capacity": "07 Cum", "price": "630", "unit": "Per Cum"},
        {"name": "Argon (5.0 Grade)", "capacity": "07 Cum", "price": "390", "unit": "Per Cum"},
        {"name": "Helium (5.0 Grade)", "capacity": "07 Cum", "price": "5700", "unit": "Per Cum"},
        {"name": "Nitrogen (5.0 Grade)", "capacity": "07 Cum", "price": "135", "unit": "Per Cum"},
        {"name": "Hydrogen (5.0 Grade)", "capacity": "07 Cum", "price": "250", "unit": "Per Cum"},
        {"name": "SF6 Gas", "capacity": "50 Kg", "price": "1050", "unit": "Per Kg"},
    ]
    
    # Default terms as list of {label, value} for dynamic T&C table
    default_terms_list = [
        {"label": "1. Prices",              "value": "As Quoted"},
        {"label": "2. GST",                 "value": "@18% Extra."},
        {"label": "3. Transportation",      "value": "Inclusive Delivery."},
        {"label": "4. Payment",             "value": "30 Days"},
        {"label": "5. Valve Damage",        "value": "Rs. 1000/- per Valve"},
        {"label": "6. Cylinder Lost / Damage", "value": "Rs.10,000/- Per Cylinder"}
    ]

    return render_template('offer_form.html',
        user=session['user'],
        customer=cust,
        customer_name=customer_name,
        q_no=q_no,
        date_str=date_str,
        products=default_products,
        terms_list=default_terms_list,
        standalone=False
    )


# ── Standalone offer routes (no customer required) ──────────────────────────
@app.route('/admin/offer/new')
@admin_required
def admin_offer_new():
    """Standalone Commercial Offer form — not tied to a registered customer."""
    # Generate sequential Quotation Number: NAG/26-27/XXXX
    try:
        curr_num = int(get_setting('last_quotation_number', '6674'))
    except (ValueError, TypeError):
        curr_num = 6674
    next_num = curr_num + 1
    set_setting('last_quotation_number', str(next_num))
    q_no = f"NAG/26-27/{next_num}"
    today = date.today()
    day = today.day
    suffix = "th" if 4 <= day <= 20 or 24 <= day <= 30 else ["st", "nd", "rd"][day % 10 - 1]
    date_str = f"{day}{suffix} {today.strftime('%B %Y')}"

    default_products = [
        {"name": "Argon (5.5 Grade)",    "capacity": "07 Cum", "price": "630",  "unit": "Per Cum"},
        {"name": "Argon (5.0 Grade)",    "capacity": "07 Cum", "price": "390",  "unit": "Per Cum"},
        {"name": "Helium (5.0 Grade)",   "capacity": "07 Cum", "price": "5700", "unit": "Per Cum"},
        {"name": "Nitrogen (5.0 Grade)", "capacity": "07 Cum", "price": "135",  "unit": "Per Cum"},
        {"name": "Hydrogen (5.0 Grade)", "capacity": "07 Cum", "price": "250",  "unit": "Per Cum"},
        {"name": "SF6 Gas",              "capacity": "50 Kg",  "price": "1050", "unit": "Per Kg"},
    ]
    default_terms_list = [
        {"label": "1. Prices",               "value": "As Quoted"},
        {"label": "2. GST",                  "value": "@18% Extra."},
        {"label": "3. Transportation",       "value": "Inclusive Delivery."},
        {"label": "4. Payment",              "value": "30 Days"},
        {"label": "5. Valve Damage",         "value": "Rs. 1000/- per Valve"},
        {"label": "6. Cylinder Lost / Damage", "value": "Rs.10,000/- Per Cylinder"}
    ]
    return render_template('offer_form.html',
        user=session['user'],
        customer=None,
        customer_name='',
        q_no=q_no,
        date_str=date_str,
        products=default_products,
        terms_list=default_terms_list,
        standalone=True
    )

@app.route('/admin/offer/generate', methods=['POST'])
@admin_required
def admin_offer_generate_standalone():
    """Generates PDF for standalone offer (customer name comes from form)."""
    customer_name = request.form.get('customer_name', 'Customer').strip()
    return _generate_offer_pdf(customer_name)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/admin/customers/<path:customer_name>/offer/generate', methods=['POST'])
@admin_required
def admin_generate_offer_pdf(customer_name):
    from urllib.parse import unquote
    customer_name = unquote(customer_name).strip()
    return _generate_offer_pdf(customer_name)

def _generate_offer_pdf(customer_name):
    # Extract form values
    attn = request.form.get('attn', '').strip()
    tel = request.form.get('tel', '').strip()
    q_date = request.form.get('date', '').strip()
    q_no = request.form.get('q_no', '').strip()
    ref = request.form.get('ref', '').strip()

    # Product lists from dynamic fields
    product_names = request.form.getlist('product_name[]')
    product_contents = request.form.getlist('product_content[]')
    product_prices = request.form.getlist('product_price[]')
    product_units = request.form.getlist('product_unit[]')

    # Dynamic T&C rows (both label and value editable)
    term_labels = request.form.getlist('term_label[]')
    term_values = request.form.getlist('term_value[]')
    terms_pairs = list(zip(term_labels, term_values))
    
    # Generate PDF using ReportLab
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = BytesIO()
    # Tighten margins from 36 to 24 to maximize printable vertical space (extra 24pt total height)
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    # ReportLab Styles
    styles = getSampleStyleSheet()
    
    # Brand Typography & Colors
    blue_brand = colors.HexColor('#0c5ca8')
    green_brand = colors.HexColor('#3fb549')
    dark_gray = colors.HexColor('#2c2c2a')
    
    # Custom Styles (compressed leading and sizes slightly where appropriate)
    brand_style1 = ParagraphStyle('Brand1', fontName='Helvetica-Bold', fontSize=26, leading=30, textColor=blue_brand, alignment=1)
    brand_style2 = ParagraphStyle('Brand2', fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=green_brand, alignment=1)
    address_style = ParagraphStyle('Address', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=dark_gray, alignment=1)
    header_contact_style = ParagraphStyle('HeaderContact', fontName='Helvetica-Bold', fontSize=9.5, leading=11.5, textColor=dark_gray, alignment=1)
    
    title_style = ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=colors.black, alignment=1, spaceAfter=6)
    
    intro_style = ParagraphStyle('Intro', fontName='Helvetica-Bold', fontSize=9.5, leading=12, textColor=colors.HexColor('#1D9E75'), alignment=1, spaceBefore=4, spaceAfter=6)
    
    cell_style = ParagraphStyle('Cell', fontName='Helvetica', fontSize=9.5, leading=12, textColor=colors.black, alignment=1)
    cell_bold_style = ParagraphStyle('CellBold', fontName='Helvetica-Bold', fontSize=10, leading=12.5, textColor=colors.black, alignment=1)
    
    left_cell_style = ParagraphStyle('LeftCell', fontName='Helvetica', fontSize=9.5, leading=12, textColor=colors.black, alignment=0)
    left_cell_bold_style = ParagraphStyle('LeftCellBold', fontName='Helvetica-Bold', fontSize=10, leading=12.5, textColor=colors.black, alignment=0)
    
    terms_title_style = ParagraphStyle('TermsTitle', fontName='Helvetica-Bold', fontSize=10.5, leading=12.5, textColor=colors.black, spaceBefore=8, spaceAfter=4)
    terms_item_style = ParagraphStyle('TermsItem', fontName='Helvetica', fontSize=9.5, leading=14, textColor=colors.black, spaceAfter=2)
    
    footer_text_style = ParagraphStyle('FooterText', fontName='Helvetica', fontSize=9.5, leading=14, textColor=colors.black, alignment=0)
    
    # 1. Noble Air Gases Header Layout — use actual logo image
    import os as _os
    from reportlab.platypus import Image as RLImage

    logo_path = _os.path.join(_os.path.dirname(__file__), 'static', 'img', 'noble_logo.png')

    if _os.path.exists(logo_path):
        # Centre the logo: fit within 240pt wide × 60pt tall, keep aspect ratio
        logo_img = RLImage(logo_path, width=240, height=60, kind='proportional')
        logo_img.hAlign = 'CENTER'
        story.append(logo_img)
        story.append(Spacer(1, 8))
    else:
        # Fallback to text if logo file not found
        story.append(Paragraph("NOBLE", brand_style1))
        story.append(Paragraph("air gases", brand_style2))
        story.append(Spacer(1, 4))

    story.append(Paragraph("Plot No. A/12, MIDC Waluj, Chhatrapati Sambhajinagar", address_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph("Email: sales@nobleairgases.com &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Mobile: +91 9225309555", header_contact_style))
    story.append(Spacer(1, 8))

    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#d8d9d4'), spaceAfter=8))
    
    # 2. Document Title
    story.append(Paragraph("COMMERCIAL OFFER", title_style))
    
    # 3. Metadata block table
    meta_data = [
        [
            Paragraph("<b>To</b>", left_cell_bold_style),
            Paragraph(f": M/s {customer_name}", left_cell_style),
            Paragraph("<b>K.Attn</b>", left_cell_bold_style),
            Paragraph(f": {attn}", left_cell_style)
        ],
        [
            Paragraph("<b>Tel</b>", left_cell_bold_style),
            Paragraph(f": {tel}", left_cell_style),
            Paragraph("<b>Date</b>", left_cell_bold_style),
            Paragraph(f": {q_date}", left_cell_style)
        ],
        [
            Paragraph("<b>Q. No.</b>", left_cell_bold_style),
            Paragraph(f": {q_no}", left_cell_style),
            Paragraph("<b>Ref</b>", left_cell_bold_style),
            Paragraph(f": {ref}", left_cell_style)
        ]
    ]
    meta_table = Table(meta_data, colWidths=[50, 220, 50, 220])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#888888')),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 10))
    
    # 4. Intro text
    story.append(Paragraph("Thank you for your interest in our products & services. We are pleased to offer our most Competitive quote for your consideration with regards to your requirements", intro_style))
    
    # 5. Quote product table
    table_data = [
        [
            Paragraph("<b>No.</b>", cell_bold_style),
            Paragraph("<b>Product</b>", cell_bold_style),
            Paragraph("<b>Content per Cylinder</b>", cell_bold_style),
            Paragraph("<b>Price<br/>(Rs/unit)</b>", cell_bold_style),
            Paragraph("<b>Unit</b>", cell_bold_style)
        ]
    ]
    
    idx_no = 1
    for name, content, price, unit in zip(product_names, product_contents, product_prices, product_units):
        if not name.strip(): continue
        table_data.append([
            Paragraph(f"{idx_no:02d}", cell_style),
            Paragraph(f"<b>{name.strip()}</b>", left_cell_bold_style),
            Paragraph(content.strip(), cell_style),
            Paragraph(f"{price.strip()}/-", cell_style),
            Paragraph(unit.strip(), cell_style)
        ])
        idx_no += 1
        
    prod_table = Table(table_data, colWidths=[40, 200, 120, 90, 90])
    prod_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d8d9d4')),
        ('LINEBELOW', (0,0), (-1,0), 1.5, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,0), (-1,0), colors.white),
    ]))
    story.append(prod_table)
    story.append(Spacer(1, 10))
    
    # 6. Terms and conditions — dynamic labels and values from form
    story.append(Paragraph("<b>TERMS &amp; CONDITIONS:</b>", terms_title_style))
    for label, value in terms_pairs:
        if label.strip():
            story.append(Paragraph(f"<b>{label.strip()}</b> : {value.strip()}", terms_item_style))

    story.append(Spacer(1, 10))
    
    # 7. Footer text
    story.append(Paragraph("For any further queries, please feel free to contact us. We value your business association.", footer_text_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Thanking you,", footer_text_style))
    story.append(Spacer(1, 15))
    
    # Signature line
    sig_data = [
        [
            Paragraph("<b>For Noble Air Gases</b>", left_cell_bold_style),
            Paragraph("", cell_style)
        ],
        [
            Paragraph("<br/><br/>Authorized Signatory", left_cell_style),
            Paragraph("", cell_style)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[270, 270])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buffer.seek(0)
    
    filename = f"Commercial_Offer_{customer_name.replace(' ', '_')}_{q_no.replace('/', '_')}.pdf"
    from flask import send_file
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")



# ================================================================
#  SALES ORDER CREATOR ROUTES
# ================================================================

@app.route('/admin/orders/new')
@admin_required
def admin_orders_new():
    import random
    order_no = f"NAG/ORD/26-27/{random.randint(3000, 9999)}"
    today = date.today()
    day = today.day
    suffix = "th" if 4 <= day <= 20 or 24 <= day <= 30 else ["st", "nd", "rd"][day % 10 - 1]
    date_str = f"{day}{suffix} {today.strftime('%B %Y')}"

    customers = get_all_customer_info()
    
    default_terms = [
        {"label": "1. Delivery", "value": "Within 2-3 days"},
        {"label": "2. GST", "value": "@18% Extra."},
        {"label": "3. Transportation", "value": "Inclusive Delivery."},
        {"label": "4. Payment", "value": "30 Days"},
        {"label": "5. Valve Damage", "value": "Rs. 1000/- per Valve"},
        {"label": "6. Cylinder Lost / Damage", "value": "Rs. 10,000/- Per Cylinder"}
    ]

    return render_template('order_form.html',
        user=session['user'],
        customers=customers,
        order_no=order_no,
        date_str=date_str,
        terms_list=default_terms
    )


@app.route('/admin/orders/generate', methods=['POST'])
@admin_required
def admin_orders_generate():
    customer_name = request.form.get('customer_name', '').strip()
    if customer_name == 'custom':
        customer_name = request.form.get('custom_customer_name', 'Customer').strip()
    
    order_no = request.form.get('order_no', '').strip()
    attn = request.form.get('attn', '').strip()
    tel = request.form.get('tel', '').strip()
    o_date = request.form.get('date', '').strip()
    ref = request.form.get('ref', '').strip()

    # Product list
    product_names = request.form.getlist('product_name[]')
    product_contents = request.form.getlist('product_content[]')
    product_qtys = request.form.getlist('product_qty[]')

    # PDF generation
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []

    styles = getSampleStyleSheet()

    blue_brand = colors.HexColor('#0c5ca8')
    green_brand = colors.HexColor('#3fb549')
    dark_gray = colors.HexColor('#2c2c2a')

    brand_style1 = ParagraphStyle('Brand1', fontName='Helvetica-Bold', fontSize=26, leading=30, textColor=blue_brand, alignment=1)
    brand_style2 = ParagraphStyle('Brand2', fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=green_brand, alignment=1)
    address_style = ParagraphStyle('Address', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=dark_gray, alignment=1)
    header_contact_style = ParagraphStyle('HeaderContact', fontName='Helvetica-Bold', fontSize=9.5, leading=11.5, textColor=dark_gray, alignment=1)
    title_style = ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=colors.black, alignment=1, spaceAfter=6)
    intro_style = ParagraphStyle('Intro', fontName='Helvetica-Bold', fontSize=9.5, leading=12, textColor=colors.HexColor('#1D9E75'), alignment=1, spaceBefore=4, spaceAfter=6)
    
    cell_style = ParagraphStyle('Cell', fontName='Helvetica', fontSize=9.5, leading=12, textColor=colors.black, alignment=1)
    cell_bold_style = ParagraphStyle('CellBold', fontName='Helvetica-Bold', fontSize=10, leading=12.5, textColor=colors.black, alignment=1)
    left_cell_style = ParagraphStyle('LeftCell', fontName='Helvetica', fontSize=9.5, leading=12, textColor=colors.black, alignment=0)
    left_cell_bold_style = ParagraphStyle('LeftCellBold', fontName='Helvetica-Bold', fontSize=10, leading=12.5, textColor=colors.black, alignment=0)
    
    footer_text_style = ParagraphStyle('FooterText', fontName='Helvetica', fontSize=9.5, leading=14, textColor=colors.black, alignment=0)

    # 1. Noble Air Gases Header Layout
    import os as _os
    from reportlab.platypus import Image as RLImage

    logo_path = _os.path.join(_os.path.dirname(__file__), 'static', 'img', 'noble_logo.png')

    if _os.path.exists(logo_path):
        logo_img = RLImage(logo_path, width=240, height=60, kind='proportional')
        logo_img.hAlign = 'CENTER'
        story.append(logo_img)
        story.append(Spacer(1, 8))
    else:
        story.append(Paragraph("NOBLE", brand_style1))
        story.append(Paragraph("air gases", brand_style2))
        story.append(Spacer(1, 4))

    story.append(Paragraph("Plot No. A/12, MIDC Waluj, Chhatrapati Sambhajinagar", address_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph("Email: sales@nobleairgases.com &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Mobile: +91 9225309555", header_contact_style))
    story.append(Spacer(1, 8))

    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#d8d9d4'), spaceAfter=8))

    # 2. Document Title
    story.append(Paragraph("SALES ORDER", title_style))

    # 3. Metadata Table
    meta_data = [
        [
            Paragraph("<b>To</b>", left_cell_bold_style),
            Paragraph(f": M/s {customer_name}", left_cell_style),
            Paragraph("<b>K.Attn</b>", left_cell_bold_style),
            Paragraph(f": {attn}", left_cell_style)
        ],
        [
            Paragraph("<b>Tel</b>", left_cell_bold_style),
            Paragraph(f": {tel}", left_cell_style),
            Paragraph("<b>Date</b>", left_cell_bold_style),
            Paragraph(f": {o_date}", left_cell_style)
        ],
        [
            Paragraph("<b>Order No.</b>", left_cell_bold_style),
            Paragraph(f": {order_no}", left_cell_style),
            Paragraph("<b>Ref</b>", left_cell_bold_style),
            Paragraph(f": {ref}", left_cell_style)
        ]
    ]
    meta_table = Table(meta_data, colWidths=[55, 215, 50, 220])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#888888')),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 10))

    # 4. Intro text
    story.append(Paragraph("We are pleased to confirm registration of the following order placed with us. Product description and quantities are outlined below:", intro_style))

    # 5. Products Table
    table_data = [
        [
            Paragraph("<b>No.</b>", cell_bold_style),
            Paragraph("<b>Product Description</b>", cell_bold_style),
            Paragraph("<b>Content</b>", cell_bold_style),
            Paragraph("<b>Quantity (Qty)</b>", cell_bold_style)
        ]
    ]

    idx_no = 1
    for name, content, qty_str in zip(product_names, product_contents, product_qtys):
        if not name.strip(): continue
        
        try:
            qty = int(qty_str) if qty_str.strip() else 0
        except ValueError:
            qty = 0

        table_data.append([
            Paragraph(f"{idx_no:02d}", cell_style),
            Paragraph(f"<b>{name.strip()}</b>", left_cell_bold_style),
            Paragraph(content.strip(), cell_style),
            Paragraph(str(qty), cell_style)
        ])
        idx_no += 1

    prod_table = Table(table_data, colWidths=[40, 280, 120, 100])
    prod_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d8d9d4')),
        ('LINEBELOW', (0,0), (-1,0), 1.5, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,0), (-1,0), colors.white),
    ]))
    story.append(prod_table)
    story.append(Spacer(1, 20))

    # 7. Footer
    story.append(Paragraph("For any further queries, please feel free to contact us. We value your business association.", footer_text_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Thanking you,", footer_text_style))
    story.append(Spacer(1, 20))

    # Signature line
    sig_data = [
        [
            Paragraph("<b>For Noble Air Gases</b>", left_cell_bold_style),
            Paragraph("", cell_style)
        ],
        [
            Paragraph("<br/><br/>Authorized Signatory", left_cell_style),
            Paragraph("", cell_style)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[270, 270])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(sig_table)

    doc.build(story)
    buffer.seek(0)

    filename = f"Sales_Order_{customer_name.replace(' ', '_')}_{order_no.replace('/', '_')}.pdf"
    from flask import send_file
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")



@app.route('/')
@login_required
def home():
    user = session.get('user')
    if user and user.get('role') in ['manager', 'owner']:
        return redirect('/admin')
    return redirect('/scan')

@app.route('/scan')
@login_required
def scan_app():
    user = session.get('user')
    customers = get_customer_names()
    return render_template('scan.html', user=user, customers=customers)

@app.route('/submit', methods=['POST'])
def submit():
    global cyl_ws
    now = datetime.now()
    rows_to_append = []

    # Check and add headers in Sheet 1 if missing (Customer Col F, Gas Type Col G)
    try:
        if scan_ws is not None:
            headers = scan_ws.row_values(1)
            if len(headers) < 6:
                scan_ws.update_cell(1, 6, "Customer")
            if len(headers) < 7:
                scan_ws.update_cell(1, 7, "Gas Type")
    except Exception as e:
        print("Error checking/updating Sheet1 headers:", e)

    parsed_scans = []
    driver = ""

    # Support JSON payload (modern split-actions submission)
    if request.is_json:
        data = request.get_json()
        driver = data.get('driver', '').strip()
        customer = data.get('customer', '').strip()
        scans = data.get('scans', [])
        
        for s in scans:
            uid = s.get('uid', '').strip()
            action = s.get('action', '').strip()
            if uid and action:
                cust_val = customer if action in ['Delivery', 'Collection'] else ''
                parsed_scans.append({
                    'uid': uid,
                    'action': action,
                    'cust_val': cust_val
                })
    else:
        # Fallback to standard form-data
        action    = request.form['action']
        driver    = request.form['driver']
        customer  = request.form.get('customer', '').strip()
        cylinders = request.form.getlist('cylinders')
    
        for uid in cylinders:
            uid = uid.strip()
            if uid:
                cust_val = customer if action in ['Delivery', 'Collection'] else ''
                parsed_scans.append({
                    'uid': uid,
                    'action': action,
                    'cust_val': cust_val
                })

    # Resolve gas types for all scanned UIDs in a single batch lookup
    uids_to_lookup = list(set(s['uid'].strip().upper() for s in parsed_scans))
    cyl_gas_map = {}
    if uids_to_lookup:
        try:
            if os.environ.get('DATABASE_URL'):
                cyls = Cylinder.query.filter(Cylinder.uid.in_(uids_to_lookup)).all()
                for c in cyls:
                    cyl_gas_map[c.uid.strip().upper()] = c.gas_type or ''
            # Fallback to Sheets for missing UIDs
            missing_uids = [u for u in uids_to_lookup if u not in cyl_gas_map]
            if missing_uids:
                sheet_cyls = get_all_cylinders()
                for c in sheet_cyls:
                    uid_up = c['uid'].strip().upper()
                    if uid_up in missing_uids:
                        cyl_gas_map[uid_up] = c.get('gas_type', '')
        except Exception as ge:
            print("[scan] Error looking up gas types for submit scan batch:", ge)

    # Format rows for Sheets append
    for s in parsed_scans:
        uid = s['uid']
        gas_val = cyl_gas_map.get(uid.strip().upper(), '')
        rows_to_append.append([
            now.strftime('%d-%m-%Y'),
            now.strftime('%H:%M:%S'),
            driver,
            s['action'],
            uid,
            s['cust_val'],
            gas_val
        ])
    
    # ── VALIDATE SCANS ──────────────────────────────────────────
    if os.environ.get('DATABASE_URL'):
        # Database-first validation: Check if we are collecting a cylinder that is already in stock/empty at the Depot
        for row_data in rows_to_append:
            scan_action = row_data[3]
            scan_uid    = row_data[4]
            if scan_action == 'Collection':
                c_db = Cylinder.query.filter(Cylinder.uid.ilike(scan_uid)).first()
                if c_db:
                    if c_db.status in ['Empty', 'Filled'] or c_db.location == 'Depot':
                        return f"Validation Error: Cylinder '{scan_uid}' is already at the Depot (status: {c_db.status or 'Empty'}). Cannot collect twice.", 400
    else:
        # Fallback validation using Google Sheets
        if cyl_ws is None and doc:
            try:
                cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
            except Exception:
                cyl_ws = None
        if cyl_ws:
            try:
                cyl_rows = cyl_ws.get_all_values()
                cyl_status_map = {}
                for idx, r in enumerate(cyl_rows):
                    if idx == 0:
                        continue
                    if r and r[0].strip():
                        # Col E = status (index 4), Col F = location (index 5)
                        status   = r[4].strip() if len(r) > 4 else 'Active'
                        location = r[5].strip() if len(r) > 5 else 'Depot'
                        cyl_status_map[r[0].strip().upper()] = (status, location)
                
                for row_data in rows_to_append:
                    scan_action = row_data[3]
                    scan_uid    = row_data[4].strip().upper()
                    if scan_action == 'Collection' and scan_uid in cyl_status_map:
                        status, location = cyl_status_map[scan_uid]
                        if status in ['Empty', 'Filled'] or location == 'Depot':
                            return f"Validation Error: Cylinder '{row_data[4]}' is already at the Depot (status: {status}). Cannot collect twice.", 400
            except Exception as se:
                print("[validation] Error validating against Google Sheets registry:", se)
    # ────────────────────────────────────────────────────────────

    db_written = False
    # Write to database (PostgreSQL) first
    if os.environ.get('DATABASE_URL'):
        try:
            with db.session.no_autoflush:
                for row_data in rows_to_append:
                    s_date = row_data[0]
                    s_time = row_data[1]
                    scan_driver = row_data[2]
                    scan_action = row_data[3]
                    scan_uid = row_data[4]
                    scan_cust = row_data[5] if len(row_data) > 5 else ''
                    scan_gas = row_data[6] if len(row_data) > 6 else ''
                    
                    scan_db = Scan(
                        scan_date=s_date,
                        scan_time=s_time,
                        driver=scan_driver,
                        action=scan_action,
                        cylinder_uid=scan_uid,
                        customer=scan_cust,
                        gas_type=scan_gas
                    )
                    db.session.add(scan_db)
                    
                    c_db = Cylinder.query.filter(Cylinder.uid.ilike(scan_uid)).first()
                    if c_db:
                        if scan_action == 'Delivery':
                            c_db.status = 'Delivered'
                            c_db.location = scan_cust or 'Customer'
                        elif scan_action == 'Collection':
                            c_db.status = 'Empty'
                            c_db.location = 'Depot'
                        elif scan_action == 'Filling':
                            c_db.status = 'Filled'
                            c_db.location = 'Depot'
                        c_db.last_activity_date = s_date
            db.session.commit()
            db_written = True
            print(f"[db] Logged {len(rows_to_append)} scans and updated cylinder registries in DB.")
        except Exception as dbe:
            db.session.rollback()
            print("[db] Error writing scans to DB:", dbe)
            from sqlalchemy.exc import OperationalError, InterfaceError
            is_connection_error = isinstance(dbe, (OperationalError, InterfaceError)) or "connection" in str(dbe).lower()
            if not is_connection_error:
                return f"Database error logging scans: {str(dbe)}", 500

    # Mirror to Sheets in background
    def background_mirror_scans():
        try:
            global sheet, cyl_ws, doc
            if sheet is None and doc:
                try:
                    sheet = doc.worksheet(SCAN_SHEET_NAME)
                except Exception:
                    pass
            if cyl_ws is None and doc:
                try:
                    cyl_ws = doc.worksheet(CYLINDER_SHEET_NAME)
                except Exception:
                    pass

            if sheet and rows_to_append:
                sheets_write_with_retry(sheet.append_rows, rows_to_append)

            if cyl_ws and rows_to_append:
                cyl_rows = cyl_ws.get_all_values()
                # Build UID → row index map
                uid_row_map = {}
                for idx, r in enumerate(cyl_rows):
                    if idx == 0:
                        continue
                    if r and r[0].strip():
                        uid_row_map[r[0].strip().upper()] = idx + 1

                today_str = datetime.now().strftime('%d-%m-%Y')
                batch_data = []
                for row_data in rows_to_append:
                    scan_uid    = row_data[4].strip().upper()
                    scan_action = row_data[3].strip()
                    scan_cust   = row_data[5].strip() if len(row_data) > 5 else ''
                    row_num = uid_row_map.get(scan_uid)
                    if not row_num:
                        continue  # UID not in registry — skip silently
                    if scan_action == 'Delivery':
                        new_status   = 'Delivered'
                        new_location = scan_cust or 'Customer'
                    elif scan_action == 'Collection':
                        new_status   = 'Empty'
                        new_location = 'Depot'
                    elif scan_action == 'Filling':
                        new_status   = 'Filled'
                        new_location = 'Depot'
                    else:
                        continue
                    batch_data.append({
                        'range': f'E{row_num}:G{row_num}',
                        'values': [[new_status, new_location, today_str]]
                    })
                if batch_data:
                    sheets_write_with_retry(cyl_ws.batch_update, batch_data)
        except Exception as se:
            print("[sheets] Error mirroring scans to Google Sheets in background:", se)

    async_sheets_write(background_mirror_scans)

    clear_cache()
    return f"{len(parsed_scans)} cylinders saved successfully"

@app.route('/manager')
def manager():
    wb    = load_workbook('data.xlsx')
    ws    = wb['Scan_Log']
    scans = list(ws.iter_rows(min_row=2, values_only=True))
    scans.reverse()
    return render_template('manager.html', scans=scans)

@app.route('/customers')
def customers():
    wb        = load_workbook('data.xlsx')
    ws        = wb['Customers']
    customers = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('customers.html', customers=customers)

@app.route('/add_customer', methods=['POST'])
def add_customer():
    customer = request.form['customer']
    email    = request.form['email']
    phone    = request.form['phone']
    wb       = load_workbook('data.xlsx')
    ws       = wb['Customers']
    ws.append([customer, email, phone])
    wb.save('data.xlsx')
    return redirect('/customers')


# ================================================================
#  INVENTORY & BULK TANKS ROUTES
# ================================================================

@app.route('/admin/inventory')
@admin_required
def admin_inventory():
    selected_date = request.args.get('date', '')
    parsed = parse_date(selected_date) if selected_date else date.today()
    target_date_str = parsed.strftime('%d-%m-%Y')
    target_date_iso = parsed.strftime('%Y-%m-%d')
    
    t1 = calculate_table1_filled_inventory()
    t2 = calculate_table2_bulk_inventory(target_date_str)
    dispatch_report = calculate_daily_dispatch_report(target_date_str)
    
    hydro_alerts = [c for c in merge_cylinder_data() if c.get('hydro_badge') in ('Overdue', 'Due Soon')]
    hydro_alerts.sort(key=lambda x: parse_date(x.get('next_hydro_due')) or date.max)
    
    return render_template('inventory.html',
        user=session['user'],
        target_date=target_date_str,
        target_date_iso=target_date_iso,
        t1=t1,
        t2=t2,
        dispatch_report=dispatch_report,
        hydro_alerts=hydro_alerts
    )

@app.route('/admin/tanks/update', methods=['POST'])
@admin_required
def admin_tanks_update():
    target_date = request.form.get('date', date.today().strftime('%d-%m-%Y'))
    
    gases = ['Argon', 'CO2', 'N2', 'Oxygen']
    success = True
    for g in gases:
        opening = request.form.get(f'{g}_opening', '0').strip() or '0'
        capacity = request.form.get(f'{g}_capacity', '0').strip() or '0'
        dead_volume = request.form.get(f'{g}_dead', '0').strip() or '0'
        unit = 'KG' if g == 'CO2' else 'Cum'
        
        ok = update_tank_opening_stock(target_date, g, opening, capacity, dead_volume, unit)
        if not ok:
            success = False
            
    if success:
        clear_cache()
        return redirect(f'/admin/inventory?date={target_date}')
    else:
        return "Error updating some tank levels. Please verify your sheets connection.", 500

@app.route('/admin/inventory/export/excel')
@admin_required
def export_excel():
    target_date = request.args.get('date', date.today().strftime('%d-%m-%Y'))
    
    t1 = calculate_table1_filled_inventory()
    t2 = calculate_table2_bulk_inventory(target_date)
    dr = calculate_daily_dispatch_report(target_date)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # ── Tab 1: Inventory Report ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Inventory Report"
    ws1.views.sheetView[0].showGridLines = True
    
    title_font = Font(name='Arial', size=16, bold=True, color='0F6E56')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    sub_font = Font(name='Arial', size=11, bold=True, color='0F6E56')
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    
    green_fill = PatternFill(start_color='0F6E56', end_color='0F6E56', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    orange_fill = PatternFill(start_color='C25E3B', end_color='C25E3B', fill_type='solid')
    gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    ws1['A1'] = "Daily Dispatch & Tank Status Report"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Report Date: {target_date}"
    ws1['A2'].font = Font(name='Arial', size=11, italic=True)
    
    ws1['A4'] = "Cyl Status Dispatch Stock (Filled Cylinder Inventory)"
    ws1['A4'].font = sub_font
    
    t1_headers = ["Product", "Full Feeling Cyl", "Gas per Cyl", "Total Gas", "Unit"]
    for col_idx, h in enumerate(t1_headers, 1):
        cell = ws1.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='center')
        
    row_idx = 6
    for row in t1['rows']:
        ws1.cell(row=row_idx, column=1, value=row['name']).font = normal_font
        ws1.cell(row=row_idx, column=2, value=row['filled_count']).font = normal_font
        ws1.cell(row=row_idx, column=3, value=row['gas_per_cyl']).font = normal_font
        ws1.cell(row=row_idx, column=4, value=row['total_gas']).font = normal_font
        ws1.cell(row=row_idx, column=5, value=row['unit']).font = normal_font
        
        for col_idx in range(1, 6):
            ws1.cell(row=row_idx, column=col_idx).border = border_thin
            
        row_idx += 1
        
    total_cell = ws1.cell(row=row_idx, column=1, value="TOTAL FILLED")
    total_cell.font = bold_font
    total_cell.fill = yellow_fill
    
    cnt_cell = ws1.cell(row=row_idx, column=2, value=t1['total_filled'])
    cnt_cell.font = bold_font
    cnt_cell.fill = yellow_fill
    ws1.cell(row=row_idx, column=3, value="").fill = yellow_fill
    
    gas_tot_cell = ws1.cell(row=row_idx, column=4, value=f"{t1['total_cum']} Cum + {t1['total_kg']} KG")
    gas_tot_cell.font = bold_font
    gas_tot_cell.fill = yellow_fill
    
    ws1.cell(row=row_idx, column=5, value="Mixed").font = bold_font
    ws1.cell(row=row_idx, column=5).fill = yellow_fill
    
    for col_idx in range(1, 6):
        ws1.cell(row=row_idx, column=col_idx).border = border_thin
        
    row_idx += 3
    
    ws1.cell(row=row_idx, column=1, value="Total USED Stock (Bulk Tank Gas Inventory)").font = sub_font
    row_idx += 1
    
    t2_headers = ["Gas", "Opening Stock", "Used Today", "Closing Stock", "Fleet Size", "Usable Stock", "Unit"]
    for col_idx, h in enumerate(t2_headers, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = orange_fill
        cell.alignment = Alignment(horizontal='center')
        
    row_idx += 1
    for row in t2:
        ws1.cell(row=row_idx, column=1, value=row['gas']).font = normal_font
        ws1.cell(row=row_idx, column=2, value=row['opening']).font = normal_font
        ws1.cell(row=row_idx, column=3, value=row['used_today']).font = normal_font
        ws1.cell(row=row_idx, column=4, value=row['closing']).font = normal_font
        ws1.cell(row=row_idx, column=5, value=row['fleet']).font = normal_font
        ws1.cell(row=row_idx, column=6, value=row['usable']).font = normal_font
        ws1.cell(row=row_idx, column=7, value=row['unit']).font = normal_font
        
        for col_idx in range(1, 8):
            ws1.cell(row=row_idx, column=col_idx).border = border_thin
            
        row_idx += 1
        
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ── Tab 2: Dispatch Report ──────────────────────────────────
    ws2 = wb.create_sheet(title="Dispatch Report")
    ws2.views.sheetView[0].showGridLines = True
    
    def style_range(ws, cell_range, font=None, fill=None, border=None, alignment=None):
        for row in ws[cell_range]:
            for cell in row:
                if font: cell.font = font
                if fill: cell.fill = fill
                if border: cell.border = border
                if alignment: cell.alignment = alignment
                
    ws2['A1'] = "Daily Dispatch Report"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Report Date: {target_date}"
    ws2['A2'].font = Font(name='Arial', size=11, italic=True)
    
    ws2.merge_cells('A4:A5')
    ws2['A4'] = "Today Dispatch & Empty Collection Party Name"
    ws2['A4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2['A4'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    ws2['A4'].fill = green_fill
    
    ws2.merge_cells('B4:I4')
    ws2['B4'] = "Today Dispatch Cyld."
    ws2['B4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['B4'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    ws2['B4'].fill = green_fill
    
    ws2.merge_cells('J4:Q4')
    ws2['J4'] = "Today Empty Collection Cyld."
    ws2['J4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['J4'].font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    ws2['J4'].fill = green_fill
    
    cols = ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']
    for idx, col_name in enumerate(cols):
        c1 = ws2.cell(row=5, column=2 + idx, value=col_name)
        c1.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c1.fill = green_fill
        c1.alignment = Alignment(horizontal='center')
        
        c2 = ws2.cell(row=5, column=10 + idx, value=col_name)
        c2.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        c2.fill = green_fill
        c2.alignment = Alignment(horizontal='center')
        
    for r_idx in (4, 5):
        for c_idx in range(1, 18):
            ws2.cell(row=r_idx, column=c_idx).border = border_thin
            
    row_idx = 6
    if dr.get('company_rows'):
        for row in dr['company_rows']:
            ws2.cell(row=row_idx, column=1, value=row['customer']).font = normal_font
            for col_idx, k in enumerate(cols):
                v1 = row['dispatch'][k]
                ws2.cell(row=row_idx, column=2 + col_idx, value=v1 if v1 != 0 else '').font = normal_font
                ws2.cell(row=row_idx, column=2 + col_idx).alignment = Alignment(horizontal='center')
                
                v2 = row['collection'][k]
                ws2.cell(row=row_idx, column=10 + col_idx, value=v2 if v2 != 0 else '').font = normal_font
                ws2.cell(row=row_idx, column=10 + col_idx).alignment = Alignment(horizontal='center')
            for c_idx in range(1, 18):
                ws2.cell(row=row_idx, column=c_idx).border = border_thin
            row_idx += 1
            
    c_tot_cell = ws2.cell(row=row_idx, column=1, value="Dispatch Cyld")
    c_tot_cell.font = bold_font
    c_tot_cell.fill = yellow_fill
    
    comp_t = dr.get('company_totals', {'dispatch': {}, 'collection': {}, 'dispatch_total': 0, 'collection_total': 0})
    for col_idx, k in enumerate(cols):
        v1 = comp_t['dispatch'].get(k, 0)
        ws2.cell(row=row_idx, column=2 + col_idx, value=v1 if v1 != 0 else '').font = bold_font
        ws2.cell(row=row_idx, column=2 + col_idx).fill = yellow_fill
        ws2.cell(row=row_idx, column=2 + col_idx).alignment = Alignment(horizontal='center')
        
        v2 = comp_t['collection'].get(k, 0)
        ws2.cell(row=row_idx, column=10 + col_idx, value=v2 if v2 != 0 else '').font = bold_font
        ws2.cell(row=row_idx, column=10 + col_idx).fill = yellow_fill
        ws2.cell(row=row_idx, column=10 + col_idx).alignment = Alignment(horizontal='center')
        
    for c_idx in range(1, 18):
        ws2.cell(row=row_idx, column=c_idx).border = border_thin
    row_idx += 1
    
    ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    ws2.cell(row=row_idx, column=1, value=f"Today Dispatch Cyld: {comp_t['dispatch_total']}").font = bold_font
    ws2.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
    style_range(ws2, f'A{row_idx}:I{row_idx}', fill=yellow_fill)
    
    ws2.merge_cells(start_row=row_idx, start_column=10, end_row=row_idx, end_column=17)
    ws2.cell(row=row_idx, column=10, value=f"Today Empty Collection Cyld: {comp_t['collection_total']}").font = bold_font
    ws2.cell(row=row_idx, column=10).alignment = Alignment(horizontal='center')
    style_range(ws2, f'J{row_idx}:Q{row_idx}', fill=yellow_fill)
    
    for c_idx in range(1, 18):
        ws2.cell(row=row_idx, column=c_idx).border = border_thin
    row_idx += 1
    
    # Party Header Row
    header_party_row = row_idx
    ws2.merge_cells(start_row=header_party_row, start_column=1, end_row=header_party_row, end_column=9)
    ws2.cell(row=header_party_row, column=1, value="Party Name & Today Dispatch Party Cyld").font = bold_font
    style_range(ws2, f'A{header_party_row}:I{header_party_row}', fill=gray_fill)
    
    ws2.merge_cells(start_row=header_party_row, start_column=10, end_row=header_party_row, end_column=17)
    ws2.cell(row=header_party_row, column=10, value="").fill = gray_fill
    
    for c_idx in range(1, 18):
        ws2.cell(row=header_party_row, column=c_idx).border = border_thin
    row_idx += 1
    
    if dr.get('party_rows'):
        for row in dr['party_rows']:
            ws2.cell(row=row_idx, column=1, value=row['customer']).font = normal_font
            for col_idx, k in enumerate(cols):
                v1 = row['dispatch'][k]
                ws2.cell(row=row_idx, column=2 + col_idx, value=v1 if v1 != 0 else '').font = normal_font
                ws2.cell(row=row_idx, column=2 + col_idx).alignment = Alignment(horizontal='center')
                
                v2 = row['collection'][k]
                ws2.cell(row=row_idx, column=10 + col_idx, value=v2 if v2 != 0 else '').font = normal_font
                ws2.cell(row=row_idx, column=10 + col_idx).alignment = Alignment(horizontal='center')
            for c_idx in range(1, 18):
                ws2.cell(row=row_idx, column=c_idx).border = border_thin
            row_idx += 1
            
    p_tot_cell = ws2.cell(row=row_idx, column=1, value="Total")
    p_tot_cell.font = bold_font
    p_tot_cell.fill = gray_fill
    
    party_t = dr.get('party_totals', {'dispatch': {}, 'collection': {}, 'dispatch_total': 0, 'collection_total': 0})
    for col_idx, k in enumerate(cols):
        v1 = party_t['dispatch'].get(k, 0)
        ws2.cell(row=row_idx, column=2 + col_idx, value=v1 if v1 != 0 else '').font = bold_font
        ws2.cell(row=row_idx, column=2 + col_idx).fill = gray_fill
        ws2.cell(row=row_idx, column=2 + col_idx).alignment = Alignment(horizontal='center')
        
        v2 = party_t['collection'].get(k, 0)
        ws2.cell(row=row_idx, column=10 + col_idx, value=v2 if v2 != 0 else '').font = bold_font
        ws2.cell(row=row_idx, column=10 + col_idx).fill = gray_fill
        ws2.cell(row=row_idx, column=10 + col_idx).alignment = Alignment(horizontal='center')
        
    for c_idx in range(1, 18):
        ws2.cell(row=row_idx, column=c_idx).border = border_thin
    row_idx += 1
    
    ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    ws2.cell(row=row_idx, column=1, value=f"Today Party Cyld Dispatch Total: {party_t['dispatch_total']}").font = bold_font
    ws2.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
    style_range(ws2, f'A{row_idx}:I{row_idx}', fill=yellow_fill)
    
    ws2.merge_cells(start_row=row_idx, start_column=10, end_row=row_idx, end_column=17)
    ws2.cell(row=row_idx, column=10, value=f"Today Party Cyld Collection Total: {party_t['collection_total']}").font = bold_font
    ws2.cell(row=row_idx, column=10).alignment = Alignment(horizontal='center')
    style_range(ws2, f'J{row_idx}:Q{row_idx}', fill=yellow_fill)
    
    for c_idx in range(1, 18):
        ws2.cell(row=row_idx, column=c_idx).border = border_thin
    row_idx += 1
    
    # Grand Totals
    grand_t = dr.get('grand_totals', {'dispatch_total': 0, 'collection_total': 0})
    ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    ws2.cell(row=row_idx, column=1, value=f"Today Total Dispatch Cylinders: {grand_t['dispatch_total']}").font = Font(name='Arial', size=11, bold=True)
    ws2.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
    style_range(ws2, f'A{row_idx}:I{row_idx}', fill=orange_fill)
    
    ws2.merge_cells(start_row=row_idx, start_column=10, end_row=row_idx, end_column=17)
    ws2.cell(row=row_idx, column=10, value=f"Today Total Collection Cylinders: {grand_t['collection_total']}").font = Font(name='Arial', size=11, bold=True)
    ws2.cell(row=row_idx, column=10).alignment = Alignment(horizontal='center')
    style_range(ws2, f'J{row_idx}:Q{row_idx}', fill=orange_fill)
    
    border_double = Border(
        top=Side(style='medium', color='000000'),
        bottom=Side(style='double', color='000000'),
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD')
    )
    for c_idx in range(1, 18):
        ws2.cell(row=row_idx, column=c_idx).border = border_double
        
    ws2.column_dimensions['A'].width = 32
    for c_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']:
        ws2.column_dimensions[c_letter].width = 8
        
    from io import BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"Inventory_Report_{target_date}.xlsx"
    from flask import send_file
    return send_file(out, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/inventory/export/pdf')
@admin_required
def export_pdf():
    target_date = request.args.get('date', date.today().strftime('%d-%m-%Y'))
    
    t1 = calculate_table1_filled_inventory()
    t2 = calculate_table2_bulk_inventory(target_date)
    dr = calculate_daily_dispatch_report(target_date)
    
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    styles = getSampleStyleSheet()
    primary_color = colors.HexColor('#0F6E56')
    dark_text = colors.HexColor('#2C2C2A')
    orange_color = colors.HexColor('#C25E3B')
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=primary_color,
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=dark_text,
        spaceAfter=12
    )
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=primary_color,
        spaceBefore=10,
        spaceAfter=6
    )
    
    # Page 1: Stock Status & Bulk Tanks
    story.append(Paragraph("Cylinder Tracker — Daily Inventory & Tanks Report", title_style))
    story.append(Paragraph(f"Report Generated for Date: {target_date}", subtitle_style))
    
    story.append(Paragraph("Cyl Status Dispatch Stock (Filled Cylinder Inventory)", section_style))
    
    t1_data = [["Product", "Full Feeling Cyl", "Gas/Cyl", "Total Gas", "Unit"]]
    for row in t1['rows']:
        t1_data.append([
            row['name'],
            str(row['filled_count']),
            f"{row['gas_per_cyl']:.4f}".rstrip('0').rstrip('.'),
            f"{row['total_gas']:.4f}".rstrip('0').rstrip('.'),
            row['unit']
        ])
    t1_data.append([
        "TOTAL FILLED",
        str(t1['total_filled']),
        "",
        f"{t1['total_cum']} Cum + {t1['total_kg']} KG",
        "Mixed"
    ])
    
    table1 = Table(t1_data, colWidths=[150, 100, 100, 120, 60])
    table1.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('ALIGN', (1,1), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,1), (-1,-2), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-2), 9),
        ('GRID', (0,0), (-1,-2), 0.5, colors.lightgrey),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#FFFFCC')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,-1), (-1,-1), 9),
        ('GRID', (0,-1), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,-1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,-1), (-1,-1), 5),
    ]))
    story.append(table1)
    
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Total USED Stock (Bulk Tank Gas Inventory)", section_style))
    
    t2_data = [["Gas", "Opening Stock", "Used Today", "Closing Stock", "Fleet Size", "Usable Stock", "Unit"]]
    for row in t2:
        t2_data.append([
            row['gas'],
            f"{row['opening']:.2f}",
            f"{row['used_today']:.2f}",
            f"{row['closing']:.2f}",
            str(row['fleet']),
            f"{row['usable']:.2f}",
            row['unit']
        ])
        
    table2 = Table(t2_data, colWidths=[90, 85, 75, 85, 65, 85, 45])
    table2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), orange_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('ALIGN', (1,1), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FAF5F3')])
    ]))
    story.append(table2)
    
    # Page 2: Daily Dispatch Report Matrix
    story.append(PageBreak())
    story.append(Paragraph("Cylinder Tracker — Daily Dispatch Report", title_style))
    story.append(Paragraph(f"Report Generated for Date: {target_date}", subtitle_style))
    story.append(Paragraph("Today Dispatch & Empty Collection Party-wise Matrix", section_style))
    
    dp_data = []
    dp_data.append(["Today Dispatch & Empty Collection Party Name", "Today Dispatch Cyld.", "", "", "", "", "", "", "", "Today Empty Collection Cyld.", "", "", "", "", "", "", ""])
    dp_data.append(["", "ACM", "ARG", "CO2", "N2", "Oxy", "Hel", "DA", "Dur", "ACM", "ARG", "CO2", "N2", "Oxy", "Hel", "DA", "Dur"])
    
    cols = ['ACM', 'ARG', 'CO2', 'N2', 'Oxy', 'Helium', 'DA', 'Dura']
    
    # 1. Company Rows
    if dr.get('company_rows'):
        for r in dr['company_rows']:
            dp_data.append([
                r['customer'],
                str(r['dispatch']['ACM'] or ''),
                str(r['dispatch']['ARG'] or ''),
                str(r['dispatch']['CO2'] or ''),
                str(r['dispatch']['N2'] or ''),
                str(r['dispatch']['Oxy'] or ''),
                str(r['dispatch']['Helium'] or ''),
                str(r['dispatch']['DA'] or ''),
                str(r['dispatch']['Dura'] or ''),
                str(r['collection']['ACM'] or ''),
                str(r['collection']['ARG'] or ''),
                str(r['collection']['CO2'] or ''),
                str(r['collection']['N2'] or ''),
                str(r['collection']['Oxy'] or ''),
                str(r['collection']['Helium'] or ''),
                str(r['collection']['DA'] or ''),
                str(r['collection']['Dura'] or '')
            ])
            
    comp_t = dr.get('company_totals', {'dispatch': {}, 'collection': {}, 'dispatch_total': 0, 'collection_total': 0})
    dp_data.append([
        "Dispatch Cyld",
        str(comp_t['dispatch'].get('ACM', '') or ''),
        str(comp_t['dispatch'].get('ARG', '') or ''),
        str(comp_t['dispatch'].get('CO2', '') or ''),
        str(comp_t['dispatch'].get('N2', '') or ''),
        str(comp_t['dispatch'].get('Oxy', '') or ''),
        str(comp_t['dispatch'].get('Helium', '') or ''),
        str(comp_t['dispatch'].get('DA', '') or ''),
        str(comp_t['dispatch'].get('Dura', '') or ''),
        str(comp_t['collection'].get('ACM', '') or ''),
        str(comp_t['collection'].get('ARG', '') or ''),
        str(comp_t['collection'].get('CO2', '') or ''),
        str(comp_t['collection'].get('N2', '') or ''),
        str(comp_t['collection'].get('Oxy', '') or ''),
        str(comp_t['collection'].get('Helium', '') or ''),
        str(comp_t['collection'].get('DA', '') or ''),
        str(comp_t['collection'].get('Dura', '') or '')
    ])
    
    dp_data.append([
        f"Today Dispatch Cyld: {comp_t['dispatch_total']}", "", "", "", "", "", "", "", "",
        f"Today Empty Collection Cyld: {comp_t['collection_total']}", "", "", "", "", "", "", ""
    ])
    
    # Party Header Row
    dp_data.append(["Party Name & Today Dispatch Party Cyld", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    
    # 2. Party Rows
    if dr.get('party_rows'):
        for r in dr['party_rows']:
            dp_data.append([
                r['customer'],
                str(r['dispatch']['ACM'] or ''),
                str(r['dispatch']['ARG'] or ''),
                str(r['dispatch']['CO2'] or ''),
                str(r['dispatch']['N2'] or ''),
                str(r['dispatch']['Oxy'] or ''),
                str(r['dispatch']['Helium'] or ''),
                str(r['dispatch']['DA'] or ''),
                str(r['dispatch']['Dura'] or ''),
                str(r['collection']['ACM'] or ''),
                str(r['collection']['ARG'] or ''),
                str(r['collection']['CO2'] or ''),
                str(r['collection']['N2'] or ''),
                str(r['collection']['Oxy'] or ''),
                str(r['collection']['Helium'] or ''),
                str(r['collection']['DA'] or ''),
                str(r['collection']['Dura'] or '')
            ])
            
    party_t = dr.get('party_totals', {'dispatch': {}, 'collection': {}, 'dispatch_total': 0, 'collection_total': 0})
    dp_data.append([
        "Total",
        str(party_t['dispatch'].get('ACM', '') or ''),
        str(party_t['dispatch'].get('ARG', '') or ''),
        str(party_t['dispatch'].get('CO2', '') or ''),
        str(party_t['dispatch'].get('N2', '') or ''),
        str(party_t['dispatch'].get('Oxy', '') or ''),
        str(party_t['dispatch'].get('Helium', '') or ''),
        str(party_t['dispatch'].get('DA', '') or ''),
        str(party_t['dispatch'].get('Dura', '') or ''),
        str(party_t['collection'].get('ACM', '') or ''),
        str(party_t['collection'].get('ARG', '') or ''),
        str(party_t['collection'].get('CO2', '') or ''),
        str(party_t['collection'].get('N2', '') or ''),
        str(party_t['collection'].get('Oxy', '') or ''),
        str(party_t['collection'].get('Helium', '') or ''),
        str(party_t['collection'].get('DA', '') or ''),
        str(party_t['collection'].get('Dura', '') or '')
    ])
    
    dp_data.append([
        f"Today Party Cyld Dispatch Total: {party_t['dispatch_total']}", "", "", "", "", "", "", "", "",
        f"Today Party Cyld Collection Total: {party_t['collection_total']}", "", "", "", "", "", "", ""
    ])
    
    # Grand Totals
    grand_t = dr.get('grand_totals', {'dispatch_total': 0, 'collection_total': 0})
    dp_data.append([
        f"Today Total Dispatch Cylinders: {grand_t['dispatch_total']}", "", "", "", "", "", "", "", "",
        f"Today Total Collection Cylinders: {grand_t['collection_total']}", "", "", "", "", "", "", ""
    ])
    
    col_widths = [124] + [26] * 16
    
    comp_len = len(dr.get('company_rows', []))
    party_len = len(dr.get('party_rows', []))
    
    comp_sub_idx = 2 + comp_len
    comp_tot_idx = 3 + comp_len
    party_head_idx = 4 + comp_len
    party_sub_idx = 5 + comp_len + party_len
    party_tot_idx = 6 + comp_len + party_len
    grand_tot_idx = 7 + comp_len + party_len
    
    t_style = [
        ('SPAN', (0,0), (0,1)),
        ('SPAN', (1,0), (8,0)),
        ('SPAN', (9,0), (16,0)),
        ('BACKGROUND', (0,0), (-1,1), primary_color),
        ('TEXTCOLOR', (0,0), (-1,1), colors.white),
        ('FONTNAME', (0,0), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,1), 8),
        ('ALIGN', (0,0), (-1,1), 'CENTER'),
        ('VALIGN', (0,0), (-1,1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('ALIGN', (1,2), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,2), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,2), (-1,-1), 7.5),
        ('VALIGN', (0,2), (-1,-1), 'MIDDLE'),
        
        ('SPAN', (0, comp_tot_idx), (8, comp_tot_idx)),
        ('SPAN', (9, comp_tot_idx), (16, comp_tot_idx)),
        ('SPAN', (0, party_head_idx), (8, party_head_idx)),
        ('SPAN', (9, party_head_idx), (16, party_head_idx)),
        ('SPAN', (0, party_tot_idx), (8, party_tot_idx)),
        ('SPAN', (9, party_tot_idx), (16, party_tot_idx)),
        ('SPAN', (0, grand_tot_idx), (8, grand_tot_idx)),
        ('SPAN', (9, grand_tot_idx), (16, grand_tot_idx)),
        
        ('BACKGROUND', (0, comp_sub_idx), (-1, comp_sub_idx), colors.HexColor('#FFFFCC')),
        ('FONTNAME', (0, comp_sub_idx), (-1, comp_sub_idx), 'Helvetica-Bold'),
        
        ('BACKGROUND', (0, comp_tot_idx), (-1, comp_tot_idx), colors.HexColor('#FFFF99')),
        ('FONTNAME', (0, comp_tot_idx), (-1, comp_tot_idx), 'Helvetica-Bold'),
        ('ALIGN', (0, comp_tot_idx), (-1, comp_tot_idx), 'CENTER'),
        
        ('BACKGROUND', (0, party_head_idx), (-1, party_head_idx), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, party_head_idx), (-1, party_head_idx), 'Helvetica-Bold'),
        ('ALIGN', (0, party_head_idx), (-1, party_head_idx), 'CENTER'),
        
        ('BACKGROUND', (0, party_sub_idx), (-1, party_sub_idx), colors.HexColor('#EAEAEA')),
        ('FONTNAME', (0, party_sub_idx), (-1, party_sub_idx), 'Helvetica-Bold'),
        
        ('BACKGROUND', (0, party_tot_idx), (-1, party_tot_idx), colors.HexColor('#FFFF99')),
        ('FONTNAME', (0, party_tot_idx), (-1, party_tot_idx), 'Helvetica-Bold'),
        ('ALIGN', (0, party_tot_idx), (-1, party_tot_idx), 'CENTER'),
        
        ('BACKGROUND', (0, grand_tot_idx), (-1, grand_tot_idx), colors.HexColor('#FFE6CC')),
        ('FONTNAME', (0, grand_tot_idx), (-1, grand_tot_idx), 'Helvetica-Bold'),
        ('FONTSIZE', (0, grand_tot_idx), (-1, grand_tot_idx), 8),
        ('ALIGN', (0, grand_tot_idx), (-1, grand_tot_idx), 'CENTER'),
    ]
    
    table_dispatch = Table(dp_data, colWidths=col_widths)
    table_dispatch.setStyle(TableStyle(t_style))
    story.append(table_dispatch)
    
    doc.build(story)
    buffer.seek(0)
    
    filename = f"Inventory_Report_{target_date}.pdf"
    from flask import send_file
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")
def start_scheduler():
    from sync import sync_sheets_to_db
    scheduler = BackgroundScheduler(daemon=True)
    def run_sync():
        with app.app_context():
            try:
                sync_sheets_to_db(doc)
            except Exception as e:
                print("Scheduler sync error:", e)
    # Run sync job every 5 minutes
    scheduler.add_job(run_sync, 'interval', minutes=5)
    scheduler.start()
    print("Background sync scheduler started successfully.")

# ── Admin Scanning (Internal Logs) ──────────────────────────────────────────────
@app.route('/admin/scanner')
@admin_required
def admin_scanner_page():
    ensure_customer_columns()
    all_customers = [c['name'] for c in get_all_customer_info() if c['name']]
    return render_template('admin_scanner.html', customers=sorted(list(set(all_customers))))

@app.route('/admin/scanner/logs')
@admin_required
def admin_scanner_logs():
    logs = []
    if os.environ.get('DATABASE_URL'):
        from models import AdminScanLog
        logs = AdminScanLog.query.order_by(AdminScanLog.id.desc()).all()
        logs = [log.to_dict() for log in logs]
    return render_template('admin_scan_logs.html', logs=logs)

@app.route('/admin/scanner/submit', methods=['POST'])
@admin_required
def admin_scanner_submit():
    data = request.json
    action = data.get('action')
    customer = data.get('customer')
    uids = data.get('uids', [])
    if not uids or not action:
        return jsonify({'error': 'Missing required fields'}), 400

    admin_name = session.get('user', {}).get('name', 'Admin')
    now = datetime.now()
    d_str = now.strftime('%d-%m-%Y')
    t_str = now.strftime('%H:%M:%S')

    # Resolve gas types for each uid
    cyls_dict = {}
    if os.environ.get('DATABASE_URL'):
        db_cyls = Cylinder.query.filter(Cylinder.uid.in_(uids)).all()
        cyls_dict = {c.uid.upper(): c.gas_type for c in db_cyls}
    else:
        sheet_cyls = get_all_cylinders()
        for c in sheet_cyls:
            if c['uid'].upper() in [u.upper() for u in uids]:
                cyls_dict[c['uid'].upper()] = c.get('gas_type', '')

    rows_to_append = []
    if os.environ.get('DATABASE_URL'):
        from models import AdminScanLog
        try:
            with db.session.no_autoflush:
                for uid in uids:
                    uid_upper = uid.upper()
                    gas = cyls_dict.get(uid_upper, '')
                    log = AdminScanLog(
                        scan_date=d_str,
                        scan_time=t_str,
                        cylinder_uid=uid_upper,
                        gas_type=gas,
                        customer=customer,
                        action=action,
                        admin_name=admin_name
                    )
                    db.session.add(log)
                    rows_to_append.append([d_str, t_str, uid_upper, gas, customer, action, admin_name])
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # Append to Google Sheets 'AdminScans' tab
    def _append_admin_scans():
        try:
            client = gspread.authorize(creds)
            sheet = client.open(SHEET_NAME)
            try:
                ws = sheet.worksheet("AdminScans")
            except gspread.exceptions.WorksheetNotFound:
                ws = sheet.add_worksheet(title="AdminScans", rows=1000, cols=10)
                ws.append_row(["Date", "Time", "UID", "Gas Type", "Customer", "Action", "Admin Name"])
            sheets_write_with_retry(ws.append_rows, rows_to_append)
        except Exception as e:
            print("[admin_scanner] Error appending to Sheets:", e)
            
    threading.Thread(target=_append_admin_scans).start()

    return jsonify({'success': True, 'count': len(uids)})


if __name__ == '__main__':
    start_scheduler()
    app.run(debug=True)
else:
    # Also start scheduler when running under WSGI server (like Gunicorn), but not during migrations
    if not os.environ.get('RUN_MIGRATION') == '1':
        start_scheduler()